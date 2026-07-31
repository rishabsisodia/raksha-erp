import openpyxl
import csv
import io
import requests
import sys
import os

BASE = "https://raksha-erp-deploy.onrender.com"
XLSX = r"C:\Users\BusinessIntelligence\Desktop\Order & Sales (U).xlsx"

def sheet_to_csv(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ""
    buf = io.StringIO()
    w = csv.writer(buf)
    for r in rows:
        cleaned = []
        for v in r:
            if v is None:
                cleaned.append("")
            else:
                cleaned.append(str(v))
        w.writerow(cleaned)
    return buf.getvalue()

def upload_csv(csv_text, endpoint, label):
    files = {"file": ("data.csv", csv_text.encode("utf-8-sig"), "text/csv")}
    print(f"  Uploading to {endpoint} ...")
    r = requests.post(f"{BASE}{endpoint}", files=files, timeout=120)
    print(f"  [{r.status_code}] {r.text[:200]}")
    return r.status_code == 200

wb = openpyxl.load_workbook(XLSX, read_only=True)

# 1. Products
print("\n=== IMPORTING PRODUCTS ===")
ws = wb["Products"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Header: Sl No., Part No., Description, Category, Size (in mm), Load, MRP
# We need: Part No, Description (Name), Category, Size, Load Rating, MRP
buf = io.StringIO()
w = csv.writer(buf)
w.writerow(["Part No", "Description", "Category", "Size", "Load Rating", "MRP"])
for r in rows[1:]:
    sl, part_no, desc, cat, size, load, mrp = (list(r) + [None]*7)[:7]
    if not part_no:
        continue
    w.writerow([
        str(part_no).strip() if part_no else "",
        str(desc).strip() if desc else "",
        str(cat).strip() if cat else "",
        str(size).strip() if size else "",
        str(load).strip() if load else "",
        mrp if mrp else 0
    ])
upload_csv(buf.getvalue(), "/api/import/products", "Products")

# 2. Customers
print("\n=== IMPORTING CUSTOMERS ===")
ws = wb["Customers"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Header: Sl No., SD Code, GSTIN, Customer Name, City, State, Phone, Executive, Status
# The import expects: Customer ID, GSTIN, State, City, Contact Name, Contact Number
buf = io.StringIO()
w = csv.writer(buf)
w.writerow(["Customer ID", "GSTIN", "City", "State", "Contact Name", "Contact Number"])
for r in rows[1:]:
    sl, sd_code, gstin, name, city, state, phone, exec_name, status = (list(r) + [None]*9)[:9]
    cust_id = str(sd_code).strip() if sd_code else ""
    if not cust_id and not gstin:
        continue
    w.writerow([
        cust_id,
        str(gstin).strip() if gstin else "",
        str(city).strip() if city else "",
        str(state).strip() if state else "",
        str(name).strip() if name else "",
        str(int(phone)) if phone and isinstance(phone, (int, float)) else str(phone).strip() if phone else ""
    ])
upload_csv(buf.getvalue(), "/api/import/customers", "Customers")

# 3. Transporters
print("\n=== IMPORTING TRANSPORTERS ===")
ws = wb["Transporters"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Header: Sl No, Transporter ID, Name, Contact Person, Phone, State, GST Number, PAN Number
buf = io.StringIO()
w = csv.writer(buf)
w.writerow(["Transporter ID", "Name", "Contact Person", "Phone", "State", "GST Number", "PAN Number"])
for r in rows[1:]:
    sl, tid, name, cp, phone, state, gst, pan = (list(r) + [None]*8)[:8]
    if not tid or not name:
        continue
    w.writerow([
        str(tid).strip() if tid else "",
        str(name).strip() if name else "",
        str(cp).strip() if cp else "",
        str(int(phone)) if phone and isinstance(phone, (int, float)) else str(phone).strip() if phone else "",
        str(state).strip() if state else "",
        str(gst).strip() if gst else "",
        str(pan).strip() if pan else ""
    ])
upload_csv(buf.getvalue(), "/api/import/transporters", "Transporters")

# 4. Orders
print("\n=== IMPORTING ORDERS ===")
ws = wb["Orders"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Header: Sl No., PO Date, Billing Site, Shipping Site, No. Of Boxes,
# Value (excl. GST & Freight), Invoice No., Invoice Date, Invoice Amount (ex. GST),
# Weight (Kg), Freight (Rate / Kg), Transport Charges, Invoice Amount, Transporter
buf = io.StringIO()
w = csv.writer(buf)
w.writerow(["Sl No.", "PO Date", "Customer Name", "Billing Site", "Shipping Site",
            "No. Of Boxes", "Value (excl. GST & Freight)", "Invoice No.", "Invoice Date",
            "Invoice Amount (ex. GST)", "Weight (Kg)", "Freight (Rate / Kg)",
            "Transport Charges", "Invoice Amount", "Transporter"])
for r in rows[1:]:
    vals = list(r) + [None]*14
    sl, po_date, bill, ship, boxes, value, inv_no, inv_date, inv_amt, weight, frt_rate, trn_charges, inv_amt2, transporter = vals[:14]
    if sl is None and not bill:
        continue
    def fmt_date(d):
        if d is None:
            return ""
        if hasattr(d, 'strftime'):
            return d.strftime('%Y-%m-%d')
        return str(d)
    w.writerow([
        sl if sl else "",
        fmt_date(po_date),
        "",
        str(bill).strip() if bill else "",
        str(ship).strip() if ship else "",
        int(boxes) if boxes else 0,
        value if value else 0,
        str(inv_no).strip() if inv_no else "",
        fmt_date(inv_date),
        inv_amt if inv_amt else 0,
        weight if weight else 0,
        frt_rate if frt_rate else 0,
        trn_charges if trn_charges else 0,
        inv_amt2 if inv_amt2 else 0,
        str(transporter).strip() if transporter else ""
    ])
upload_csv(buf.getvalue(), "/api/import/orders", "Orders")

# 5. Sales (2)
print("\n=== IMPORTING SALES (2) ===")
ws = wb["Sales (2)"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Header: Invoice, Date, Party, Location, Transporter, Freight, Invoice Value, Weight, LR Tracking
# Import expects: Sl No., Date, Raksha Invoice NO, Party Name, Raksha Invoice Value,
# Payment Terms, Location, Pincode, State, Transporter Name, LR No, Freight, Weight
buf = io.StringIO()
w = csv.writer(buf)
w.writerow(["Sl No.", "Date", "Raksha Invoice NO", "Party Name", "Raksha Invoice Value",
            "Payment Terms", "Location", "Pincode", "State", "Transporter Name",
            "LR No", "Freight", "Weight"])
for i, r in enumerate(rows[1:], 1):
    vals = list(r) + [None]*9
    inv, date, party, loc, transporter, freight, inv_val, weight, lr = vals[:9]
    def fmt_date(d):
        if d is None:
            return ""
        if hasattr(d, 'strftime'):
            return d.strftime('%Y-%m-%d')
        return str(d)
    lr_str = str(lr).strip() if lr else ""
    if lr_str in ("-", "–"):
        lr_str = ""
    w.writerow([
        i,
        fmt_date(date),
        str(inv).strip() if inv else "",
        str(party).strip() if party else "",
        inv_val if inv_val else 0,
        "",
        str(loc).strip() if loc else "",
        "",
        "",
        str(transporter).strip() if transporter else "",
        lr_str,
        freight if freight else 0,
        weight if weight else 0
    ])
upload_csv(buf.getvalue(), "/api/import/sales", "Sales (2)")

# 6. Expenses
print("\n=== IMPORTING EXPENSES ===")
ws = wb["Expenses"]
rows = list(ws.iter_rows(values_only=True))
header = rows[0]
# Check what headers exist
print(f"  Expenses headers: {list(header)}")
if rows[1:]:
    print(f"  Sample row: {list(rows[1])}")
    # Create minimal CSV
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["Category", "Amount", "Description"])
    for r in rows[1:]:
        vals = list(r) + [None]*5
        cat = str(vals[0]).strip() if vals[0] else ""
        amt = vals[1] if vals[1] else 0
        desc = str(vals[2]).strip() if vals[2] else ""
        if cat and amt:
            w.writerow([cat, amt, desc])
    if buf.tell() > 20:
        upload_csv(buf.getvalue(), "/api/import/expenses", "Expenses")
    else:
        print("  No expenses to import")
else:
    print("  No expense data rows")

wb.close()
print("\n=== IMPORT COMPLETE ===")
