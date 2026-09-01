import io
import csv
import re
from datetime import datetime
from fastapi import APIRouter, Depends
from fastapi.responses import Response, StreamingResponse
from fpdf import FPDF

from models import Order, Sale, ProformaOrder, ProformaOrderItem, User, Customer
from auth import get_current_user, require_permission
from database import SessionLocal

router = APIRouter(tags=["exports"])


@router.get("/api/export/orders")
def export_orders(format: str = "csv", user: User = Depends(require_permission("orders", "export"))):
    db = SessionLocal()
    try:
        rows = db.query(Order).order_by(Order.sl_no).all()
        headers = ["Sl No.", "PO No.", "PO Date", "Customer Name", "Billing Site", "Shipping Site", "No. Of Boxes",
                    "Value (excl. GST & Freight)", "Invoice No.", "Invoice Date",
                    "Invoice Amount (ex. GST)", "Weight (Kg)", "Freight (Rate / Kg)",
                    "Transport Charges", "Invoice Amount", "E-way Bill No", "LR No.",
                    "Entry Date", "Credit Note Amount", "Credit Note No.", "Transporter"]
        data = []
        for o in rows:
            data.append([o.sl_no, o.po_no or "", o.po_date or "", o.customer_name or "", o.billing_site or "", o.shipping_site or "",
                         o.no_of_boxes or 0, o.value_excl_gst_freight or 0, o.invoice_no or "",
                         o.invoice_date or "", o.invoice_amount_excl_gst or 0, o.weight_kgs or 0,
                         o.freight_rate_per_kg or 0, o.transport_charges or 0, o.invoice_amount or 0,
                         o.eway_bill_no or "", o.lr_no or "", o.entry_date or "",
                         o.credit_note_amount or 0, o.credit_note_no or "", o.transporter or ""])

        if format == "xlsx":
            return export_xlsx("Orders", headers, data)
        elif format == "pdf":
            return export_pdf("Orders", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


@router.get("/api/export/sales")
def export_sales(format: str = "csv", user: User = Depends(require_permission("sales", "export"))):
    db = SessionLocal()
    try:
        rows = db.query(Sale).order_by(Sale.id.desc()).all()
        headers = ["Invoice No.", "Date", "Party Name", "Location", "State", "Transporter",
                    "Freight", "Weight", "Weight PG Fiber", "Invoice Value", "GP", "GP%",
                    "Payment Terms", "Sales Person", "PG Fiber Invoice No", "PG Fiber Invoice Value"]
        data = []
        for s in rows:
            dt = ""
            if s.sale_date:
                try:
                    dt = s.sale_date.strftime("%Y-%m-%d")
                except Exception:
                    dt = str(s.sale_date)[:10]
            data.append([s.invoice_no or "", dt, s.party_name or "", s.location or "",
                         s.state or "", s.transporter_name or "", s.freight_amount or 0,
                         s.weight_kgs or 0, s.weight_pg_fiber or 0, s.invoice_value or 0,
                         s.gp or 0, s.gp_percent or 0, s.payment_terms or "",
                         s.sales_person or "", s.pg_fiber_invoice_no or "",
                         s.pg_fiber_invoice_value or 0])

        if format == "xlsx":
            return export_xlsx("Sales", headers, data)
        elif format == "pdf":
            return export_pdf("Sales", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


@router.get("/api/export/proforma-orders")
def export_proforma_orders(format: str = "csv", order_type: str = None, user: User = Depends(require_permission("proforma_orders", "export"))):
    db = SessionLocal()
    try:
        query = db.query(ProformaOrder)
        if order_type:
            query = query.filter(ProformaOrder.order_type == order_type)
        rows = query.order_by(ProformaOrder.created_at.desc()).all()
        headers = ["PI No", "Date", "Customer", "Type", "Billing Site", "Shipping Site",
                    "Boxes", "Total Qty", "Value (excl GST)", "GST", "Freight",
                    "Total Amount", "Payment Status", "Delivery Days"]
        data = []
        for o in rows:
            cust = db.query(Customer).filter(Customer.id == o.customer_id).first()
            data.append([o.pi_no or "", o.pi_date.strftime("%Y-%m-%d") if o.pi_date else "",
                         cust.contact_name if cust else "", o.order_type or "",
                         o.billing_site or "", o.shipping_site or "",
                         o.no_of_boxes or 0, o.total_qty or 0, o.value_excl_gst or 0,
                         o.gst_amount or 0, o.freight_amount or 0, o.total_amount or 0,
                         o.payment_status or "", o.delivery_days or 0])

        if format == "xlsx":
            return export_xlsx("PI-PO Orders", headers, data)
        elif format == "pdf":
            return export_pdf("PI-PO Orders", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


def export_csv(headers, data):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    csv_bytes = output.getvalue().encode('utf-8-sig')
    return Response(content=csv_bytes, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=export.csv"})


def export_xlsx(sheet_name, headers, data):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in data:
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={sheet_name}.xlsx"})


def export_pdf(title, headers, data):
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Raksha ERP - {title}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", ln=True, align="C")
    pdf.ln(4)
    num_cols = len(headers)
    col_width = max(277 / num_cols, 20)
    pdf.set_font("Helvetica", "B", 7)
    for h in headers:
        short_h = str(h)[:20]
        pdf.cell(col_width, 7, short_h, border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "", 6)
    for row in data:
        for val in row:
            s = str(val)[:22]
            pdf.cell(col_width, 5, s, border=1)
        pdf.ln()
    pdf_bytes = pdf.output()
    return Response(content=bytes(pdf_bytes), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={title}.pdf"})
