from fastapi import APIRouter, HTTPException, UploadFile, File, Depends
from fastapi.security import HTTPBearer
from ..models import (
    Order, Sale, Product, Pricing, Customer, Transporter, Expense, User
)
from ..schemas import (
    ProductIn, PricingIn
)
from ..auth import get_current_user, require_permission
from ..database import SessionLocal
from sqlalchemy import text
from ..config import ALLOWED_EXTENSIONS, DEFAULT_GST_RATE
from datetime import datetime, timezone
from typing import Optional
import os
import csv
import io
import logging

logger = logging.getLogger(__name__)

router = APIRouter(tags=["imports"])


# ---- CSV IMPORT HELPERS ----
def parse_csv_amount(val):
    if not val or str(val).strip() in ('-', '–', '', 'None', 'none'):
        return 0
    return float(str(val).replace('₹', '').replace(',', '').replace(' ', '').strip() or 0)


def parse_csv_date(val):
    if not val or val.strip() in ('-', '–', ''):
        return None
    from dateutil import parser as dateparser
    try:
        return dateparser.parse(val.strip(), dayfirst=False).strftime('%Y-%m-%d')
    except Exception:
        return val.strip() if val else None


def map_csv_col(row, keys, default=""):
    for k in keys:
        v = row.get(k, "")
        if v and v.strip() and v.strip() not in ('-', '–'):
            return v.strip()
    return default


# ---- XLSX IMPORT HELPERS ----
def read_xlsx_sheet(file_content, sheet_name=None):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = None
    if sheet_name:
        exact = [s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()]
        if exact:
            ws = wb[exact[0]]
        else:
            partial = [s for s in wb.sheetnames if sheet_name.strip().lower() in s.strip().lower()]
            if partial:
                ws = wb[partial[0]]
    if ws is None:
        ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    return rows


def rows_to_csv_string(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


# ---- CSV IMPORT ENDPOINTS ----
@router.post("/api/import/orders")
async def import_orders_csv(file: UploadFile = File(...), user: User = Depends(require_permission("orders", "import"))):
    content = await file.read()
    text_content = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text_content))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            sl_raw = row.get('Sl No.', '').strip()
            if not sl_raw:
                skipped += 1
                continue
            try:
                sl_no = int(sl_raw)
            except ValueError:
                skipped += 1
                continue
            existing = db.query(Order).filter(Order.sl_no == sl_no).first()
            data = {
                "sl_no": sl_no,
                "po_no": "",
                "po_date": parse_csv_date(row.get('PO Date', '')),
                "customer_name": row.get('Customer Name', '').strip(),
                "billing_site": row.get('Billing Site', '').strip(),
                "shipping_site": row.get('Shipping Site', '').strip(),
                "no_of_boxes": int(parse_csv_amount(row.get('No. Of Boxes', '0'))),
                "value_excl_gst_freight": parse_csv_amount(row.get('Value (excl. GST & Freight)', '0')),
                "invoice_no": row.get('Invoice No.', '').strip().strip('-– ') if row.get('Invoice No.', '').strip() not in ('-', '–', '') else '',
                "invoice_date": parse_csv_date(row.get('Invoice Date', '')),
                "invoice_amount_excl_gst": parse_csv_amount(row.get('Invoice Amount (ex. GST)', '0')),
                "weight_kgs": parse_csv_amount(row.get('Weight (Kg)', '0')),
                "freight_rate_per_kg": parse_csv_amount(row.get('Freight (Rate / Kg)', '0')),
                "transport_charges": parse_csv_amount(row.get('Transport Charges', '0')),
                "invoice_amount": parse_csv_amount(row.get('Invoice Amount', '0')),
                "eway_bill_no": row.get('E-way Bill No', '').strip() if row.get('E-way Bill No', '').strip() not in ('-', '–', '') else '',
                "lr_no": row.get('LR Copy', '').strip() if row.get('LR Copy', '').strip() not in ('-', '–', '') else '',
                "entry_date": parse_csv_date(row.get('ERP Entry Date', '')),
                "credit_note_amount": parse_csv_amount(row.get('Credit Note Amount (If any)', '0')),
                "credit_note_no": row.get('Credit Note No.', '').strip() if row.get('Credit Note No.', '').strip() not in ('-', '–', '') else '',
                "transporter": row.get('Transporter', '').strip(),
                "transporter_no": "",
            }
            if existing:
                for k, v in data.items():
                    if k != "sl_no":
                        setattr(existing, k, v)
            else:
                o = Order(**data)
                db.add(o)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} orders, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@router.post("/api/import/sales")
async def import_sales_csv(file: UploadFile = File(...), user: User = Depends(require_permission("sales", "import"))):
    content = await file.read()
    text_content = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text_content))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            sl_raw = row.get('Sl No.', '').strip()
            if not sl_raw:
                skipped += 1
                continue
            try:
                sl_no = int(sl_raw)
            except ValueError:
                skipped += 1
                continue
            invoice_no = row.get('Raksha Invoice NO', '').strip()
            if invoice_no in ('-', '–', ''):
                invoice_no = f"INDORE-{sl_no:04d}"
            sale_date = parse_csv_date(row.get('Date ', '') or row.get('Date', ''))
            sale_date_dt = None
            if sale_date:
                try:
                    from datetime import datetime as dt
                    sale_date_dt = dt.strptime(sale_date, '%Y-%m-%d')
                except Exception:
                    logger.warning("Invalid sale_date in CSV import: %s", sale_date)
            freight = parse_csv_amount(row.get('Freight', '0'))
            gp = parse_csv_amount(row.get('GP', '0'))
            gp_pct_raw = row.get('GP%', '0').replace('%', '').strip()
            try:
                gp_pct = float(gp_pct_raw) if gp_pct_raw and gp_pct_raw not in ('-', '–', '', 'None') else 0
            except ValueError:
                gp_pct = 0
            invoice_value = parse_csv_amount(row.get('Raksha Invoice Value', '') or row.get('Invoice Value', '') or row.get('invoice_value', '0'))
            s = Sale(
                invoice_no=invoice_no,
                sale_date=sale_date_dt,
                payment_terms=row.get('Payment Terms', '').strip(),
                party_name=row.get('Party Name ', '') or row.get('Party Name', '').strip(),
                location=row.get('Location', '').strip(),
                pincode=row.get('Pincode', '').strip(),
                state=row.get('State', '').strip(),
                transporter_name=row.get('Transporter Name', '').strip(),
                lr_no=row.get('LR No', '').strip(),
                freight_amount=freight,
                weight_kgs=parse_csv_amount(row.get('Weight', '0')),
                weight_pg_fiber=parse_csv_amount(row.get('Weight on PG Fiber Bill', '0')),
                sales_person=row.get('Sales Ex Person / Person In-Charge', '').strip(),
                pg_fiber_invoice_no=(row.get('P.G.Fiber Invoice No', '') or row.get('P.G.Fiber Invoice No ', '') or '').strip(),
                pg_fiber_invoice_value=parse_csv_amount(row.get('P.G.Fiber Invoice Value', '') or row.get('P.G.Fiber Invoice Value ', '0')),
                gp=gp,
                gp_percent=gp_pct,
                invoice_value=invoice_value,
                total_amount=invoice_value if invoice_value else 0,
                source_csv="From Indore",
            )
            db.add(s)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} sales, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@router.post("/api/import/products")
async def import_products_csv(file: UploadFile = File(...), user: User = Depends(require_permission("products", "import"))):
    content = await file.read()
    text_content = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text_content))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            name = map_csv_col(row, ['Name', 'Description', 'Product Name', 'name', 'description'])
            if not name:
                skipped += 1
                continue
            part_no = map_csv_col(row, ['Part No', 'Part Number', 'SKU', 'part_no'])
            category = map_csv_col(row, ['Category', 'category'], 'Manhole Cover')
            size = map_csv_col(row, ['Size', 'size'])
            load_rating = map_csv_col(row, ['Load Rating', 'Load', 'load_rating'], '5 Ton')
            material = map_csv_col(row, ['Material', 'material'], 'FRP')
            color = map_csv_col(row, ['Color', 'color'], 'Grey')
            unit = map_csv_col(row, ['Unit', 'unit'], 'Nos')
            hsn_code = map_csv_col(row, ['HSN Code', 'HSN', 'hsn_code'])
            mrp = parse_csv_amount(map_csv_col(row, ['MRP', 'Price', 'mrp'], '0'))

            if part_no:
                existing = db.query(Product).filter(Product.part_no == part_no).first()
            else:
                existing = db.query(Product).filter(Product.name == name).first()
            if existing:
                existing.name = name
                existing.category = category
                existing.size = size
                existing.load_rating = load_rating
                existing.material = material
                existing.color = color
                existing.unit = unit
                existing.hsn_code = hsn_code
                if part_no:
                    existing.part_no = part_no
                if mrp and existing.pricing:
                    existing.pricing.mrp = mrp
                    existing.pricing.raw_material_cost = mrp
                    existing.pricing.total_cost = mrp
                imported += 1
                continue

            p = Product(part_no=part_no, name=name, category=category,
                        size=size, load_rating=load_rating, material=material,
                        color=color, unit=unit, hsn_code=hsn_code)
            db.add(p)
            db.flush()
            db.add(Pricing(product_id=p.id, raw_material_cost=mrp, total_cost=mrp, mrp=mrp))
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} products, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@router.post("/api/import/customers")
async def import_customers_csv(file: UploadFile = File(...), user: User = Depends(require_permission("customers", "import"))):
    content = await file.read()
    text_content = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text_content))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            customer_id = map_csv_col(row, ['Customer ID', 'ID', 'Cust ID', 'customer_id'])
            if not customer_id:
                skipped += 1
                continue
            existing = db.query(Customer).filter(Customer.customer_id == customer_id).first()
            data = {
                "customer_id": customer_id,
                "name": map_csv_col(row, ['Contact Name', 'Name', 'Customer Name', 'contact_name']),
                "gstin": map_csv_col(row, ['GSTIN', 'GST Number', 'GST', 'gstin']),
                "billing_address": map_csv_col(row, ['Billing Address', 'Address', 'billing_address']),
                "shipping_address": map_csv_col(row, ['Shipping Address', 'shipping_address']),
                "state": map_csv_col(row, ['State', 'state']),
                "district": map_csv_col(row, ['District', 'district']),
                "city": map_csv_col(row, ['City', 'city']),
                "pincode": map_csv_col(row, ['Pincode', 'Pin Code', 'ZIP', 'pincode']),
                "contact_name": map_csv_col(row, ['Contact Name', 'Name', 'contact_name']),
                "contact_number": map_csv_col(row, ['Contact Number', 'Phone', 'Mobile', 'contact_number']),
                "contact_email": map_csv_col(row, ['Contact Email', 'Email', 'contact_email']),
                "exec_code": map_csv_col(row, ['Executive Code', 'Exec Code', 'exec_code']),
                "exec_name": map_csv_col(row, ['Executive Name', 'Exec Name', 'exec_name']),
                "exec_number": map_csv_col(row, ['Executive Number', 'Exec Number', 'exec_number']),
                "exec_email": map_csv_col(row, ['Executive Email', 'Exec Email', 'exec_email']),
            }
            if existing:
                for k, v in data.items():
                    if k != "customer_id":
                        setattr(existing, k, v)
            else:
                c = Customer(**data)
                db.add(c)
                try:
                    db.flush()
                except Exception:
                    db.rollback()
                    ex = db.query(Customer).filter(Customer.customer_id == customer_id).first()
                    if ex:
                        for k, v in data.items():
                            if k != "customer_id":
                                setattr(ex, k, v)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} customers, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@router.post("/api/import/transporters")
async def import_transporters_csv(file: UploadFile = File(...), user: User = Depends(require_permission("transporters", "import"))):
    content = await file.read()
    text_content = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text_content))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            transporter_id = map_csv_col(row, ['Transporter ID', 'ID', 'transporter_id'])
            name = map_csv_col(row, ['Name', 'Transporter Name', 'name'])
            if not transporter_id or not name:
                skipped += 1
                continue
            existing = db.query(Transporter).filter(Transporter.transporter_id == transporter_id).first()
            data = {
                "transporter_id": transporter_id,
                "name": name,
                "phone": map_csv_col(row, ['Phone', 'Mobile', 'phone']),
                "email": map_csv_col(row, ['Email', 'email']),
                "address": map_csv_col(row, ['Address', 'address']),
                "state": map_csv_col(row, ['State', 'state']),
                "district": map_csv_col(row, ['District', 'district']),
                "city": map_csv_col(row, ['City', 'city']),
                "pincode": map_csv_col(row, ['Pincode', 'Pin Code', 'pincode']),
                "gst_number": map_csv_col(row, ['GST Number', 'GSTIN', 'GST', 'gst_number']),
                "pan_number": map_csv_col(row, ['PAN Number', 'PAN', 'pan_number']),
                "contact_person": map_csv_col(row, ['Contact Person', 'contact_person']),
                "contact_number": map_csv_col(row, ['Contact Number', 'contact_number']),
            }
            if existing:
                for k, v in data.items():
                    if k != "transporter_id":
                        setattr(existing, k, v)
            else:
                t = Transporter(**data)
                db.add(t)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} transporters, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@router.post("/api/import/expenses")
async def import_expenses_csv(file: UploadFile = File(...), user: User = Depends(require_permission("expenses", "import"))):
    content = await file.read()
    text_content = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text_content))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            category = map_csv_col(row, ['Category', 'category'])
            amount_raw = map_csv_col(row, ['Amount', 'amount'], '0')
            amount = parse_csv_amount(amount_raw)
            if not category or amount == 0:
                skipped += 1
                continue
            date_str = map_csv_col(row, ['Date', 'Expense Date', 'expense_date'])
            dt = None
            if date_str:
                try:
                    from dateutil import parser as dateparser
                    dt = dateparser.parse(date_str, dayfirst=False)
                except Exception:
                    logger.warning("Invalid expense date in CSV import: %s", date_str)
            if not dt:
                dt = datetime.now(timezone.utc)
            e = Expense(
                category=category,
                description=map_csv_col(row, ['Description', 'description']),
                amount=amount,
                vendor=map_csv_col(row, ['Vendor', 'vendor']),
                expense_date=dt
            )
            db.add(e)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} expenses, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@router.post("/api/upload")
async def upload_file(file: UploadFile = File(...), user: User = Depends(require_permission("products", "create"))):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, "Only .jpg, .png, .pdf files allowed")
    try:
        import cloudinary.uploader
        content = await file.read()
        rtype = "raw" if ext == ".pdf" else "image"
        result = cloudinary.uploader.upload(
            content,
            folder="raksha_erp",
            resource_type=rtype
        )
        url = result["secure_url"]
        if rtype == "raw" and "/image/" in url:
            url = url.replace("/image/upload/", "/raw/upload/")
        return {"filename": result["public_id"], "url": url, "original": file.filename}
    except Exception as e:
        raise HTTPException(500, f"Upload failed: {str(e)}")


@router.post("/api/import-standard-packaging")
async def import_standard_packaging(file: UploadFile = File(...), user: User = Depends(require_permission("products", "import"))):
    db = SessionLocal()
    try:
        content = await file.read()
        filename = (file.filename or "").lower()

        if filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        else:
            text_content = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text_content))
            rows = list(reader)

        if not rows:
            raise HTTPException(400, "File is empty")

        header_row_idx = 0
        for i, row in enumerate(rows):
            row_lower = [str(c).strip().lower() for c in row]
            if any("part no" in c for c in row_lower):
                header_row_idx = i
                break

        header = [h.strip().lower() for h in rows[header_row_idx]]
        part_no_idx = None
        box_idx = None
        mrp_idx = None
        desc_idx = None
        size_idx = None
        grp_idx = None
        for i, h in enumerate(header):
            if "part no" in h:
                part_no_idx = i
            if h in ("box", "boxes") or "std pack" in h or "std pkg" in h:
                box_idx = i
            if h in ("gen", "mrp") or h == "gen":
                mrp_idx = i
            if "product spec" in h or "description" in h:
                desc_idx = i
            if "size" in h:
                size_idx = i
            if "item grp" in h or "item group" in h or "grp" in h:
                grp_idx = i
        if part_no_idx is None or box_idx is None:
            raise HTTPException(400, f"File must have 'Part No' and 'Box'/'Std Packing' columns. Found: {header}")

        desc_idx = desc_idx or 1
        size_idx = size_idx or 2
        grp_idx = grp_idx or 4

        created = 0
        updated = 0
        not_found = []
        for row in rows[header_row_idx+1:]:
            if len(row) <= max(part_no_idx, box_idx):
                continue
            part_no = row[part_no_idx].strip().replace(" ", "")
            box_val = row[box_idx].strip()
            if not part_no or not box_val:
                continue
            try:
                pieces = int(float(box_val))
            except ValueError:
                logger.debug(f"Skipping row: invalid box value '{box_val}' for part_no '{part_no}'")
                continue

            desc = str(row[desc_idx] or "").strip() if len(row) > desc_idx else ""
            size_raw = str(row[size_idx] or "").strip() if len(row) > size_idx else ""
            item_grp = str(row[grp_idx] or "").strip().upper() if len(row) > grp_idx else "FRP"

            size_mm = size_raw.lower().replace(" ", "").replace("x", "x")
            tonnage = "10 Ton" if "10 ton" in desc.lower() else "5 Ton"
            color = "White" if "white" in desc.lower() or part_no.upper().endswith("-WH") else "Grey"
            has_lock = "lock" in desc.lower() or part_no.upper().endswith("L") and "GRY" not in part_no.upper()
            has_hinges = "hinge" in desc.lower()
            category = "Gully Cover" if "gully" in desc.lower() or item_grp == "GULLY" else "Manhole Cover"

            product = db.query(Product).filter(Product.part_no == part_no).first()
            if product:
                product.pieces_per_box = pieces
                product.std_packaging = pieces
                if mrp_idx is not None and len(row) > mrp_idx:
                    mrp_val = row[mrp_idx].strip().replace(",", "").replace("₹", "").replace("?", "")
                    try:
                        mrp = float(mrp_val)
                        if mrp > 0:
                            pr = db.query(Pricing).filter(Pricing.product_id == product.id).first()
                            if pr:
                                pr.mrp = mrp
                            else:
                                db.add(Pricing(product_id=product.id, mrp=mrp, gst_rate=DEFAULT_GST_RATE))
                    except ValueError:
                        logger.warning("Invalid MRP value '%s' for product %s", mrp_val, part_no)
                updated += 1
            else:
                if not desc:
                    continue
                new_p = Product(
                    part_no=part_no, name=desc, category=category,
                    size=size_mm, load_rating=tonnage, material="FRP",
                    color=color, hsn_code="39259090",
                    pieces_per_box=pieces, std_packaging=pieces
                )
                db.add(new_p)
                db.flush()
                mrp_val = 0
                if mrp_idx is not None and len(row) > mrp_idx:
                    try:
                        mrp_val = float(row[mrp_idx].strip().replace(",", "").replace("₹", "").replace("?", ""))
                    except ValueError:
                        logger.warning("Invalid MRP value for new product: %s", row[mrp_idx])
                db.add(Pricing(product_id=new_p.id, mrp=mrp_val, gst_rate=DEFAULT_GST_RATE))
                created += 1

        db.commit()
        return {"updated": updated, "created": created, "not_found": not_found, "total_rows": len(rows) - header_row_idx - 1}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


# ---- XLSX IMPORT ENDPOINTS ----
@router.post("/api/import/orders-xlsx")
async def import_orders_xlsx(file: UploadFile = File(...), user: User = Depends(require_permission("orders", "import"))):
    content = await file.read()
    try:
        rows = read_xlsx_sheet(content, sheet_name='Orders')
        if not rows:
            raise HTTPException(400, "File is empty")
        csv_text = rows_to_csv_string(rows)
        text_content = csv_text
        reader = csv.DictReader(io.StringIO(text_content))
        db = SessionLocal()
        imported = 0
        skipped = 0
        try:
            for row in reader:
                sl_raw = row.get('Sl No.', '').strip()
                if not sl_raw:
                    skipped += 1
                    continue
                try:
                    sl_no = int(float(sl_raw))
                except ValueError:
                    skipped += 1
                    continue
                existing = db.query(Order).filter(Order.sl_no == sl_no).first()
                data = {
                    "sl_no": sl_no,
                    "po_no": "",
                    "po_date": parse_csv_date(row.get('PO Date', '')),
                    "customer_name": (row.get('Customer Name', '') or row.get('Billing Site', '') or '').strip(),
                    "billing_site": row.get('Billing Site', '').strip(),
                    "shipping_site": row.get('Shipping Site', '').strip(),
                    "no_of_boxes": int(parse_csv_amount(row.get('No. Of Boxes', '0'))),
                    "value_excl_gst_freight": parse_csv_amount(row.get('Value (excl. GST & Freight)', '0')),
                    "invoice_no": row.get('Invoice No.', '').strip() if row.get('Invoice No.', '').strip() not in ('-', '–', '') else '',
                    "invoice_date": parse_csv_date(row.get('Invoice Date', '')),
                    "invoice_amount_excl_gst": parse_csv_amount(row.get('Invoice Amount (ex. GST)', '0')),
                    "weight_kgs": parse_csv_amount(row.get('Weight (Kg)', '0')),
                    "freight_rate_per_kg": parse_csv_amount(row.get('Freight (Rate / Kg)', '0')),
                    "transport_charges": parse_csv_amount(row.get('Transport Charges', '0')),
                    "invoice_amount": parse_csv_amount(row.get('Invoice Amount', '0')),
                    "eway_bill_no": row.get('E-way Bill No', '').strip() if row.get('E-way Bill No', '').strip() not in ('-', '–', '') else '',
                    "lr_no": row.get('LR Copy', '').strip() if row.get('LR Copy', '').strip() not in ('-', '–', '') else '',
                    "entry_date": parse_csv_date(row.get('ERP Entry Date', '')),
                    "credit_note_amount": parse_csv_amount(row.get('Credit Note Amount (If any)', '0')),
                    "credit_note_no": row.get('Credit Note No.', '').strip() if row.get('Credit Note No.', '').strip() not in ('-', '–', '') else '',
                    "transporter": row.get('Transporter', '').strip(),
                    "transporter_no": "",
                }
                if existing:
                    for k, v in data.items():
                        if k != "sl_no":
                            setattr(existing, k, v)
                else:
                    o = Order(**data)
                    db.add(o)
                imported += 1
            db.commit()
            return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} orders from XLSX"}
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to read XLSX: {str(e)}")


@router.post("/api/import/sales-xlsx")
async def import_sales_xlsx(file: UploadFile = File(...), user: User = Depends(require_permission("sales", "import"))):
    content = await file.read()
    try:
        rows = read_xlsx_sheet(content, sheet_name='Sales')
        if not rows:
            raise HTTPException(400, "File is empty")
        csv_text = rows_to_csv_string(rows)
        reader = csv.DictReader(io.StringIO(csv_text))
        db = SessionLocal()
        imported = 0
        skipped = 0
        try:
            for row in reader:
                sl_raw = row.get('Sl No.', '').strip()
                if not sl_raw:
                    skipped += 1
                    continue
                try:
                    int(float(sl_raw))
                except ValueError:
                    skipped += 1
                    continue
                invoice_no = row.get('Raksha Invoice NO', '').strip()
                if invoice_no in ('-', '–', ''):
                    invoice_no = None
                sale_date = parse_csv_date(row.get('Date ', '') or row.get('Date', ''))
                sale_date_dt = None
                if sale_date:
                    try:
                        sale_date_dt = datetime.strptime(sale_date, '%Y-%m-%d')
                    except Exception:
                        logger.warning("Invalid sale_date in CSV update: %s", sale_date)
                freight = parse_csv_amount(row.get('Freight', '0'))
                gp = parse_csv_amount(row.get('GP', '0'))
                gp_pct_raw = row.get('GP%', '0').replace('%', '').strip()
                try:
                    gp_pct = float(gp_pct_raw) if gp_pct_raw and gp_pct_raw not in ('-', '–', '', 'None') else 0
                except ValueError:
                    gp_pct = 0
                invoice_value = parse_csv_amount(row.get('Raksha Invoice Value', '') or row.get('Invoice Value', '') or row.get('invoice_value', '0'))

                existing = None
                if invoice_no:
                    existing = db.query(Sale).filter(Sale.invoice_no == invoice_no).first()
                if not existing:
                    existing = db.query(Sale).filter(
                        Sale.party_name == (row.get('Party Name ', '') or row.get('Party Name', '') or '').strip(),
                        Sale.sale_date == sale_date_dt
                    ).first() if sale_date_dt else None

                data = dict(
                    invoice_no=invoice_no,
                    sale_date=sale_date_dt,
                    payment_terms=row.get('Payment Terms', '').strip(),
                    party_name=(row.get('Party Name ', '') or row.get('Party Name', '') or '').strip(),
                    location=row.get('Location', '').strip(),
                    pincode=row.get('Pincode', '').strip(),
                    state=row.get('State', '').strip(),
                    transporter_name=row.get('Transporter Name', '').strip(),
                    lr_no=row.get('LR No', '').strip(),
                    freight_amount=freight,
                    weight_kgs=parse_csv_amount(row.get('Weight', '0')),
                    weight_pg_fiber=parse_csv_amount(row.get('Weight on PG Fiber Bill', '0')),
                    sales_person=row.get('Sales Ex Person / Person In-Charge', '').strip(),
                    pg_fiber_invoice_no=(row.get('P.G.Fiber Invoice No', '') or row.get('P.G.Fiber Invoice No ', '') or '').strip(),
                    pg_fiber_invoice_value=parse_csv_amount(row.get('P.G.Fiber Invoice Value', '') or row.get('P.G.Fiber Invoice Value ', '0')),
                    gp=gp,
                    gp_percent=gp_pct,
                    invoice_value=invoice_value,
                    total_amount=invoice_value if invoice_value else 0,
                    source_csv="From Indore",
                )

                if existing:
                    for k, v in data.items():
                        setattr(existing, k, v)
                else:
                    s = Sale(**data)
                    db.add(s)
                imported += 1
            db.commit()
            return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} sales from XLSX"}
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to read XLSX: {str(e)}")
