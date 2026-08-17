from fastapi import FastAPI, HTTPException, UploadFile, File, Query, Depends, status, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, HTMLResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware
from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Date, ForeignKey, Text, text, func
from sqlalchemy.orm import declarative_base, sessionmaker, relationship
from pydantic import BaseModel, Field, field_validator
from typing import Optional, List, Literal
from datetime import datetime, date, timedelta, timezone
import os
import logging
import cloudinary
import cloudinary.uploader
import urllib.request
import csv
import io
import re
import time
import tempfile
import uuid
import bcrypt
import jwt
import requests
from fpdf import FPDF
from html import escape as escape_html

logger = logging.getLogger("raksha-erp")
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s: %(message)s")

app = FastAPI(title="Raksha ERP")

# Rate limiter (fix #3)
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, lambda request, exc: JSONResponse(
    status_code=429, content={"detail": "Too many requests. Please try again later."}
))
app.add_middleware(SlowAPIMiddleware)

# CORS from env var (fix #5) with preflight cache (fix #19)
_cors_origins = os.environ.get("CORS_ORIGINS", "https://raksha-erp-deploy.onrender.com").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in _cors_origins if o.strip()],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Authorization", "Content-Type"],
    max_age=600,
)

# Security headers middleware (fix #12) + HTTPS enforcement (fix #13)
@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    response = await call_next(request)
    if request.url.path.startswith("/api/"):
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["X-XSS-Protection"] = "0"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=()"
        response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.tailwindcss.com https://cdn.jsdelivr.net; style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://fonts.googleapis.com; font-src 'self' https://cdnjs.cloudflare.com https://fonts.gstatic.com; img-src 'self' data: blob: https:; connect-src 'self';"
        if not request.url.scheme == "https" and request.headers.get("x-forwarded-proto", "https") != "https":
            if os.environ.get("ENVIRONMENT") == "production":
                return JSONResponse(status_code=301, content={"detail": "HTTPS required"}, headers={"Strict-Transport-Security": "max-age=31536000; includeSubDomains"})
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response

JWT_SECRET = os.environ.get("JWT_SECRET")
if not JWT_SECRET:
    logger.critical("JWT_SECRET environment variable is not set. Refusing to start without a secure secret.")
    raise RuntimeError("JWT_SECRET environment variable is required")
JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30
REFRESH_TOKEN_EXPIRE_DAYS = 7

CLOUDINARY_URL = os.environ.get("CLOUDINARY_URL", "")
if CLOUDINARY_URL and "@" in CLOUDINARY_URL:
    parts = CLOUDINARY_URL.replace("cloudinary://", "").split("@")
    creds = parts[0].split(":")
    cloudinary.config(
        api_key=creds[0],
        api_secret=creds[1],
        cloud_name=parts[1],
        secure=True
    )

# WhatsApp Cloud API Config
WHATSAPP_TOKEN = os.environ.get("WHATSAPP_TOKEN", "")
WHATSAPP_PHONE_ID = os.environ.get("WHATSAPP_PHONE_ID", "")
WHATSAPP_BUSINESS_ACCOUNT_ID = os.environ.get("WHATSAPP_BUSINESS_ACCOUNT_ID", "")
WHATSAPP_API_URL = "https://graph.facebook.com/v18.0"

DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./raksha_erp.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
if DATABASE_URL.startswith("postgresql://"):
    engine = create_engine(DATABASE_URL, pool_size=20, max_overflow=10, pool_timeout=30, pool_pre_ping=True)
elif os.environ.get("ENVIRONMENT") == "production":
    logger.critical("DATABASE_URL must be set to PostgreSQL in production. SQLite is not supported.")
    raise RuntimeError("DATABASE_URL must be set to PostgreSQL in production")
else:
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

def get_setting(key, default=""):
    try:
        db = SessionLocal()
        row = db.query(Settings).filter(Settings.key == key).first()
        db.close()
        return row.value if row else default
    except Exception:
        return default

def get_gst_rate():
    val = get_setting("default_gst_rate", "18")
    try:
        return float(val)
    except Exception:
        return 18.0

# Temp PDF storage for WhatsApp (no Cloudinary dependency)
_TEMP_PDFS = {}
_TEMP_PDFS_MAX_AGE = 3600  # 1 hour
_TEMP_PDFS_MAX_SIZE = 100  # max entries

@app.get("/api/whatsapp/temp-pdf/{pdf_id}")
def serve_temp_pdf(pdf_id: str, user: User = Depends(get_current_user)):
    # Cleanup old entries on access
    now = time.time()
    expired = [k for k, v in _TEMP_PDFS.items() if now - v.get("created_at", 0) > _TEMP_PDFS_MAX_AGE]
    for k in expired:
        del _TEMP_PDFS[k]
    
    data = _TEMP_PDFS.get(pdf_id)
    if not data:
        raise HTTPException(404, "PDF expired or not found")
    del _TEMP_PDFS[pdf_id]
    safe_filename = re.sub(r'[^\w\-.]', '_', data["filename"])
    return Response(content=data["bytes"], media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'})

ROLE_PERMISSIONS = {
    "admin": {
        "dashboard": ["view"],
        "products": ["view", "create", "edit", "delete", "import", "export"],
        "orders": ["view", "create", "edit", "delete", "import", "export"],
        "proforma_orders": ["view", "create", "edit", "delete", "export"],
        "customers": ["view", "create", "edit", "delete", "import"],
        "transporters": ["view", "create", "edit", "delete", "import"],
        "sales": ["view", "create", "edit", "delete", "import", "export", "bulk_edit"],
        "expenses": ["view", "create", "edit", "delete", "import"],
        "reports": ["view", "export"],
        "settings": ["view", "edit"],
        "users": ["view", "create", "edit", "delete"],
    },
    "manager": {
        "dashboard": ["view"],
        "products": ["view", "create", "edit"],
        "orders": ["view", "create", "edit"],
        "proforma_orders": ["view", "create", "edit"],
        "customers": ["view", "create", "edit"],
        "transporters": ["view", "create", "edit"],
        "sales": ["view", "create", "edit"],
        "expenses": ["view", "create", "edit"],
        "reports": ["view"],
        "settings": ["view"],
        "users": [],
    },
    "viewer": {
        "dashboard": ["view"],
        "products": ["view"],
        "orders": ["view"],
        "proforma_orders": ["view"],
        "customers": ["view"],
        "transporters": ["view"],
        "sales": ["view"],
        "expenses": ["view"],
        "reports": ["view"],
        "settings": ["view"],
        "users": [],
    },
}

security = HTTPBearer(auto_error=False)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user or not user.is_active:
            raise HTTPException(status_code=401, detail="User not found or inactive")
        return user
    finally:
        db.close()

def require_permission(module, action):
    def dependency(user: User = Depends(get_current_user)):
        perms = ROLE_PERMISSIONS.get(user.role, {}).get(module, [])
        if action not in perms:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return dependency


class Product(Base):
    __tablename__ = "products"
    id = Column(Integer, primary_key=True, index=True)
    part_no = Column(String, default="")
    name = Column(String, nullable=False)
    category = Column(String, default="")
    size = Column(String, default="")
    load_rating = Column(String, default="")
    material = Column(String, default="FRP")
    color = Column(String, default="Grey")
    unit = Column(String, default="Nos")
    hsn_code = Column(String, default="")
    pieces_per_box = Column(Integer, default=1)
    std_packaging = Column(Integer, default=1)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    pricing = relationship("Pricing", back_populates="product", uselist=False, cascade="all,delete-orphan")


class Pricing(Base):
    __tablename__ = "pricing"
    id = Column(Integer, primary_key=True, index=True)
    product_id = Column(Integer, ForeignKey("products.id"), unique=True)
    raw_material_cost = Column(Float, default=0)
    labor_cost = Column(Float, default=0)
    overhead_cost = Column(Float, default=0)
    packing_cost = Column(Float, default=0)
    total_cost = Column(Float, default=0)
    profit_margin = Column(Float, default=20)
    mrp = Column(Float, default=0)
    dealer_price = Column(Float, default=0)
    distributor_price = Column(Float, default=0)
    gst_rate = Column(Float, default=18)
    product = relationship("Product", back_populates="pricing")


class Customer(Base):
    __tablename__ = "customers"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, default="")
    customer_id = Column(String, unique=True, nullable=False)
    gstin = Column(String, default="")
    billing_address = Column(String, default="")
    shipping_address = Column(String, default="")
    state = Column(String, default="")
    district = Column(String, default="")
    city = Column(String, default="")
    pincode = Column(String, default="")
    contact_name = Column(String, default="")
    contact_number = Column(String, default="")
    contact_email = Column(String, default="")
    exec_code = Column(String, default="")
    exec_name = Column(String, default="")
    exec_number = Column(String, default="")
    exec_email = Column(String, default="")
    blacklisted = Column(Integer, default=0)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class Transporter(Base):
    __tablename__ = "transporters"
    id = Column(Integer, primary_key=True, index=True)
    transporter_id = Column(String, unique=True, nullable=False)
    name = Column(String, nullable=False)
    phone = Column(String, default="")
    email = Column(String, default="")
    address = Column(Text, default="")
    state = Column(String, default="")
    district = Column(String, default="")
    city = Column(String, default="")
    pincode = Column(String, default="")
    gst_number = Column(String, default="")
    pan_number = Column(String, default="")
    gst_certificate = Column(String, default="")
    pan_card = Column(String, default="")
    contact_person = Column(String, default="")
    contact_number = Column(String, default="")
    tracking_url_pattern = Column(String, default="")
    blacklisted = Column(Integer, default=0)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class Sale(Base):
    __tablename__ = "sales"
    id = Column(Integer, primary_key=True, index=True)
    invoice_no = Column(String)
    customer_id = Column(Integer, ForeignKey("customers.id"), nullable=True)
    product_id = Column(Integer, ForeignKey("products.id"), nullable=True)
    quantity = Column(Integer, nullable=True)
    unit_price = Column(Float, nullable=True)
    discount_percent = Column(Float, default=0)
    discount_amount = Column(Float, default=0)
    taxable_amount = Column(Float, nullable=True)
    cgst_rate = Column(Float, default=9)
    cgst_amount = Column(Float, default=0)
    sgst_rate = Column(Float, default=9)
    sgst_amount = Column(Float, default=0)
    freight_amount = Column(Float, default=0)
    total_amount = Column(Float, default=0)
    payment_status = Column(String, default="Pending")
    payment_method = Column(String, default="Cash")
    sale_date = Column(DateTime, nullable=True)
    notes = Column(String, default="")
    party_name = Column(String, default="")
    payment_terms = Column(String, default="")
    location = Column(String, default="")
    pincode = Column(String, default="")
    state = Column(String, default="")
    transporter_name = Column(String, default="")
    lr_no = Column(String, default="")
    weight_kgs = Column(Float, default=0)
    weight_pg_fiber = Column(Float, default=0)
    sales_person = Column(String, default="")
    pg_fiber_invoice_no = Column(String, default="")
    pg_fiber_invoice_value = Column(Float, default=0)
    gp = Column(Float, default=0)
    gp_percent = Column(Float, default=0)
    invoice_value = Column(Float, default=0)
    source_csv = Column(String, default="")
    lr_tracking_status = Column(String, default="")
    lr_tracking_url = Column(String, default="")
    lr_last_checked = Column(DateTime, nullable=True)
    customer = relationship("Customer")
    product = relationship("Product")
    items = relationship("SaleItem", cascade="all,delete-orphan", back_populates="sale")


class SaleItem(Base):
    __tablename__ = "sale_items"
    id = Column(Integer, primary_key=True, index=True)
    sale_id = Column(Integer, ForeignKey("sales.id"))
    sl_no = Column(Integer, default=1)
    product_id = Column(Integer, ForeignKey("products.id"))
    quantity = Column(Integer, default=0)
    unit_price = Column(Float, default=0)
    discount_percent = Column(Float, default=0)
    discount_amount = Column(Float, default=0)
    taxable_amount = Column(Float, default=0)
    gst_rate = Column(Float, default=18)
    cgst_amount = Column(Float, default=0)
    sgst_amount = Column(Float, default=0)
    total_amount = Column(Float, default=0)
    basic_amount = Column(Float, default=0)
    sale = relationship("Sale", back_populates="items")
    product = relationship("Product")


class Expense(Base):
    __tablename__ = "expenses"
    id = Column(Integer, primary_key=True, index=True)
    category = Column(String)
    description = Column(String, default="")
    amount = Column(Float)
    vendor = Column(String, default="")
    expense_date = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class Settings(Base):
    __tablename__ = "settings"
    id = Column(Integer, primary_key=True, index=True)
    key = Column(String, unique=True)
    value = Column(String, default="")


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, nullable=False)
    password_hash = Column(String, nullable=False)
    full_name = Column(String, default="")
    email = Column(String, default="")
    role = Column(String, default="viewer")
    is_active = Column(Integer, default=1)
    last_login = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))


class Order(Base):
    __tablename__ = "orders"
    id = Column(Integer, primary_key=True, index=True)
    sl_no = Column(Integer, default=0)
    po_no = Column(String, default="")
    po_date = Column(String, default="")
    customer_name = Column(String, default="")
    billing_site = Column(String, default="")
    shipping_site = Column(String, default="")
    no_of_boxes = Column(Integer, default=0)
    value_excl_gst_freight = Column(Float, default=0)
    invoice_no = Column(String, default="")
    invoice_date = Column(String, default="")
    invoice_amount_excl_gst = Column(Float, default=0)
    weight_kgs = Column(Float, default=0)
    freight_rate_per_kg = Column(Float, default=0)
    transport_charges = Column(Float, default=0)
    invoice_amount = Column(Float, default=0)
    eway_bill_no = Column(String, default="")
    lr_no = Column(String, default="")
    entry_date = Column(String, default="")
    credit_note_amount = Column(Float, default=0)
    credit_note_no = Column(String, default="")
    transporter = Column(String, default="")
    transporter_no = Column(String, default="")


class ProformaOrder(Base):
    __tablename__ = "proforma_orders"
    id = Column(Integer, primary_key=True, index=True)
    pi_no = Column(String, unique=True)
    pi_date = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    customer_id = Column(Integer, ForeignKey("customers.id"))
    billing_site = Column(String, default="")
    shipping_site = Column(String, default="")
    no_of_boxes = Column(Integer, default=0)
    total_qty = Column(Integer, default=0)
    value_excl_gst = Column(Float, default=0)
    gst_amount = Column(Float, default=0)
    total_amount = Column(Float, default=0)
    freight_amount = Column(Float, default=0)
    payment_status = Column(String, default="Pending")
    payment_method = Column(String, default="Cash")
    transport_mode = Column(String, default="")
    delivery_days = Column(Integer, default=30)
    notes = Column(String, default="")
    terms = Column(Text, default="")
    order_type = Column(String, default="PI")
    po_no = Column(String, default="")
    po_date = Column(DateTime, nullable=True)
    purchase_total = Column(Float, default=0)
    transport_cost = Column(Float, default=0)
    gross_profit = Column(Float, default=0)
    net_profit = Column(Float, default=0)
    transporter_id = Column(Integer, ForeignKey("transporters.id"), nullable=True)
    whatsapp_status = Column(String, default="pending")
    status = Column(String, default="draft")
    discount_scheme_applied = Column(Integer, default=0)
    discount_percent = Column(Float, default=0)
    discount_amount = Column(Float, default=0)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    updated_at = Column(DateTime, default=lambda: datetime.now(timezone.utc), onupdate=lambda: datetime.now(timezone.utc))
    customer = relationship("Customer")
    items = relationship("ProformaOrderItem", back_populates="proforma_order", cascade="all,delete-orphan")


class ProformaOrderItem(Base):
    __tablename__ = "proforma_order_items"
    id = Column(Integer, primary_key=True, index=True)
    proforma_order_id = Column(Integer, ForeignKey("proforma_orders.id"))
    sl_no = Column(Integer)
    product_id = Column(Integer, ForeignKey("products.id"))
    part_no = Column(String, default="")
    description = Column(String, default="")
    size = Column(String, default="")
    category = Column(String, default="")
    qty_boxes = Column(Integer, default=1)
    std_packaging = Column(Integer, default=1)
    pieces_per_box = Column(Integer, default=1)
    final_qty = Column(Integer, default=0)
    mrp = Column(Float, default=0)
    d1 = Column(Float, default=0)
    d2 = Column(Float, default=0)
    d3 = Column(Float, default=0)
    d4 = Column(Float, default=0)
    d5 = Column(Float, default=0)
    cd = Column(Float, default=0)
    discount_percent = Column(Float, default=0)
    net_rate = Column(Float, default=0)
    lock_hinge = Column(Integer, default=0)
    basic_amount = Column(Float, default=0)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    proforma_order = relationship("ProformaOrder", back_populates="items")
    product = relationship("Product")


class BillingSite(Base):
    __tablename__ = "billing_sites"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    address = Column(String, default="")
    phone = Column(String, default="")
    email = Column(String, default="")
    website = Column(String, default="")
    gstin = Column(String, default="")
    state_code = Column(String, default="")
    pan = Column(String, default="")


class PurchaseRate(Base):
    __tablename__ = "purchase_rates"
    id = Column(Integer, primary_key=True, index=True)
    product_id = Column(Integer, ForeignKey("products.id"))
    rate = Column(Float, default=0)
    supplier = Column(String, default="")
    effective_date = Column(Date, default=date.today)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    product = relationship("Product")


class TransporterQuote(Base):
    __tablename__ = "transporter_quotes"
    id = Column(Integer, primary_key=True, index=True)
    order_id = Column(Integer, ForeignKey("proforma_orders.id"))
    transporter_id = Column(Integer, ForeignKey("transporters.id"))
    rate_per_kg = Column(Float, default=0)
    total_cost = Column(Float, default=0)
    status = Column(String, default="pending")
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))
    proforma_order = relationship("ProformaOrder")
    transporter = relationship("Transporter")


class AuditLog(Base):
    __tablename__ = "audit_logs"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, nullable=True)
    username = Column(String, default="")
    action = Column(String, nullable=False)
    resource = Column(String, default="")
    resource_id = Column(String, default="")
    details = Column(Text, default="")
    ip_address = Column(String, default="")
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


class TokenBlacklist(Base):
    __tablename__ = "token_blacklist"
    id = Column(Integer, primary_key=True, index=True)
    token = Column(Text, nullable=False)
    user_id = Column(Integer, nullable=True)
    expires_at = Column(DateTime, nullable=False)
    created_at = Column(DateTime, default=lambda: datetime.now(timezone.utc))


def audit_log(user, action, resource="", resource_id="", details="", request=None):
    try:
        db = SessionLocal()
        ip = ""
        if request:
            ip = request.headers.get("x-forwarded-for", request.client.host if request.client else "")
        log = AuditLog(
            user_id=user.id if user else None,
            username=user.username if user else "system",
            action=action, resource=resource, resource_id=str(resource_id),
            details=details, ip_address=ip
        )
        db.add(log)
        db.commit()
        db.close()
    except Exception as e:
        logger.warning(f"Audit log failed: {e}")


@app.get("/api/db-info")
def db_info(user: User = Depends(require_permission("settings", "view"))):
    return {
        "has_database_url_key": "DATABASE_URL" in os.environ,
        "env_key_count": len(os.environ)
    }


@app.on_event("startup")
def startup_event():
    Base.metadata.create_all(bind=engine)
    with engine.connect() as conn:
        def safe_ddl(sql):
            try:
                conn.execute(text(sql))
                conn.commit()
            except Exception:
                conn.rollback()
        safe_ddl("CREATE TABLE IF NOT EXISTS proforma_orders (id SERIAL PRIMARY KEY, pi_no VARCHAR UNIQUE, pi_date TIMESTAMP, customer_id INTEGER REFERENCES customers(id), billing_site VARCHAR DEFAULT '', shipping_site VARCHAR DEFAULT '', no_of_boxes INTEGER DEFAULT 0, total_qty INTEGER DEFAULT 0, value_excl_gst FLOAT DEFAULT 0, gst_amount FLOAT DEFAULT 0, total_amount FLOAT DEFAULT 0, freight_amount FLOAT DEFAULT 0, payment_status VARCHAR DEFAULT 'Pending', payment_method VARCHAR DEFAULT 'Cash', transport_mode VARCHAR DEFAULT '', delivery_days INTEGER DEFAULT 30, notes VARCHAR DEFAULT '', terms TEXT DEFAULT '', order_type VARCHAR DEFAULT 'PI', created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        safe_ddl("CREATE TABLE IF NOT EXISTS proforma_order_items (id SERIAL PRIMARY KEY, proforma_order_id INTEGER REFERENCES proforma_orders(id), sl_no INTEGER, product_id INTEGER REFERENCES products(id), part_no VARCHAR DEFAULT '', description VARCHAR DEFAULT '', size VARCHAR DEFAULT '', category VARCHAR DEFAULT '', qty_boxes INTEGER DEFAULT 1, std_packaging INTEGER DEFAULT 1, pieces_per_box INTEGER DEFAULT 1, final_qty INTEGER DEFAULT 0, mrp FLOAT DEFAULT 0, d1 FLOAT DEFAULT 0, d2 FLOAT DEFAULT 0, d3 FLOAT DEFAULT 0, d4 FLOAT DEFAULT 0, d5 FLOAT DEFAULT 0, cd FLOAT DEFAULT 0, discount_percent FLOAT DEFAULT 0, net_rate FLOAT DEFAULT 0, lock_hinge INTEGER DEFAULT 0, basic_amount FLOAT DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)")
        safe_ddl("ALTER TABLE products ADD COLUMN IF NOT EXISTS part_no VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE products ADD COLUMN IF NOT EXISTS pieces_per_box INTEGER DEFAULT 1")
        safe_ddl("ALTER TABLE products ADD COLUMN IF NOT EXISTS std_packaging INTEGER DEFAULT 1")
        safe_ddl("ALTER TABLE orders ADD COLUMN IF NOT EXISTS customer_name VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE sales DROP CONSTRAINT IF EXISTS sales_invoice_no_key")
        safe_ddl("ALTER TABLE transporters ADD COLUMN IF NOT EXISTS tracking_url_pattern VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE sales ADD COLUMN IF NOT EXISTS lr_tracking_status VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE sales ADD COLUMN IF NOT EXISTS lr_tracking_url VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE sales ADD COLUMN IF NOT EXISTS lr_last_checked TIMESTAMP")
        new_sale_cols = [
            "party_name", "payment_terms", "location", "pincode", "state",
            "transporter_name", "lr_no", "weight_kgs", "weight_pg_fiber",
            "sales_person", "pg_fiber_invoice_no", "pg_fiber_invoice_value",
            "gp", "gp_percent", "source_csv"
        ]
        for col in new_sale_cols:
            col_type = "FLOAT" if col in ("weight_kgs","weight_pg_fiber","pg_fiber_invoice_value","gp","gp_percent") else "VARCHAR DEFAULT ''"
            safe_ddl(f"ALTER TABLE sales ADD COLUMN IF NOT EXISTS {col} {col_type}")
        safe_ddl("UPDATE sales SET invoice_value = '0' WHERE invoice_value IS NULL OR invoice_value = '' OR invoice_value = 'None' OR invoice_value = '\\u2013'")
        safe_ddl("ALTER TABLE sales ALTER COLUMN invoice_value TYPE FLOAT USING invoice_value::float")
        safe_ddl("ALTER TABLE sales ADD COLUMN IF NOT EXISTS invoice_value FLOAT DEFAULT 0")
        safe_ddl("UPDATE sales SET invoice_value = total_amount WHERE invoice_value = 0 AND total_amount > 0")
        safe_ddl("UPDATE sales SET total_amount = invoice_value WHERE invoice_value > 0 AND (total_amount = 0 OR total_amount = freight_amount)")
        safe_ddl("UPDATE sales SET total_amount = invoice_value WHERE invoice_value > freight_amount AND freight_amount > 0 AND total_amount = freight_amount")
        customer_cols = [
            ("customer_id", "VARCHAR DEFAULT ''"),
            ("gstin", "VARCHAR DEFAULT ''"),
            ("billing_address", "VARCHAR DEFAULT ''"),
            ("shipping_address", "VARCHAR DEFAULT ''"),
            ("state", "VARCHAR DEFAULT ''"),
            ("district", "VARCHAR DEFAULT ''"),
            ("city", "VARCHAR DEFAULT ''"),
            ("pincode", "VARCHAR DEFAULT ''"),
            ("contact_name", "VARCHAR DEFAULT ''"),
            ("contact_number", "VARCHAR DEFAULT ''"),
            ("contact_email", "VARCHAR DEFAULT ''"),
            ("exec_code", "VARCHAR DEFAULT ''"),
            ("exec_name", "VARCHAR DEFAULT ''"),
            ("exec_number", "VARCHAR DEFAULT ''"),
            ("exec_email", "VARCHAR DEFAULT ''"),
            ("blacklisted", "INTEGER DEFAULT 0"),
            ("created_at", "TIMESTAMP DEFAULT CURRENT_TIMESTAMP"),
        ]
        for col_name, col_type in customer_cols:
            safe_ddl(f"ALTER TABLE customers ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
        safe_ddl("UPDATE customers SET customer_id = 'C' || id WHERE customer_id IS NULL OR customer_id = ''")
        safe_ddl("ALTER TABLE customers ADD CONSTRAINT customers_customer_id_unique UNIQUE (customer_id)")
        safe_ddl("ALTER TABLE users ADD COLUMN IF NOT EXISTS full_name VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE users ADD COLUMN IF NOT EXISTS email VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE users ADD COLUMN IF NOT EXISTS is_active INTEGER DEFAULT 1")
        safe_ddl("ALTER TABLE users ADD COLUMN IF NOT EXISTS last_login TIMESTAMP")
        safe_ddl("ALTER TABLE users ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP")
        safe_ddl("UPDATE users SET role = 'admin' WHERE username = 'admin' AND (role = 'user' OR role IS NULL OR role = '')")
        # Proforma Orders new columns
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS po_no VARCHAR DEFAULT ''")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS po_date TIMESTAMP")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS purchase_total FLOAT DEFAULT 0")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS transport_cost FLOAT DEFAULT 0")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS gross_profit FLOAT DEFAULT 0")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS net_profit FLOAT DEFAULT 0")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS transporter_id INTEGER REFERENCES transporters(id)")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS whatsapp_status VARCHAR DEFAULT 'pending'")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS status VARCHAR DEFAULT 'draft'")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS discount_scheme_applied INTEGER DEFAULT 0")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS discount_percent FLOAT DEFAULT 0")
        safe_ddl("ALTER TABLE proforma_orders ADD COLUMN IF NOT EXISTS discount_amount FLOAT DEFAULT 0")
        # Sale items table
        safe_ddl("""CREATE TABLE IF NOT EXISTS sale_items (
            id SERIAL PRIMARY KEY,
            sale_id INTEGER REFERENCES sales(id),
            sl_no INTEGER DEFAULT 1,
            product_id INTEGER REFERENCES products(id),
            quantity INTEGER DEFAULT 0,
            unit_price FLOAT DEFAULT 0,
            discount_percent FLOAT DEFAULT 0,
            discount_amount FLOAT DEFAULT 0,
            taxable_amount FLOAT DEFAULT 0,
            gst_rate FLOAT DEFAULT 18,
            cgst_amount FLOAT DEFAULT 0,
            sgst_amount FLOAT DEFAULT 0,
            total_amount FLOAT DEFAULT 0,
            basic_amount FLOAT DEFAULT 0
        )""")
    backfill_part_numbers()
    backfill_pieces_per_box()
    backfill_product_names()
    seed_data()
    seed_billing_sites()


PIECES_PER_BOX_MAP = {
    "10x10": 12, "12x12": 6, "15x15": 6, "18x18": 4,
    "21x21": 3, "24x24": 2, "26x26": 1, "28x28": 1,
    "30x30": 1, "36x36": 1, "42x42": 1,
    "12x18": 6, "12x24": 5, "18x24": 3,
    "250x250": 12, "300x300": 6, "380x380": 6, "450x450": 4,
    "530x530": 3, "600x600": 2, "660x660": 1, "700x700": 1, "710x710": 1,
    "750x750": 1, "900x900": 1, "1060x1060": 1, "1065x1065": 1,
    "300x450": 6, "300x600": 5, "450x600": 3,
}


def backfill_pieces_per_box():
    db = SessionLocal()
    try:
        products = db.query(Product).all()
        updated = 0
        for p in products:
            size_lower = (p.size or "").lower().replace(" ", "")
            ppb = PIECES_PER_BOX_MAP.get(size_lower)
            if ppb and p.pieces_per_box != ppb:
                p.pieces_per_box = ppb
                p.std_packaging = ppb
                updated += 1
            pn = (p.part_no or "").upper().replace(" ", "")
            if pn.startswith("FRP012"):
                if p.load_rating != "10 Ton":
                    p.load_rating = "10 Ton"
                    updated += 1
            elif pn.startswith("FRP01") or pn.startswith("FRP04") or pn.startswith("FRP10"):
                if p.load_rating != "5 Ton" or not p.load_rating:
                    p.load_rating = "5 Ton"
                    updated += 1
        if updated:
            db.commit()
            print(f"Backfilled pieces_per_box and tonnage for {updated} products")
    finally:
        db.close()


PART_NO_CSV = [
    ("FRP01101-GRY", "Raksha FRP Manhole Cover 10x10 - 5 Ton Grey"),
    ("FRP01103-GRY", "Raksha FRP Manhole Cover 12x12 - 5 Ton Grey"),
    ("FRP01106-GRY", "Raksha FRP Manhole Cover 15x15 - 5 Ton Grey"),
    ("FRP01109-GRY", "Raksha FRP Manhole Cover 18x18 - 5 Ton Grey"),
    ("FRP01112-GRY", "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey"),
    ("FRP01115-GRY", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey"),
    ("FRP01117-GRY", "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey"),
    ("FRP01119-GRY", "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey"),
    ("FRP01121-GRY", "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey"),
    ("FRP01127-GRY", "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey"),
    ("FRP04106-GRY", "Raksha FRP Manhole Cover 12x18 - 5 Ton Grey"),
    ("FRP04112-GRY", "Raksha FRP Manhole Cover 12x24 - 5 Ton Grey"),
    ("FRP10106-GRY", "Raksha FRP Manhole Cover 18x24 - 5 Ton Grey"),
    ("FRP01101-WH", "Raksha FRP Manhole Cover 10x10 - 5 Ton White"),
    ("FRP01103-WH", "Raksha FRP Manhole Cover 12x12 - 5 Ton White"),
    ("FRP01106-WH", "Raksha FRP Manhole Cover 15x15 - 5 Ton White"),
    ("FRP01109-WH", "Raksha FRP Manhole Cover 18x18 - 5 Ton White"),
    ("FRP01112-WH", "Raksha FRP Manhole Cover 21x21 - 5 Ton White"),
    ("FRP01115-WH", "Raksha FRP Manhole Cover 24x24 - 5 Ton White"),
    ("FRP01117-WH", "Raksha FRP Manhole Cover 26x26 - 5 Ton White"),
    ("FRP01119-WH", "Raksha FRP Manhole Cover 28x28 - 5 Ton White"),
    ("FRP01121-WH", "Raksha FRP Manhole Cover 30x30 - 5 Ton White"),
    ("FRP01127-WH", "Raksha FRP Manhole Cover 36x36 - 5 Ton White"),
    ("FRP04106-WH", "Raksha FRP Manhole Cover 12x18 - 5 Ton White"),
    ("FRP04112-WH", "Raksha FRP Manhole Cover 12x24 - 5 Ton White"),
    ("FRP10106-WH", "Raksha FRP Manhole Cover 18x24 - 5 Ton White"),
    ("FRP01112-GRYL", "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey (with Lock)"),
    ("FRP01115-GRYL", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (with Lock)"),
    ("FRP01117-GRYL", "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey (with Lock)"),
    ("FRP01119-GRYL", "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey (with Lock)"),
    ("FRP01121-GRYL", "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey (with Lock)"),
    ("FRP01127-GRYL", "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey (with Lock)"),
    ("FRP01112-WHL", "Raksha FRP Manhole Cover 21x21 - 5 Ton White (with Lock)"),
    ("FRP01115-WHL", "Raksha FRP Manhole Cover 24x24 - 5 Ton White (with Lock)"),
    ("FRP01117-WHL", "Raksha FRP Manhole Cover 26x26 - 5 Ton White (with Lock)"),
    ("FRP01119-WHL", "Raksha FRP Manhole Cover 28x28 - 5 Ton White (with Lock)"),
    ("FRP01121-WHL", "Raksha FRP Manhole Cover 30x30 - 5 Ton White (with Lock)"),
    ("FRP01127-WHL", "Raksha FRP Manhole Cover 36x36 - 5 Ton White (with Lock)"),
    ("FRP01115-GRYH", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges)"),
    ("FRP01115-WHH", "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges)"),
    ("FRP01115-GRY/H&L", "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges & Lock)"),
    ("FRP01115-WH/H&L", "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges & Lock)"),
    ("FRP01209-GRY", "Raksha FRP Manhole Cover 18x18 - 10 Ton Grey"),
    ("FRP01215-GRY", "Raksha FRP Manhole Cover 24x24 - 10 Ton Grey"),
    ("FRP01219-GRY", "Raksha FRP Manhole Cover 28x28 - 10 Ton Grey"),
    ("FRP01221-GRY", "Raksha FRP Manhole Cover 30x30 - 10 Ton Grey"),
    ("FRP01233-GRY", "Raksha FRP Manhole Cover 42x42 - 10 Ton Grey"),
    ("FRP01209-WH", "Raksha FRP Manhole Cover 18x18 - 10 Ton White"),
    ("FRP01215-WH", "Raksha FRP Manhole Cover 24x24 - 10 Ton White"),
    ("FRP01219-WH", "Raksha FRP Manhole Cover 28x28 - 10 Ton White"),
    ("FRP01221-WH", "Raksha FRP Manhole Cover 30x30 - 10 Ton White"),
    ("FRP01233-WH", "Raksha FRP Manhole Cover 42x42 - 10 Ton White"),
    ("RGC00001-GRY", "Raksha FRP Gully Cover 10x10 - Grey"),
    ("RGC00002-GRY", "Raksha FRP Gully Cover 12x12 - Grey"),
    ("RGC00003-GRY", "Raksha FRP Gully Cover 15x15 - Grey"),
    ("RGC00004-GRY", "Raksha FRP Gully Cover 18x18 - Grey"),
    ("RGC00005-GRY", "Raksha FRP Gully Cover 24x24 - Grey"),
    ("RGC00001-WH", "Raksha FRP Gully Cover 10x10 - White"),
    ("RGC00002-WH", "Raksha FRP Gully Cover 12x12 - White"),
    ("RGC00003-WH", "Raksha FRP Gully Cover 15x15 - White"),
    ("RGC00004-WH", "Raksha FRP Gully Cover 18x18 - White"),
    ("RGC00005-WH", "Raksha FRP Gully Cover 24x24 - White"),
]


def generate_part_no(product):
    if product.part_no:
        return product.part_no
    name_lower = (product.name or "").lower()
    size = (product.size or "").lower().replace(" ", "")
    color = (product.color or "").lower()
    has_lock = "lock" in name_lower
    has_hinges = "hinge" in name_lower
    is_gully = "gully" in name_lower
    is_10ton = "10 ton" in name_lower

    size_dim_map = {
        "250x250": ("10x10", "10x10"), "300x300": ("12x12", "12x12"),
        "380x380": ("15x15", "15x15"), "450x450": ("18x18", "18x18"),
        "535x535": ("21x21", "21x21"), "600x600": ("24x24", "24x24"),
        "660x660": ("26x26", "26x26"), "710x710": ("28x28", "28x28"),
        "760x760": ("30x30", "30x30"), "900x900": ("36x36", "36x36"),
        "1065x1065": ("42x42", "42x42"),
        "300x450": ("12x18", "12x18"), "300x600": ("12x24", "12x24"),
        "450x600": ("18x24", "18x24"),
    }
    nominal = size_dim_map.get(size)
    if not nominal:
        m = re.search(r'(\d+)\s*x\s*(\d+)', product.size or product.name or "")
        if m:
            nominal = (f"{m.group(1)}x{m.group(2)}", f"{m.group(1)}x{m.group(2)}")
        else:
            return ""

    nom_lower = nominal[0]
    color_name = "White" if "white" in color else "Grey"
    tonnage = "10 Ton" if is_10ton else "5 Ton"

    suffix = ""
    if has_hinges and has_lock:
        suffix = " (Double Hinges & Lock)"
    elif has_hinges:
        suffix = " (Double Hinges)"
    elif has_lock:
        suffix = " (with Lock)"

    prefix = "RGC" if is_gully else "FRP"
    if is_gully:
        csv_desc = f"Raksha FRP Gully Cover {nom_lower} - {color_name}{suffix}"
    else:
        csv_desc = f"Raksha FRP Manhole Cover {nom_lower} - {tonnage} {color_name}{suffix}"

    for pn, desc in PART_NO_CSV:
        if desc.strip().lower() == csv_desc.strip().lower():
            return pn

    return ""


def backfill_part_numbers():
    db = SessionLocal()
    try:
        updated = 0
        for p in db.query(Product).filter((Product.part_no == "") | (Product.part_no.is_(None))).all():
            pn = generate_part_no(p)
            if pn:
                p.part_no = pn
                updated += 1
        if updated:
            db.commit()
            print(f"Backfilled part_no for {updated} products")
    finally:
        db.close()


def get_new_product_name(part_no, old_name=""):
    pn = (part_no or "").upper().replace(" ", "")
    is_gully = pn.startswith("RGC")
    is_10ton = pn.startswith("FRP012")

    suffix = ""
    if pn.endswith("/H&L"):
        suffix = " (Double Hinges & Lock)"
    elif pn.endswith("H") and not pn.endswith("-WH"):
        suffix = " (Double Hinges)"
    elif pn.endswith("L") and "GRY" not in pn[-4:]:
        suffix = " (with Lock)"

    color = "White" if "-WH" in pn else "Grey"

    size_map = {
        "FRP01101": "10x10", "FRP01103": "12x12", "FRP01106": "15x15",
        "FRP01109": "18x18", "FRP01112": "21x21", "FRP01115": "24x24",
        "FRP01117": "26x26", "FRP01119": "28x28", "FRP01121": "30x30",
        "FRP01127": "36x36", "FRP04106": "12x18", "FRP04112": "12x24",
        "FRP10106": "18x24", "FRP01209": "18x18", "FRP01215": "24x24",
        "FRP01219": "28x28", "FRP01221": "30x30", "FRP01233": "42x42",
        "RGC00001": "10x10", "RGC00002": "12x12", "RGC00003": "15x15",
        "RGC00004": "18x18", "RGC00005": "24x24",
    }

    size = ""
    for prefix, sz in size_map.items():
        if pn.startswith(prefix):
            size = sz
            break

    if not size:
        m = re.search(r"(\d+)\s*[xX]\s*(\d+)", old_name or "")
        if m:
            size = f"{m.group(1)}x{m.group(2)}"

    if is_gully:
        return f"Raksha FRP Gully Cover {size} - {color}{suffix}"

    tonnage = "10 Ton" if is_10ton else "5 Ton"
    return f"Raksha FRP Manhole Cover {size} - {tonnage} {color}{suffix}"


def backfill_product_names():
    db = SessionLocal()
    try:
        updated = 0
        for p in db.query(Product).all():
            new_name = get_new_product_name(p.part_no, p.name)
            if new_name and new_name != p.name:
                p.name = new_name
                updated += 1
        if updated:
            db.commit()
            print(f"Backfilled product names for {updated} products")
    finally:
        db.close()


def seed_data():
    db = SessionLocal()
    try:
        admin = db.query(User).filter(User.username == "admin").first()
        admin_password = os.environ.get("ADMIN_PASSWORD")
        if not admin:
            if not admin_password:
                logger.critical("ADMIN_PASSWORD env var not set and no admin user exists. Cannot seed admin account.")
                raise RuntimeError("ADMIN_PASSWORD environment variable is required on first run")
            pw_hash = bcrypt.hashpw(admin_password.encode(), bcrypt.gensalt()).decode()
            admin = User(username="admin", password_hash=pw_hash, full_name="Administrator", email="admin@raksha.com", role="admin", is_active=1)
            db.add(admin)
            db.commit()
            logger.info("Admin user seeded from ADMIN_PASSWORD env var")
        elif not admin.password_hash.startswith("$2"):
            if not admin_password:
                logger.warning("Admin password hash is not bcrypt format, but ADMIN_PASSWORD not set. Skipping rehash.")
            else:
                admin.password_hash = bcrypt.hashpw(admin_password.encode(), bcrypt.gensalt()).decode()
                admin.role = "admin"
                admin.full_name = admin.full_name or "Administrator"
                db.commit()
                logger.info("Admin password rehashed from ADMIN_PASSWORD env var")

        if db.query(Product).count() > 0:
            return

        products = [
            # Manhole Cover - Grey (5 Ton)
            {"part_no": "FRP01101-GRY", "name": "Raksha FRP Manhole Cover 10x10 - 5 Ton Grey", "category": "Manhole Cover", "size": "10x10", "color": "Grey", "rate": 190, "mrp": 686, "ppb": 12, "tonnage": "5 Ton"},
            {"part_no": "FRP01103-GRY", "name": "Raksha FRP Manhole Cover 12x12 - 5 Ton Grey", "category": "Manhole Cover", "size": "12x12", "color": "Grey", "rate": 242, "mrp": 830, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01106-GRY", "name": "Raksha FRP Manhole Cover 15x15 - 5 Ton Grey", "category": "Manhole Cover", "size": "15x15", "color": "Grey", "rate": 310, "mrp": 1046, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01109-GRY", "name": "Raksha FRP Manhole Cover 18x18 - 5 Ton Grey", "category": "Manhole Cover", "size": "18x18", "color": "Grey", "rate": 455, "mrp": 1536, "ppb": 4, "tonnage": "5 Ton"},
            {"part_no": "FRP01112-GRY", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey", "category": "Manhole Cover", "size": "21x21", "color": "Grey", "rate": 640, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-GRY", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 765, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-GRY", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey", "category": "Manhole Cover", "size": "26x26", "color": "Grey", "rate": 1130, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-GRY", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey", "category": "Manhole Cover", "size": "28x28", "color": "Grey", "rate": 1500, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-GRY", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey", "category": "Manhole Cover", "size": "30x30", "color": "Grey", "rate": 1750, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-GRY", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey", "category": "Manhole Cover", "size": "36x36", "color": "Grey", "rate": 3200, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP04106-GRY", "name": "Raksha FRP Manhole Cover 12x18 - 5 Ton Grey", "category": "Manhole Cover", "size": "12x18", "color": "Grey", "rate": 350, "mrp": 1154, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP04112-GRY", "name": "Raksha FRP Manhole Cover 12x24 - 5 Ton Grey", "category": "Manhole Cover", "size": "12x24", "color": "Grey", "rate": 500, "mrp": 1624, "ppb": 5, "tonnage": "5 Ton"},
            {"part_no": "FRP10106-GRY", "name": "Raksha FRP Manhole Cover 18x24 - 5 Ton Grey", "category": "Manhole Cover", "size": "18x24", "color": "Grey", "rate": 620, "mrp": 2020, "ppb": 3, "tonnage": "5 Ton"},
            # Manhole Cover - White (5 Ton)
            {"part_no": "FRP01101-WH", "name": "Raksha FRP Manhole Cover 10x10 - 5 Ton White", "category": "Manhole Cover", "size": "10x10", "color": "White", "rate": 190, "mrp": 686, "ppb": 12, "tonnage": "5 Ton"},
            {"part_no": "FRP01103-WH", "name": "Raksha FRP Manhole Cover 12x12 - 5 Ton White", "category": "Manhole Cover", "size": "12x12", "color": "White", "rate": 242, "mrp": 830, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01106-WH", "name": "Raksha FRP Manhole Cover 15x15 - 5 Ton White", "category": "Manhole Cover", "size": "15x15", "color": "White", "rate": 310, "mrp": 1046, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP01109-WH", "name": "Raksha FRP Manhole Cover 18x18 - 5 Ton White", "category": "Manhole Cover", "size": "18x18", "color": "White", "rate": 455, "mrp": 1536, "ppb": 4, "tonnage": "5 Ton"},
            {"part_no": "FRP01112-WH", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton White", "category": "Manhole Cover", "size": "21x21", "color": "White", "rate": 640, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WH", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 765, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-WH", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton White", "category": "Manhole Cover", "size": "26x26", "color": "White", "rate": 1130, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-WH", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton White", "category": "Manhole Cover", "size": "28x28", "color": "White", "rate": 1500, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-WH", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton White", "category": "Manhole Cover", "size": "30x30", "color": "White", "rate": 1750, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-WH", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton White", "category": "Manhole Cover", "size": "36x36", "color": "White", "rate": 3200, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP04106-WH", "name": "Raksha FRP Manhole Cover 12x18 - 5 Ton White", "category": "Manhole Cover", "size": "12x18", "color": "White", "rate": 350, "mrp": 1154, "ppb": 6, "tonnage": "5 Ton"},
            {"part_no": "FRP04112-WH", "name": "Raksha FRP Manhole Cover 12x24 - 5 Ton White", "category": "Manhole Cover", "size": "12x24", "color": "White", "rate": 500, "mrp": 1624, "ppb": 5, "tonnage": "5 Ton"},
            {"part_no": "FRP10106-WH", "name": "Raksha FRP Manhole Cover 18x24 - 5 Ton White", "category": "Manhole Cover", "size": "18x24", "color": "White", "rate": 620, "mrp": 2020, "ppb": 3, "tonnage": "5 Ton"},
            # Manhole Cover - Grey With Lock (5 Ton)
            {"part_no": "FRP01112-GRYL", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "21x21", "color": "Grey", "rate": 710, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-GRYL", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 835, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-GRYL", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "26x26", "color": "Grey", "rate": 1270, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-GRYL", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "28x28", "color": "Grey", "rate": 1640, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-GRYL", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "30x30", "color": "Grey", "rate": 1890, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-GRYL", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton Grey (with Lock)", "category": "Manhole Cover", "size": "36x36", "color": "Grey", "rate": 3340, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            # Manhole Cover - White With Lock (5 Ton)
            {"part_no": "FRP01112-WHL", "name": "Raksha FRP Manhole Cover 21x21 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "21x21", "color": "White", "rate": 710, "mrp": 2130, "ppb": 3, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WHL", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 835, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01117-WHL", "name": "Raksha FRP Manhole Cover 26x26 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "26x26", "color": "White", "rate": 1270, "mrp": 3266, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01119-WHL", "name": "Raksha FRP Manhole Cover 28x28 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "28x28", "color": "White", "rate": 1640, "mrp": 4934, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01121-WHL", "name": "Raksha FRP Manhole Cover 30x30 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "30x30", "color": "White", "rate": 1890, "mrp": 5854, "ppb": 1, "tonnage": "5 Ton"},
            {"part_no": "FRP01127-WHL", "name": "Raksha FRP Manhole Cover 36x36 - 5 Ton White (with Lock)", "category": "Manhole Cover", "size": "36x36", "color": "White", "rate": 3340, "mrp": 11454, "ppb": 1, "tonnage": "5 Ton"},
            # Manhole Cover - Double Hinges (5 Ton)
            {"part_no": "FRP01115-GRYH", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges)", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 965, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WHH", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges)", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 965, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            # Manhole Cover - Double Hinges & Lock (5 Ton)
            {"part_no": "FRP01115-GRY/H&L", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton Grey (Double Hinges & Lock)", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 1065, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            {"part_no": "FRP01115-WH/H&L", "name": "Raksha FRP Manhole Cover 24x24 - 5 Ton White (Double Hinges & Lock)", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 1065, "mrp": 2560, "ppb": 2, "tonnage": "5 Ton"},
            # Heavy Duty Manhole Cover - Grey (10 Ton)
            {"part_no": "FRP01209-GRY", "name": "Raksha FRP Manhole Cover 18x18 - 10 Ton Grey", "category": "Manhole Cover", "size": "18x18", "color": "Grey", "rate": 1200, "mrp": 3340, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01215-GRY", "name": "Raksha FRP Manhole Cover 24x24 - 10 Ton Grey", "category": "Manhole Cover", "size": "24x24", "color": "Grey", "rate": 2200, "mrp": 6042, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01219-GRY", "name": "Raksha FRP Manhole Cover 28x28 - 10 Ton Grey", "category": "Manhole Cover", "size": "28x28", "color": "Grey", "rate": 3100, "mrp": 8504, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01221-GRY", "name": "Raksha FRP Manhole Cover 30x30 - 10 Ton Grey", "category": "Manhole Cover", "size": "30x30", "color": "Grey", "rate": 3800, "mrp": 10414, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01233-GRY", "name": "Raksha FRP Manhole Cover 42x42 - 10 Ton Grey", "category": "Manhole Cover", "size": "42x42", "color": "Grey", "rate": 11000, "mrp": 30208, "ppb": 1, "tonnage": "10 Ton"},
            # Heavy Duty Manhole Cover - White (10 Ton)
            {"part_no": "FRP01209-WH", "name": "Raksha FRP Manhole Cover 18x18 - 10 Ton White", "category": "Manhole Cover", "size": "18x18", "color": "White", "rate": 1200, "mrp": 3340, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01215-WH", "name": "Raksha FRP Manhole Cover 24x24 - 10 Ton White", "category": "Manhole Cover", "size": "24x24", "color": "White", "rate": 2200, "mrp": 6042, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01219-WH", "name": "Raksha FRP Manhole Cover 28x28 - 10 Ton White", "category": "Manhole Cover", "size": "28x28", "color": "White", "rate": 3100, "mrp": 8504, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01221-WH", "name": "Raksha FRP Manhole Cover 30x30 - 10 Ton White", "category": "Manhole Cover", "size": "30x30", "color": "White", "rate": 3800, "mrp": 10414, "ppb": 1, "tonnage": "10 Ton"},
            {"part_no": "FRP01233-WH", "name": "Raksha FRP Manhole Cover 42x42 - 10 Ton White", "category": "Manhole Cover", "size": "42x42", "color": "White", "rate": 11000, "mrp": 30208, "ppb": 1, "tonnage": "10 Ton"},
            # Gully Cover - Grey
            {"part_no": "RGC00001-GRY", "name": "Raksha FRP Gully Cover 10x10 - Grey", "category": "Gully Cover", "size": "10x10", "color": "Grey", "rate": 240, "mrp": 806, "ppb": 12},
            {"part_no": "RGC00002-GRY", "name": "Raksha FRP Gully Cover 12x12 - Grey", "category": "Gully Cover", "size": "12x12", "color": "Grey", "rate": 325, "mrp": 984, "ppb": 6},
            {"part_no": "RGC00003-GRY", "name": "Raksha FRP Gully Cover 15x15 - Grey", "category": "Gully Cover", "size": "15x15", "color": "Grey", "rate": 440, "mrp": 1380, "ppb": 6},
            {"part_no": "RGC00004-GRY", "name": "Raksha FRP Gully Cover 18x18 - Grey", "category": "Gully Cover", "size": "18x18", "color": "Grey", "rate": 570, "mrp": 2012, "ppb": 4},
            {"part_no": "RGC00005-GRY", "name": "Raksha FRP Gully Cover 24x24 - Grey", "category": "Gully Cover", "size": "24x24", "color": "Grey", "rate": 1160, "mrp": 3910, "ppb": 2},
            # Gully Cover - White
            {"part_no": "RGC00001-WH", "name": "Raksha FRP Gully Cover 10x10 - White", "category": "Gully Cover", "size": "10x10", "color": "White", "rate": 240, "mrp": 806, "ppb": 12},
            {"part_no": "RGC00002-WH", "name": "Raksha FRP Gully Cover 12x12 - White", "category": "Gully Cover", "size": "12x12", "color": "White", "rate": 325, "mrp": 984, "ppb": 6},
            {"part_no": "RGC00003-WH", "name": "Raksha FRP Gully Cover 15x15 - White", "category": "Gully Cover", "size": "15x15", "color": "White", "rate": 440, "mrp": 1380, "ppb": 6},
            {"part_no": "RGC00004-WH", "name": "Raksha FRP Gully Cover 18x18 - White", "category": "Gully Cover", "size": "18x18", "color": "White", "rate": 570, "mrp": 2012, "ppb": 4},
            {"part_no": "RGC00005-WH", "name": "Raksha FRP Gully Cover 24x24 - White", "category": "Gully Cover", "size": "24x24", "color": "White", "rate": 1160, "mrp": 3910, "ppb": 2},
        ]

        pid = 1
        for prod in products:
            p = Product(id=pid, part_no=prod["part_no"], name=prod["name"], category=prod["category"], size=prod["size"], load_rating=prod.get("tonnage", "5 Ton"), material="FRP", color=prod["color"], hsn_code="39259090", pieces_per_box=prod.get("ppb", 1), std_packaging=prod.get("ppb", 1))
            db.add(p)
            db.flush()
            db.add(Pricing(product_id=p.id, raw_material_cost=prod["rate"], total_cost=prod["rate"], profit_margin=20, gst_rate=18, mrp=prod["mrp"]))
            pid += 1

        db.commit()
        print(f"Seeded {pid - 1} products")
    finally:
        db.close()


def seed_billing_sites():
    db = SessionLocal()
    try:
        if db.query(BillingSite).count() > 0:
            return
        sites = [
            {"name": "Diamond Pipes & Tubes Private Limited - Unit 1", "address": "No - 209/394, Hosur Main Road, Chandapur, Bangalore - 560081", "phone": "9341329825", "email": "spi@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "29AAECS9353F1Z0", "state_code": "29", "pan": "AAECS9353F"},
            {"name": "Diamond Pipes & Tubes Private Limited - Unit 2", "address": "No.195/2, 74/74/1, 115/67/1, 81/81/1, Chikkanahalli Road, Bommanahalli, Bengaluru - 560068", "phone": "9341985236", "email": "pi@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "29AAACD5818Q1Z2", "state_code": "29", "pan": "AAACD5818Q"},
            {"name": "Diamond Pipes & Tubes Private Limited - Unit 3", "address": "S.No.70 / 2B,Daman Industrial Estate, Kadaiya, Daman - 396210", "phone": "8306340710", "email": "bds@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "26AAACD5818Q1Z8", "state_code": "26", "pan": "AAACD5818Q"},
            {"name": "Shand Pipe Industry Private Limited - Unit 1", "address": "Industrial Area, Hosur Road, Bommasandra, Bangalore - 560099", "phone": "7815833361", "email": "dpt@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "29ABJPS4076L1ZV", "state_code": "29", "pan": "ABJPS4076L"},
            {"name": "Shand Pipe Industry Private Limited - Unit 2", "address": "Sy. No. 168, Madivala Village, Kasba Hobli, Anekal Taluk, Bengaluru - 562106", "phone": "7815833361", "email": "dpt@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "29ABJPS4076L1ZV", "state_code": "29", "pan": "ABJPS4076L"},
            {"name": "Raksha Pipes Private Limited - Ernakulam", "address": "36/337A, Chettu Kudiyil House, Pullepady Kathrikadavu Road, Kochi, Ernakulam - 682017", "phone": "9562011100", "email": "rishabmkt@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "32AACCV0019M1ZK", "state_code": "32", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Coimbatore", "address": "D.No. 509/1A,Maniakarampalayam Road,Nallampalayam,Coimbatore,Tamil Nadu - 641006", "phone": "9345157327", "email": "raksha_tcy@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "33AACCV0019M1ZI", "state_code": "33", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Vijayawada", "address": "# 25/96, 4th Cross, Ramraj Nagar, Kabela Centre, Vijayawada, Andhra Pradesh - 520012", "phone": "7032959106", "email": "prajay@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "37AACCV0019M1ZA", "state_code": "37", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - JC Road", "address": "No-11, New No - 11/1, 1st B Cross, Fireworks Colony, J C Road, Bangalore - 560002", "phone": "9342209496", "email": "galaxyblr@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "29AACCV0019M1Z7", "state_code": "29", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Belgaum", "address": "C.T.S No-4927/29, Sambhaji Galli Mahadwar Road, Belgaum - 590002", "phone": "9343943148", "email": "galaxy_bel@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "29AACCV0019M1Z7", "state_code": "29", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Pune", "address": "Gate No.1150, Opp.Hotel Abhiruchi, Near Modak International School, Pune-Saswad Road, 10th Mile, Pune - 412308", "phone": "9325411100", "email": "bhavana@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "27AACCV0019M1Z7", "state_code": "27", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Indore", "address": "520, Shekhar Central, A B Road, Manorama Ganj, Indore, Madhya Pradesh", "phone": "9770851100", "email": "desana_ind@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "23AACCV0019M1ZJ", "state_code": "23", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Jaipur", "address": "M - 1, Opp. V K I, Road No. - 03, Near Sharada Sec.School, Kalyan Nagar, Jaipur - 302039", "phone": "9529937091", "email": "emerald@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "08AACCV0019M1ZJ", "state_code": "08", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Ludhiana", "address": "Block No - 30, No - 3450, Street No- 3, Heera Nagar, P.O.Moti Nagar, Ludhiana - 141010", "phone": "9356292307", "email": "desana_ldh@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "03AACCV0019M1ZJ", "state_code": "03", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Kolkata", "address": "NH2 Chakundi Dankuni, Post Dankuni, Dankuni, Hooghly - 712310", "phone": "9339711100", "email": "raksha_kol@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "19AACCV0019M1ZJ", "state_code": "19", "pan": "AACCV0019M"},
            {"name": "Raksha Pipes Private Limited - Patna", "address": "ANA Ware House, TRS Campus, New Byepass Road, Patna, Bihar - 800002", "phone": "7991607272", "email": "raksha_patna@shandgroup.com", "website": "www.rakshapipes.com", "gstin": "10AACCV0019M1ZJ", "state_code": "10", "pan": "AACCV0019M"},
        ]
        for s in sites:
            db.add(BillingSite(**s))
        db.commit()
        print(f"Seeded {len(sites)} billing sites")
    finally:
        db.close()


class ProductIn(BaseModel):
    part_no: str = ""
    name: str
    category: str = ""
    size: str = ""
    load_rating: str = ""
    material: str = "FRP"
    color: str = "Grey"
    unit: str = "Nos"
    hsn_code: str = ""


class PricingIn(BaseModel):
    raw_material_cost: float = Field(0, ge=0)
    mrp: float = Field(0, ge=0)
    labor_cost: float = Field(0, ge=0)
    overhead_cost: float = Field(0, ge=0)
    packing_cost: float = Field(0, ge=0)
    profit_margin: float = Field(20, ge=0)
    gst_rate: float = Field(18, ge=0)


class CustomerIn(BaseModel):
    customer_id: str
    name: str = ""
    gstin: str = ""
    billing_address: str = ""
    shipping_address: str = ""
    state: str = ""
    district: str = ""
    city: str = ""
    pincode: str = ""
    contact_name: str = ""
    contact_number: str = ""
    contact_email: str = ""
    exec_code: str = ""
    exec_name: str = ""
    exec_number: str = ""
    exec_email: str = ""
    blacklisted: int = 0


class TransporterIn(BaseModel):
    transporter_id: str
    name: str
    phone: str
    email: str
    address: str
    state: str
    district: str
    city: str
    pincode: str
    gst_number: str
    pan_number: str
    gst_certificate: str = ""
    pan_card: str = ""
    contact_person: str
    contact_number: str
    tracking_url_pattern: str = ""
    blacklisted: int = 0


class SaleItemIn(BaseModel):
    product_id: int
    quantity: int = Field(0, ge=0)
    unit_price: float = Field(0, ge=0)
    discount_percent: float = Field(0, ge=0, le=100)


class SaleIn(BaseModel):
    customer_id: int
    product_id: int = 0
    quantity: int = Field(0, ge=0)
    unit_price: float = Field(0, ge=0)
    discount_percent: float = Field(0, ge=0, le=100)
    freight_amount: float = Field(0, ge=0)
    invoice_value: float = Field(0, ge=0)
    payment_status: str = "Pending"
    payment_method: str = "Cash"
    notes: str = ""
    transporter_name: str = ""
    lr_no: str = ""
    items: List[SaleItemIn] = []


class ExpenseIn(BaseModel):
    category: str
    description: str = ""
    amount: float = Field(0, ge=0)
    vendor: str = ""
    expense_date: Optional[str] = None


class OrderIn(BaseModel):
    sl_no: int = 0
    po_no: str = ""
    po_date: str = ""
    customer_name: str = ""
    billing_site: str = ""
    shipping_site: str = ""
    no_of_boxes: int = Field(0, ge=0)
    value_excl_gst_freight: float = Field(0, ge=0)
    invoice_no: str = ""
    invoice_date: str = ""
    invoice_amount_excl_gst: float = Field(0, ge=0)
    weight_kgs: float = Field(0, ge=0)
    freight_rate_per_kg: float = Field(0, ge=0)
    transport_charges: float = Field(0, ge=0)
    invoice_amount: float = Field(0, ge=0)
    eway_bill_no: str = ""
    lr_no: str = ""
    entry_date: str = ""
    credit_note_amount: float = Field(0, ge=0)
    credit_note_no: str = ""
    transporter: str = ""
    transporter_no: str = ""


class ProformaOrderItemIn(BaseModel):
    product_id: int
    part_no: str = ""
    description: str = ""
    size: str = ""
    category: str = ""
    qty_boxes: int = Field(1, ge=0)
    std_packaging: int = Field(1, ge=0)
    pieces_per_box: int = Field(1, ge=0)
    final_qty: int = Field(0, ge=0)
    mrp: float = Field(0, ge=0)
    d1: float = Field(0, ge=0, le=100)
    d2: float = Field(0, ge=0, le=100)
    d3: float = Field(0, ge=0, le=100)
    d4: float = Field(0, ge=0, le=100)
    d5: float = Field(0, ge=0, le=100)
    cd: float = Field(0, ge=0, le=100)
    discount_percent: float = Field(0, ge=0, le=100)
    net_rate: float = Field(0, ge=0)
    lock_hinge: int = Field(0, ge=0)
    basic_amount: float = Field(0, ge=0)


class ProformaOrderIn(BaseModel):
    customer_id: int
    billing_site: str = ""
    shipping_site: str = ""
    freight_amount: float = Field(0, ge=0)
    payment_status: str = "Pending"
    payment_method: str = "Cash"
    transport_mode: str = ""
    delivery_days: int = Field(30, ge=0)
    notes: str = ""
    terms: str = ""
    order_type: str = "PI"
    discount_scheme_applied: int = 0
    items: List[ProformaOrderItemIn] = []


# ---- VALIDATION MODELS (replacing raw dict endpoints) ----
class LoginIn(BaseModel):
    username: str
    password: str

class RefreshIn(BaseModel):
    refresh_token: str

class UserCreateIn(BaseModel):
    username: str = Field(..., min_length=3, max_length=50)
    password: str = Field(..., min_length=8)
    full_name: str = ""
    email: str = ""
    role: Literal["admin", "manager", "viewer"] = "viewer"

    @field_validator("password")
    @classmethod
    def validate_password_strength(cls, v):
        if not re.search(r"[A-Z]", v):
            raise ValueError("Password must contain at least one uppercase letter")
        if not re.search(r"[a-z]", v):
            raise ValueError("Password must contain at least one lowercase letter")
        if not re.search(r"\d", v):
            raise ValueError("Password must contain at least one digit")
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", v):
            raise ValueError("Password must contain at least one special character")
        return v

    @field_validator("email")
    @classmethod
    def validate_email(cls, v):
        if v and not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", v):
            raise ValueError("Invalid email format")
        return v

class UserUpdateIn(BaseModel):
    full_name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[int] = None
    password: Optional[str] = None

class ChangePasswordIn(BaseModel):
    current_password: str
    new_password: str = Field(..., min_length=8)

    @field_validator("new_password")
    @classmethod
    def validate_password_strength(cls, v):
        if not re.search(r"[A-Z]", v):
            raise ValueError("Password must contain at least one uppercase letter")
        if not re.search(r"[a-z]", v):
            raise ValueError("Password must contain at least one lowercase letter")
        if not re.search(r"\d", v):
            raise ValueError("Password must contain at least one digit")
        if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", v):
            raise ValueError("Password must contain at least one special character")
        return v

class PurchaseRateIn(BaseModel):
    product_id: int
    rate: float = Field(0, ge=0)
    supplier: str = ""
    effective_date: Optional[str] = None

class PurchaseRateUpdateIn(BaseModel):
    rate: float = Field(0, ge=0)
    supplier: str = ""
    effective_date: Optional[str] = None

class BulkPurchaseRateIn(BaseModel):
    rates: List[PurchaseRateIn]

class TransportUpdateIn(BaseModel):
    transport_cost: float = Field(0, ge=0)

class OrderStatusIn(BaseModel):
    status: str

class WhatsAppSendIn(BaseModel):
    phone: str = Field(..., pattern=r"^\d{10,15}$")
    message: str = ""

class WhatsAppSendPIIn(BaseModel):
    phone: str = Field(..., pattern=r"^\d{10,15}$")

class WhatsAppSendPOIn(BaseModel):
    phone: str = Field(..., pattern=r"^\d{10,15}$")

class WhatsAppTestIn(BaseModel):
    phone: str = Field(..., pattern=r"^\d{10,15}$")

class SaleInvoiceIn(BaseModel):
    invoice_no: str = ""
    invoice_value: float = Field(0, ge=0)

class BulkPaymentIn(BaseModel):
    ids: List[int]
    status: str

class BulkLRIn(BaseModel):
    ids: List[int]
    lr_no: str = ""

class LRTrackingIn(BaseModel):
    lr_no: str = ""
    tracking_url: str = ""

class SettingsUpdateIn(BaseModel):
    settings: dict


# ---- PRODUCTS ----
SIZE_ORDER = {
    "10x10": 1, "250x250": 1,
    "12x12": 2, "300x300": 2,
    "15x15": 3, "380x380": 3,
    "18x18": 4, "450x450": 4,
    "21x21": 5, "535x535": 5,
    "24x24": 6, "600x600": 6,
    "26x26": 7, "660x660": 7,
    "28x28": 8, "710x710": 8,
    "30x30": 9, "760x760": 9,
    "36x36": 10, "900x900": 10,
    "42x42": 11, "1065x1065": 11,
    "12x18": 12, "300x450": 12,
    "12x24": 13, "300x600": 13,
    "18x24": 14, "450x600": 14,
}
COLOR_ORDER = {"Grey": 0, "White": 1}
CATEGORY_ORDER = {"Manhole Cover": 0, "Gully Cover": 1}

CSV_ORDER = {pn: i for i, (pn, _) in enumerate(PART_NO_CSV)}

# Discount Scheme (Aug 1 - Oct 31, 2026)
DISCOUNT_SCHEME = {
    "base_discount": 54,
    "slabs": [
        {"min": 50100, "max": 75000, "additional": 2.50},
        {"min": 75100, "max": 100000, "additional": 5.00},
        {"min": 100001, "max": 200000, "additional": 7.00},
        {"min": 200001, "max": float('inf'), "additional": 9.00},
    ]
}

def calculate_discount_scheme(basic_value):
    """Calculate discount based on the slab scheme. Returns (total_discount_percent, additional_percent, slab_info)"""
    if basic_value < 50100:
        return (0, 0, None)
    
    base = DISCOUNT_SCHEME["base_discount"]
    for slab in DISCOUNT_SCHEME["slabs"]:
        if slab["min"] <= basic_value <= slab["max"]:
            total = base + slab["additional"]
            return (total, slab["additional"], f"₹{slab['min']:,} to ₹{slab['max']:,}" if slab["max"] != float('inf') else f"₹{slab['min']:,} & Above")
    
    return (0, 0, None)

@app.get("/api/discount-scheme")
def get_discount_scheme(user: User = Depends(get_current_user)):
    return DISCOUNT_SCHEME

@app.get("/api/discount-calculate/{basic_value}")
def get_discount_calculate(basic_value: float, user: User = Depends(get_current_user)):
    total_pct, additional_pct, slab_info = calculate_discount_scheme(basic_value)
    discount_amount = basic_value * total_pct / 100 if total_pct > 0 else 0
    return {
        "basic_value": basic_value,
        "total_discount_percent": total_pct,
        "additional_percent": additional_pct,
        "slab_info": slab_info,
        "discount_amount": round(discount_amount, 2),
        "final_value": round(basic_value - discount_amount, 2)
    }

@app.get("/api/products")
def list_products(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Product).all()
        out = []
        for p in rows:
            mrp = p.pricing.mrp if p.pricing else 0
            out.append({
                "id": p.id, "part_no": p.part_no, "name": p.name, "category": p.category,
                "size": p.size, "load_rating": p.load_rating,
                "material": p.material, "color": p.color, "unit": p.unit, "hsn_code": p.hsn_code,
                "mrp": mrp
            })
        out.sort(key=lambda p: CSV_ORDER.get(p["part_no"], 999))
        return out
    finally:
        db.close()


@app.post("/api/products")
def create_product(inp: ProductIn, user: User = Depends(require_permission("products", "create"))):
    db = SessionLocal()
    try:
        p = Product(**inp.dict())
        db.add(p)
        db.commit()
        db.refresh(p)
        db.add(Pricing(product_id=p.id))
        db.commit()
        return {"id": p.id, "message": "Product created"}
    finally:
        db.close()


@app.put("/api/products/{pid}")
def update_product(pid: int, inp: ProductIn, user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        p = db.query(Product).filter(Product.id == pid).first()
        if not p:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(p, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/products/{pid}")
def delete_product(pid: int, user: User = Depends(require_permission("products", "delete"))):
    db = SessionLocal()
    try:
        p = db.query(Product).filter(Product.id == pid).first()
        if not p:
            raise HTTPException(404, "Not found")
        db.delete(p)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- PRICING ----
@app.get("/api/products/{pid}/pricing")
def get_pricing(pid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        pr = db.query(Pricing).filter(Pricing.product_id == pid).first()
        if not pr:
            raise HTTPException(404, "Not found")
        return {
            "raw_material_cost": pr.raw_material_cost,
            "labor_cost": pr.labor_cost,
            "overhead_cost": pr.overhead_cost,
            "packing_cost": pr.packing_cost,
            "total_cost": pr.total_cost,
            "profit_margin": pr.profit_margin,
            "mrp": pr.mrp,
            "dealer_price": pr.dealer_price,
            "distributor_price": pr.distributor_price,
            "gst_rate": pr.gst_rate
        }
    finally:
        db.close()


@app.put("/api/products/{pid}/pricing")
def update_pricing(pid: int, inp: PricingIn, user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        pr = db.query(Pricing).filter(Pricing.product_id == pid).first()
        if not pr:
            pr = Pricing(product_id=pid)
            db.add(pr)
        pr.raw_material_cost = inp.raw_material_cost
        pr.mrp = inp.mrp
        pr.gst_rate = inp.gst_rate
        pr.labor_cost = inp.labor_cost
        pr.overhead_cost = inp.overhead_cost
        pr.packing_cost = inp.packing_cost
        pr.profit_margin = inp.profit_margin
        total = inp.raw_material_cost + inp.labor_cost + inp.overhead_cost + inp.packing_cost
        pr.total_cost = total
        db.commit()
        return {"message": "Updated", "mrp": inp.mrp}
    finally:
        db.close()


# ---- ORDERS ----
@app.get("/api/orders")
def list_orders(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Order).order_by(Order.id).all()
        return [{"id": o.id, "sl_no": o.sl_no, "po_no": o.po_no, "po_date": o.po_date,
                 "customer_name": o.customer_name or "",
                 "billing_site": o.billing_site, "shipping_site": o.shipping_site,
                 "no_of_boxes": o.no_of_boxes, "value_excl_gst_freight": o.value_excl_gst_freight,
                 "invoice_no": o.invoice_no, "invoice_date": o.invoice_date,
                 "invoice_amount_excl_gst": o.invoice_amount_excl_gst,
                 "weight_kgs": o.weight_kgs, "freight_rate_per_kg": o.freight_rate_per_kg,
                 "transport_charges": o.transport_charges, "invoice_amount": o.invoice_amount,
                 "eway_bill_no": o.eway_bill_no, "lr_no": o.lr_no, "entry_date": o.entry_date,
                 "credit_note_amount": o.credit_note_amount, "credit_note_no": o.credit_note_no,
                 "transporter": o.transporter, "transporter_no": o.transporter_no}
                for o in rows]
    finally:
        db.close()


@app.get("/api/orders/{oid}")
def get_order(oid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        o = db.query(Order).filter(Order.id == oid).first()
        if not o:
            raise HTTPException(404, "Order not found")
        return {"id": o.id, "sl_no": o.sl_no, "po_no": o.po_no, "po_date": o.po_date,
                 "customer_name": o.customer_name or "",
                 "billing_site": o.billing_site, "shipping_site": o.shipping_site,
                 "no_of_boxes": o.no_of_boxes, "value_excl_gst_freight": o.value_excl_gst_freight,
                 "invoice_no": o.invoice_no, "invoice_date": o.invoice_date,
                 "invoice_amount_excl_gst": o.invoice_amount_excl_gst,
                 "weight_kgs": o.weight_kgs, "freight_rate_per_kg": o.freight_rate_per_kg,
                 "transport_charges": o.transport_charges, "invoice_amount": o.invoice_amount,
                 "eway_bill_no": o.eway_bill_no, "lr_no": o.lr_no, "entry_date": o.entry_date,
                 "credit_note_amount": o.credit_note_amount, "credit_note_no": o.credit_note_no,
                 "transporter": o.transporter, "transporter_no": o.transporter_no}
    finally:
        db.close()


@app.post("/api/orders")
def create_order(inp: OrderIn, user: User = Depends(require_permission("orders", "create"))):
    db = SessionLocal()
    try:
        max_sl = db.query(func.max(Order.sl_no)).scalar()
        next_sl = (max_sl + 1) if max_sl else 1
        data = inp.dict()
        data["sl_no"] = next_sl
        o = Order(**data)
        db.add(o)
        db.commit()
        db.refresh(o)
        return {"id": o.id, "sl_no": next_sl, "message": "Order created"}
    finally:
        db.close()


@app.put("/api/orders/{oid}")
def update_order(oid: int, inp: OrderIn, user: User = Depends(require_permission("orders", "edit"))):
    db = SessionLocal()
    try:
        o = db.query(Order).filter(Order.id == oid).first()
        if not o:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(o, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/orders/{oid}")
def delete_order(oid: int, user: User = Depends(require_permission("orders", "delete"))):
    db = SessionLocal()
    try:
        o = db.query(Order).filter(Order.id == oid).first()
        if not o:
            raise HTTPException(404, "Not found")
        db.delete(o)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- PROFORMA ORDERS (Multi-Product PI/PO) ----
@app.get("/api/proforma-orders")
def list_proforma_orders(order_type: str = None, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        query = db.query(ProformaOrder)
        if order_type:
            query = query.filter(ProformaOrder.order_type == order_type)
        rows = query.order_by(ProformaOrder.created_at.desc()).all()
        # Batch-load customers to avoid N+1
        cust_ids = list(set(o.customer_id for o in rows if o.customer_id))
        cust_map = {}
        if cust_ids:
            custs = db.query(Customer).filter(Customer.id.in_(cust_ids)).all()
            cust_map = {c.id: c.contact_name for c in custs}
        # Batch-load all items to avoid N+1
        order_ids = [o.id for o in rows]
        all_items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id.in_(order_ids)).all()
        items_count_map = {}
        for item in all_items:
            items_count_map[item.proforma_order_id] = items_count_map.get(item.proforma_order_id, 0) + 1
        out = []
        for o in rows:
            out.append({
                "id": o.id, "pi_no": o.pi_no,
                "pi_date": o.pi_date.isoformat() if o.pi_date else None,
                "customer_name": cust_map.get(o.customer_id, "?"),
                "customer_id": o.customer_id,
                "billing_site": o.billing_site, "shipping_site": o.shipping_site,
                "no_of_boxes": o.no_of_boxes, "total_qty": o.total_qty,
                "value_excl_gst": o.value_excl_gst, "gst_amount": o.gst_amount,
                "total_amount": o.total_amount, "freight_amount": o.freight_amount,
                "payment_status": o.payment_status, "payment_method": o.payment_method,
                "transport_mode": o.transport_mode, "delivery_days": o.delivery_days,
                "notes": o.notes, "terms": o.terms, "order_type": o.order_type,
                "status": o.status or "draft",
                "discount_scheme_applied": o.discount_scheme_applied or 0,
                "discount_percent": o.discount_percent or 0,
                "discount_amount": o.discount_amount or 0,
                "item_count": items_count_map.get(o.id, 0),
                "created_at": o.created_at.isoformat() if o.created_at else None
            })
        return out
    finally:
        db.close()


@app.get("/api/proforma-orders/{oid}")
def get_proforma_order(oid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        o = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not o:
            raise HTTPException(404, "Order not found")
        cust = db.query(Customer).filter(Customer.id == o.customer_id).first()
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == o.id).order_by(ProformaOrderItem.sl_no).all()
        items_out = []
        for item in items:
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            items_out.append({
                "id": item.id, "sl_no": item.sl_no,
                "product_id": item.product_id,
                "product_name": prod.name if prod else "?",
                "part_no": item.part_no, "description": item.description,
                "size": item.size, "category": item.category,
                "qty_boxes": item.qty_boxes, "std_packaging": item.std_packaging,
                "pieces_per_box": item.pieces_per_box, "final_qty": item.final_qty,
                "mrp": item.mrp, "d1": item.d1, "d2": item.d2, "d3": item.d3,
                "d4": item.d4, "d5": item.d5, "cd": item.cd,
                "discount_percent": item.discount_percent,
                "net_rate": item.net_rate, "lock_hinge": item.lock_hinge,
                "basic_amount": item.basic_amount
            })
        return {
            "id": o.id, "pi_no": o.pi_no,
            "pi_date": o.pi_date.isoformat() if o.pi_date else None,
            "customer_id": o.customer_id,
            "customer_name": cust.contact_name if cust else "?",
            "customer_gst": cust.gstin if cust else "",
            "customer_state": cust.state if cust else "",
            "customer_address": cust.billing_address if cust else "",
            "billing_site": o.billing_site, "shipping_site": o.shipping_site,
            "no_of_boxes": o.no_of_boxes, "total_qty": o.total_qty,
            "value_excl_gst": o.value_excl_gst, "gst_amount": o.gst_amount,
            "total_amount": o.total_amount, "freight_amount": o.freight_amount,
            "payment_status": o.payment_status, "payment_method": o.payment_method,
            "transport_mode": o.transport_mode, "delivery_days": o.delivery_days,
            "notes": o.notes, "terms": o.terms, "order_type": o.order_type,
            "status": o.status or "draft",
            "discount_scheme_applied": o.discount_scheme_applied or 0,
            "discount_percent": o.discount_percent or 0,
            "discount_amount": o.discount_amount or 0,
            "items": items_out,
            "created_at": o.created_at.isoformat() if o.created_at else None
        }
    finally:
        db.close()


@app.post("/api/proforma-orders")
def create_proforma_order(inp: ProformaOrderIn, user: User = Depends(require_permission("proforma_orders", "create"))):
    db = SessionLocal()
    try:
        customer = db.query(Customer).filter(Customer.id == inp.customer_id).first()
        if not customer:
            raise HTTPException(404, "Customer not found")

        # Generate pi_no with retry to handle race conditions
        max_retries = 5
        for attempt in range(max_retries):
            try:
                max_id = db.query(func.max(ProformaOrder.id)).scalar() or 0
                pi_no = f"RFC/{datetime.now().strftime('%y%m')}-{max_id + 1:03d}"
                break
            except Exception:
                if attempt == max_retries - 1:
                    raise
                db.rollback()
                time.sleep(0.1)

        total_qty = 0
        total_basic = 0

        for item in inp.items:
            net = item.mrp
            for d in [item.d1, item.d2, item.d3, item.d4, item.d5, item.cd]:
                net = net * (1 - d / 100)
            item.net_rate = round(net, 2)
            item.basic_amount = item.final_qty * item.net_rate
            total_qty += item.final_qty
            total_basic += item.basic_amount

        gst_amount = total_basic * get_gst_rate() / 100
        
        # Calculate discount scheme
        discount_pct = 0
        discount_amount = 0
        if inp.discount_scheme_applied:
            discount_pct, additional_pct, slab_info = calculate_discount_scheme(total_basic)
            if discount_pct > 0:
                discount_amount = total_basic * discount_pct / 100
        
        final_basic = total_basic - discount_amount
        gst_amount = final_basic * get_gst_rate() / 100
        total_amount = final_basic + gst_amount + inp.freight_amount

        order = ProformaOrder(
            pi_no=pi_no, customer_id=inp.customer_id,
            billing_site=inp.billing_site, shipping_site=inp.shipping_site,
            total_qty=total_qty, no_of_boxes=sum(i.qty_boxes for i in inp.items),
            value_excl_gst=total_basic, gst_amount=gst_amount,
            total_amount=total_amount, freight_amount=inp.freight_amount,
            payment_status=inp.payment_status, payment_method=inp.payment_method,
            transport_mode=inp.transport_mode, delivery_days=inp.delivery_days,
            notes=inp.notes, terms=inp.terms, order_type=inp.order_type,
            discount_scheme_applied=inp.discount_scheme_applied,
            discount_percent=discount_pct, discount_amount=discount_amount
        )
        db.add(order)
        db.flush()

        for idx, item in enumerate(inp.items):
            db.add(ProformaOrderItem(
                proforma_order_id=order.id, sl_no=idx + 1,
                product_id=item.product_id, part_no=item.part_no,
                description=item.description, size=item.size,
                category=item.category, qty_boxes=item.qty_boxes,
                std_packaging=item.std_packaging, pieces_per_box=item.pieces_per_box,
                final_qty=item.final_qty, mrp=item.mrp,
                d1=item.d1, d2=item.d2, d3=item.d3, d4=item.d4, d5=item.d5,
                cd=item.cd, discount_percent=item.discount_percent,
                net_rate=item.net_rate, lock_hinge=item.lock_hinge,
                basic_amount=item.basic_amount
            ))

        db.commit()
        return {"id": order.id, "pi_no": pi_no, "total": total_amount}
    finally:
        db.close()


@app.put("/api/proforma-orders/{oid}")
def update_proforma_order(oid: int, inp: ProformaOrderIn, user: User = Depends(require_permission("proforma_orders", "edit"))):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")

        customer = db.query(Customer).filter(Customer.id == inp.customer_id).first()
        if not customer:
            raise HTTPException(404, "Customer not found")

        order.customer_id = inp.customer_id
        order.billing_site = inp.billing_site
        order.shipping_site = inp.shipping_site
        order.freight_amount = inp.freight_amount
        order.payment_status = inp.payment_status
        order.payment_method = inp.payment_method
        order.transport_mode = inp.transport_mode
        order.delivery_days = inp.delivery_days
        order.notes = inp.notes
        order.terms = inp.terms
        order.order_type = inp.order_type
        order.discount_scheme_applied = inp.discount_scheme_applied

        db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).delete()

        total_qty = 0
        total_basic = 0
        for idx, item in enumerate(inp.items):
            net = item.mrp
            for d in [item.d1, item.d2, item.d3, item.d4, item.d5, item.cd]:
                net = net * (1 - d / 100)
            item.net_rate = round(net, 2)
            item.basic_amount = item.final_qty * item.net_rate
            total_qty += item.final_qty
            total_basic += item.basic_amount
            db.add(ProformaOrderItem(
                proforma_order_id=oid, sl_no=idx + 1,
                product_id=item.product_id, part_no=item.part_no,
                description=item.description, size=item.size,
                category=item.category, qty_boxes=item.qty_boxes,
                std_packaging=item.std_packaging, pieces_per_box=item.pieces_per_box,
                final_qty=item.final_qty, mrp=item.mrp,
                d1=item.d1, d2=item.d2, d3=item.d3, d4=item.d4, d5=item.d5,
                cd=item.cd, discount_percent=item.discount_percent,
                net_rate=item.net_rate, lock_hinge=item.lock_hinge,
                basic_amount=item.basic_amount
            ))

        # Calculate discount scheme
        discount_pct = 0
        discount_amount = 0
        if inp.discount_scheme_applied:
            discount_pct, additional_pct, slab_info = calculate_discount_scheme(total_basic)
            if discount_pct > 0:
                discount_amount = total_basic * discount_pct / 100
        
        final_basic = total_basic - discount_amount
        gst_amount = final_basic * get_gst_rate() / 100
        order.total_qty = total_qty
        order.no_of_boxes = sum(i.qty_boxes for i in inp.items)
        order.value_excl_gst = total_basic
        order.discount_percent = discount_pct
        order.discount_amount = discount_amount
        order.gst_amount = gst_amount
        order.total_amount = final_basic + gst_amount + inp.freight_amount
        order.updated_at = datetime.now(timezone.utc)

        db.commit()
        return {"message": "Order updated"}
    finally:
        db.close()


@app.delete("/api/proforma-orders/{oid}")
def delete_proforma_order(oid: int, user: User = Depends(require_permission("proforma_orders", "delete"))):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).delete()
        db.delete(order)
        db.commit()
        return {"message": "Order deleted"}
    finally:
        db.close()


@app.get("/api/products/{pid}/details")
def get_product_details(pid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        p = db.query(Product).filter(Product.id == pid).first()
        if not p:
            raise HTTPException(404, "Product not found")
        pr = db.query(Pricing).filter(Pricing.product_id == pid).first()
        return {
            "id": p.id, "name": p.name, "category": p.category,
            "size": p.size, "part_no": p.part_no,
            "mrp": pr.mrp if pr else 0,
            "pieces_per_box": p.pieces_per_box or 1,
            "std_packaging": p.std_packaging or 1
        }
    finally:
        db.close()


@app.get("/api/proforma-orders/{oid}/pdf")
def generate_proforma_order_pdf(oid: int):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).order_by(ProformaOrderItem.sl_no).all()

        pi_date = order.pi_date.strftime("%d-%b-%Y") if order.pi_date else ""

        billing_site = None
        if order.billing_site:
            try:
                billing_site = db.query(BillingSite).filter(BillingSite.id == int(order.billing_site)).first()
            except (ValueError, TypeError):
                logger.warning("Invalid billing_site ID '%s' for order %s", order.billing_site, order.id)

        if order.order_type == "PO":
            html = _generate_po_html(order, customer, items, pi_date, billing_site)
        else:
            html = _generate_pi_html(order, customer, items, pi_date, billing_site)

        return HTMLResponse(content=html)
    finally:
        db.close()


def _billing_site_header(bs=None):
    name = (bs.name if bs else "Raksha Pipes Private Limited").replace("Private Limited", "Pvt. Ltd.") if bs else "Raksha Pipes Pvt. Ltd."
    address = bs.address if bs else ""
    phone = bs.phone if bs else ""
    email = bs.email if bs else ""
    website = bs.website if bs else "www.rakshapipes.com"
    gstin = bs.gstin if bs else ""
    state_code = bs.state_code if bs else ""
    pan = bs.pan if bs else ""
    return f"""
<div style="text-align:center;border-bottom:3px double #000;padding-bottom:10px;margin-bottom:10px;">
<h1 style="margin:0;font-size:22px;font-weight:bold;letter-spacing:1px;">{escape_html(name)}</h1>
<p style="margin:2px 0;font-size:11px;">{escape_html(address)}</p>
<table style="width:100%;font-size:10px;margin-top:6px;"><tr>
<td style="text-align:left;">Phone: +91 - {escape_html(phone)}</td>
<td style="text-align:center;">Email: {escape_html(email)}</td>
<td style="text-align:right;">Website: {escape_html(website)}</td>
</tr><tr>
<td style="text-align:left;">State Code: {escape_html(state_code)}</td>
<td style="text-align:center;">GSTIN: {escape_html(gstin)}</td>
<td style="text-align:right;">PAN No: {escape_html(pan)}</td>
</tr></table>
</div>
"""

COMPANY_BANK_DETAILS = """
<div style="margin-top:15px;font-size:11px;">
<h4 style="margin:0 0 6px 0;font-size:13px;">Bank Details</h4>
<table style="font-size:11px;">
<tr><td style="font-weight:bold;padding-right:10px;">Name:</td><td>Raksha Pipes Pvt. Ltd.</td></tr>
<tr><td style="font-weight:bold;padding-right:10px;">Account Number:</td><td>004705011678</td></tr>
<tr><td style="font-weight:bold;padding-right:10px;">Bank Name:</td><td>ICICI Bank Ltd.</td></tr>
<tr><td style="font-weight:bold;padding-right:10px;">Branch Name:</td><td>Koramangala, Bengaluru</td></tr>
<tr><td style="font-weight:bold;padding-right:10px;">IFSC Code:</td><td>ICIC0000047</td></tr>
</table>
</div>
"""

COMPANY_TERMS = ""

def _generate_po_html(order, customer, items, pi_date, billing_site=None):
    cust_name = customer.contact_name if customer else (order.billing_site or "")
    cust_gstin = customer.gstin if customer else ""
    cust_state = customer.state if customer else ""

    bs_name = (billing_site.name if billing_site else "Raksha Pipes Private Limited").replace("Private Limited", "Pvt. Ltd.") if billing_site else "Raksha Pipes Pvt. Ltd."
    bs_address = billing_site.address if billing_site else ""
    bs_phone = billing_site.phone if billing_site else ""
    bs_email = billing_site.email if billing_site else ""
    bs_website = billing_site.website if billing_site else "www.rakshapipes.com"
    bs_gstin = billing_site.gstin if billing_site else ""
    bs_state_code = billing_site.state_code if billing_site else ""
    bs_pan = billing_site.pan if billing_site else ""

    items_html = ""
    total_box = 0
    total_pcs = 0
    total_amount = 0
    for item in items:
        box = item.qty_boxes or 0
        pcs = item.final_qty or 0
        amt = item.basic_amount or 0
        total_box += box
        total_pcs += pcs
        total_amount += amt
        items_html += f"""
        <tr>
            <td style="padding:5px 8px;border:1px solid #ccc;font-size:10px;">{item.part_no or ''}</td>
            <td style="padding:5px 8px;border:1px solid #ccc;font-size:10px;">{escape_html(item.description or '')} {escape_html(item.size or '')}</td>
            <td style="padding:5px 8px;border:1px solid #ccc;text-align:center;font-size:10px;">Box</td>
            <td style="padding:5px 8px;border:1px solid #ccc;text-align:center;font-size:10px;">{item.std_packaging or 0}</td>
            <td style="padding:5px 8px;border:1px solid #ccc;text-align:center;font-size:10px;">{pcs}</td>
            <td style="padding:5px 8px;border:1px solid #ccc;text-align:right;font-size:10px;">&#8377;{item.mrp:,.0f}</td>
            <td style="padding:5px 8px;border:1px solid #ccc;text-align:right;font-size:10px;">&#8377;{amt:,.0f}</td>
        </tr>"""

    gst_amount = total_amount * get_gst_rate() / 100
    grand_total = total_amount + gst_amount

    # Calculate discount scheme
    discount_html = ""
    discount_amount = 0
    if order.discount_scheme_applied and total_amount >= 50100:
        slabs = [
            (50100, 75000, 2.50),
            (75100, 100000, 5.00),
            (100001, 200000, 7.00),
            (200001, float('inf'), 9.00),
        ]
        additional_discount = 0
        for smin, smax, add in slabs:
            if smin <= total_amount <= smax:
                additional_discount = add
                break
        total_discount_pct = 54 + additional_discount
        discount_amount = total_amount * total_discount_pct / 100
        after_discount = total_amount - discount_amount
        gst_amount = after_discount * get_gst_rate() / 100
        grand_total = after_discount + gst_amount
        discount_html = f"""
        <tr style="font-weight:bold;color:#059669;">
            <td colspan="6" style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">Discount Scheme ({total_discount_pct}%)</td>
            <td style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">-&#8377;{discount_amount:,.0f}</td>
        </tr>
        <tr style="font-weight:bold;">
            <td colspan="6" style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">After Discount</td>
            <td style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">&#8377;{after_discount:,.0f}</td>
        </tr>"""

    html = f"""<!DOCTYPE html><html><head><title>Purchase Order</title>
<style>
body{{font-family:Arial,sans-serif;margin:15px;font-size:11px;color:#000;}}
table{{width:100%;border-collapse:collapse;}}
@page{{size:A4;margin:10mm;}}
@media print{{body{{margin:5mm;}}}}
</style></head><body>

{_billing_site_header(billing_site)}

<table style="margin-bottom:8px;font-size:11px;width:100%;">
<tr>
<td style="width:50%;vertical-align:top;">
<b>VENDOR / Supplier Details</b><br>
<table style="font-size:10px;margin-top:4px;">
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">Company Name:</td><td>{escape_html(cust_name)}</td></tr>
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">Address:</td><td>{escape_html(order.billing_site or '')}</td></tr>
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">State Code:</td><td>{escape_html(cust_state)}</td></tr>
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">GSTIN:</td><td>{escape_html(cust_gstin)}</td></tr>
</table>
</td>
<td style="width:50%;vertical-align:top;">
<table style="font-size:10px;float:right;">
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">Contact Person:</td><td>{escape_html(customer.contact_name if customer else '')}</td></tr>
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">Mobile No:</td><td>{escape_html(customer.contact_number if customer else '')}</td></tr>
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">Email:</td><td>{escape_html(customer.contact_email if customer else '')}</td></tr>
<tr><td style="padding:1px 8px 1px 0;font-weight:bold;">Payment Terms:</td><td>100% Advance</td></tr>
</table>
</td>
</tr>
</table>

<table style="width:100%;margin-bottom:6px;font-size:11px;">
<tr>
<td style="width:20%;font-weight:bold;">Anand No.:</td>
<td style="width:15%;"></td>
<td style="width:30%;text-align:center;font-size:16px;font-weight:bold;border-top:2px solid #000;border-bottom:2px solid #000;padding:4px;">PURCHASE ORDER</td>
<td style="width:15%;text-align:right;font-weight:bold;">PO No:</td>
<td style="width:20%;text-align:right;">{order.pi_no or ''}</td>
</tr>
<tr>
<td style="font-weight:bold;">PO Date:</td>
<td>{pi_date}</td>
<td></td>
<td style="text-align:right;font-weight:bold;">Amd No.:</td>
<td></td>
</tr>
</table>

<table style="width:100%;border:1px solid #ccc;margin-top:8px;">
<thead><tr style="background:#f0f0f0;">
<th style="padding:6px;border:1px solid #ccc;text-align:left;font-size:10px;width:18%;">Part No</th>
<th style="padding:6px;border:1px solid #ccc;text-align:left;font-size:10px;width:30%;">Description</th>
<th style="padding:6px;border:1px solid #ccc;text-align:center;font-size:10px;width:8%;">Box</th>
<th style="padding:6px;border:1px solid #ccc;text-align:center;font-size:10px;width:10%;">Pcs</th>
<th style="padding:6px;border:1px solid #ccc;text-align:center;font-size:10px;width:10%;">Total Pcs</th>
<th style="padding:6px;border:1px solid #ccc;text-align:right;font-size:10px;width:12%;">Rate</th>
<th style="padding:6px;border:1px solid #ccc;text-align:right;font-size:10px;width:12%;">Amount</th>
</tr></thead>
<tbody>{items_html}</tbody>
<tfoot>
<tr style="font-weight:bold;background:#f9f9f9;">
<td colspan="2" style="padding:6px 8px;border:1px solid #ccc;font-size:11px;">Total</td>
<td style="padding:6px 8px;border:1px solid #ccc;text-align:center;font-size:11px;">{total_box}</td>
<td style="padding:6px 8px;border:1px solid #ccc;text-align:center;font-size:11px;">{total_pcs}</td>
<td style="padding:6px 8px;border:1px solid #ccc;text-align:center;font-size:11px;">{total_pcs}</td>
<td style="padding:6px 8px;border:1px solid #ccc;"></td>
<td style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">&#8377;{total_amount:,.0f}</td>
</tr>
{discount_html}
<tr style="font-weight:bold;">
<td colspan="6" style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">GST  18%</td>
<td style="padding:6px 8px;border:1px solid #ccc;text-align:right;font-size:11px;">&#8377;{gst_amount:,.0f}</td>
</tr>
<tr style="font-weight:bold;background:#f0f0f0;">
<td colspan="6" style="padding:8px;border:1px solid #ccc;text-align:right;font-size:13px;">GRAND TOTAL</td>
<td style="padding:8px;border:1px solid #ccc;text-align:right;font-size:13px;">&#8377;{grand_total:,.0f}</td>
</tr>
</tfoot></table>

<table style="width:100%;margin-top:40px;font-size:11px;border:none;">
<tr>
<td style="width:33%;text-align:center;border-top:1px solid #000;padding-top:8px;">CREATED BY</td>
<td style="width:33%;text-align:center;border-top:1px solid #000;padding-top:8px;">REVIEWED BY</td>
<td style="width:33%;text-align:center;border-top:1px solid #000;padding-top:8px;">APPROVED BY</td>
</tr>
</table>

</body></html>"""
    return html


def _generate_pi_html(order, customer, items, pi_date, billing_site=None):
    cust_name = customer.contact_name if customer else (order.billing_site or "")
    cust_gstin = customer.gstin if customer else ""
    cust_state = customer.state if customer else ""
    cust_id = customer.customer_id if customer else ""

    items_html = ""
    total_box = 0
    total_pcs = 0
    total_basic = 0
    total_lock_hinge = 0
    for item in items:
        box = item.qty_boxes or 0
        pcs = item.final_qty or 0
        mrp = item.mrp or 0
        d1 = item.d1 or 0
        d2 = item.d2 or 0
        d3 = item.d3 or 0
        d4 = item.d4 or 0
        d5 = item.d5 or 0
        cd = item.cd or 0
        lock = item.lock_hinge or 0
        net = item.net_rate or 0
        amt = item.basic_amount or 0
        total_box += box
        total_pcs += pcs
        total_basic += amt
        total_lock_hinge += lock

        # Calculate discount chain
        base = mrp
        after_d1 = base - (base * d1 / 100) if d1 else base
        after_d2 = after_d1 - (after_d1 * d2 / 100) if d2 else after_d1
        after_d3 = after_d2 - (after_d2 * d3 / 100) if d3 else after_d2
        after_d4 = after_d3 - (after_d3 * d4 / 100) if d4 else after_d3
        after_d5 = after_d4 - (after_d4 * d5 / 100) if d5 else after_d4
        after_cd = after_d5 - (after_d5 * cd / 100) if cd else after_d5

        items_html += f"""
        <tr>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{item.sl_no}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;font-size:9px;">{escape_html(item.description or '')} ({escape_html(item.size or '')})</td>
            <td style="padding:4px 6px;border:1px solid #ccc;font-size:9px;">{escape_html(item.category or '')}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;font-size:9px;">{escape_html(item.part_no or '')}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">FRP</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">Box</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{item.std_packaging or 0}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{pcs}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">Pieces</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{mrp:,.0f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{d1}%</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{after_d1:,.2f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{d2}%</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{after_d2:,.2f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{d3}%</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{after_d3:,.2f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{d4}%</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{after_d4:,.2f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{d5}%</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{after_d5:,.2f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:center;font-size:9px;">{cd}%</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;font-weight:bold;">{after_cd:,.2f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;">{lock:,.0f}</td>
            <td style="padding:4px 6px;border:1px solid #ccc;text-align:right;font-size:9px;font-weight:bold;">{amt:,.2f}</td>
        </tr>"""

    packing_charges = 0
    sub_total = total_basic + packing_charges
    
    # Calculate discount scheme
    discount_html_row = ""
    discount_amount = 0
    if order.discount_scheme_applied and sub_total >= 50100:
        slabs = [
            (50100, 75000, 2.50),
            (75100, 100000, 5.00),
            (100001, 200000, 7.00),
            (200001, float('inf'), 9.00),
        ]
        additional_discount = 0
        for smin, smax, add in slabs:
            if smin <= sub_total <= smax:
                additional_discount = add
                break
        total_discount_pct = 54 + additional_discount
        discount_amount = sub_total * total_discount_pct / 100
        sub_total = sub_total - discount_amount
        discount_html_row = f'<tr><td style="padding:2px 8px;font-weight:bold;color:#059669;">DISCOUNT SCHEME ({total_discount_pct}%)</td><td style="text-align:right;padding:2px 8px;color:#059669;">-&#8377;{discount_amount:,.2f}</td></tr>'
    
    gst = sub_total * get_gst_rate() / 100
    total_value = sub_total + gst
    tcs_rate = 0.001  # 0.1%
    tcs_amount = total_value * tcs_rate
    final_value = total_value + tcs_amount

    html = f"""<!DOCTYPE html><html><head><title>Quotation cum Proforma Invoice</title>
<style>
body{{font-family:Arial,sans-serif;margin:10px;font-size:10px;color:#000;}}
table{{width:100%;border-collapse:collapse;}}
@page{{size:landscape A4;margin:8mm;}}
@media print{{body{{margin:5mm;}}}}
</style></head><body>

{_billing_site_header(billing_site)}

<table style="width:100%;font-size:10px;margin-bottom:8px;">
<tr>
<td style="width:50%;vertical-align:top;">
<table style="font-size:10px;">
<tr><td style="padding:1px 6px;font-weight:bold;width:120px;">CONSIGNEE ERP CODE</td><td>{escape_html(cust_id)}</td>
<td style="padding:1px 6px;font-weight:bold;width:120px;">KYC STATUS</td><td></td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">CONSIGNEE NAME</td><td>{escape_html(cust_name)}</td>
<td style="padding:1px 6px;font-weight:bold;">NAME OF SALE EXECUTIVE</td><td>{escape_html(customer.exec_name if customer else '')}</td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">LOCATION</td><td>{escape_html(customer.city if customer else '')}</td>
<td style="padding:1px 6px;font-weight:bold;">LOCATION</td><td>{escape_html(customer.city if customer else '')}</td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">STATE</td><td>{escape_html(cust_state)}</td>
<td style="padding:1px 6px;font-weight:bold;">STATE</td><td>{escape_html(cust_state)}</td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">PURCHASE ORDER NO</td><td>{escape_html(order.pi_no or '')}</td>
<td style="padding:1px 6px;font-weight:bold;">BILL TO ADDRESS CODE</td><td></td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">PURCHASE ORDER DATE</td><td>{pi_date}</td>
<td style="padding:1px 6px;font-weight:bold;">SHIP TO ADDRESS CODE</td><td></td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">PLANNING DATE</td><td>{pi_date}</td>
<td style="padding:1px 6px;font-weight:bold;">TOTAL WEIGHT</td><td></td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">DESPATCH DATE</td><td>{pi_date}</td>
<td style="padding:1px 6px;font-weight:bold;">TOTAL VALUE</td><td>&#8377;{total_basic:,.2f}</td></tr>
<tr><td style="padding:1px 6px;font-weight:bold;">STATE SR NO</td><td></td>
<td style="padding:1px 6px;font-weight:bold;">TRADE DISCOUNT</td><td></td></tr>
</table>
</td>
</tr>
</table>

<div style="overflow-x:auto;">
<table style="width:100%;border:1px solid #ccc;font-size:9px;">
<thead><tr style="background:#1a365d;color:white;">
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">SN</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Product Specification</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Size in (Inch &amp; MM)</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">ERP Part No</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Item Grp.</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Base UOM</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Std Packing</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Final Qty</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Final UOM</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Gen</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">D-1</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Net</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">D-2</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Net Rt</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">D-3</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Net Rt</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">D-4</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Net Rt</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">D-5</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Net Rt</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">CD</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Net Rt</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Lock &amp; Hings</th>
<th style="padding:4px;border:1px solid #ccc;font-size:8px;">Basic Amt Without GST</th>
</tr></thead>
<tbody>{items_html}</tbody>
<tfoot>
<tr style="font-weight:bold;background:#f0f0f0;">
<td colspan="5" style="padding:5px 8px;border:1px solid #ccc;font-size:10px;">TOTAL</td>
<td style="padding:5px 8px;border:1px solid #ccc;text-align:center;">Box</td>
<td style="padding:5px 8px;border:1px solid #ccc;text-align:center;">{total_box}</td>
<td style="padding:5px 8px;border:1px solid #ccc;text-align:center;">{total_pcs}</td>
<td style="padding:5px 8px;border:1px solid #ccc;text-align:center;">Pieces</td>
<td colspan="11" style="padding:5px 8px;border:1px solid #ccc;"></td>
<td style="padding:5px 8px;border:1px solid #ccc;text-align:right;">{total_lock_hinge:,.0f}</td>
<td style="padding:5px 8px;border:1px solid #ccc;text-align:right;">{total_basic:,.2f}</td>
</tr>
</tfoot></table>
</div>

<table style="width:100%;margin-top:10px;font-size:11px;">
<tr>
<td style="width:50%;vertical-align:top;">
{COMPANY_BANK_DETAILS}
</td>
<td style="width:50%;vertical-align:top;text-align:right;">
<table style="float:right;font-size:11px;">
<tr><td style="padding:2px 8px;font-weight:bold;">BASIC VALUE</td><td style="text-align:right;padding:2px 8px;">&#8377;{total_basic:,.2f}</td></tr>
<tr><td style="padding:2px 8px;font-weight:bold;">ADD PACKING &amp; FORWARDING CHARGES</td><td style="text-align:right;padding:2px 8px;">&#8377;{packing_charges:,.2f}</td></tr>
{discount_html_row}
<tr><td style="padding:2px 8px;font-weight:bold;">SUB TOTAL</td><td style="text-align:right;padding:2px 8px;">&#8377;{sub_total:,.2f}</td></tr>
<tr><td style="padding:2px 8px;font-weight:bold;">GST @ 18.00%</td><td style="text-align:right;padding:2px 8px;">&#8377;{gst:,.2f}</td></tr>
<tr><td style="padding:2px 8px;font-weight:bold;">TOTAL VALUE</td><td style="text-align:right;padding:2px 8px;">&#8377;{total_value:,.2f}</td></tr>
<tr><td style="padding:2px 8px;font-weight:bold;">TCS On Sales 0.1%</td><td style="text-align:right;padding:2px 8px;">&#8377;{tcs_amount:,.2f}</td></tr>
<tr><td style="padding:2px 8px;font-weight:bold;">ROUND OFF DIFF</td><td style="text-align:right;padding:2px 8px;">&#8377;0.00</td></tr>
<tr style="font-size:14px;font-weight:bold;border-top:2px solid #000;">
<td style="padding:6px 8px;">FINAL PI VALUE</td>
<td style="text-align:right;padding:6px 8px;">&#8377;{final_value:,.2f}</td></tr>
</table>
</td>
</tr>
</table>

{COMPANY_TERMS}

<div style="margin-top:40px;text-align:right;font-size:11px;">
<p>For <b>{escape_html((billing_site.name if billing_site else "Raksha Pipes Private Limited").replace("Private Limited", "Pvt. Ltd."))}</b></p>
<p style="margin-top:30px;">Authorized Signatory</p>
</div>

</body></html>"""
    return html


# ---- CUSTOMERS ----
@app.get("/api/customers")
def list_customers(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Customer).all()
        return [{"id": c.id, "customer_id": c.customer_id, "name": c.name or "", "gstin": c.gstin,
                 "billing_address": c.billing_address, "shipping_address": c.shipping_address,
                 "state": c.state, "district": c.district, "city": c.city, "pincode": c.pincode,
                 "contact_name": c.contact_name, "contact_number": c.contact_number, "contact_email": c.contact_email,
                 "exec_code": c.exec_code, "exec_name": c.exec_name, "exec_number": c.exec_number, "exec_email": c.exec_email,
                 "blacklisted": c.blacklisted}
                for c in rows]
    finally:
        db.close()


@app.post("/api/customers")
def create_customer(inp: CustomerIn, user: User = Depends(require_permission("customers", "create"))):
    db = SessionLocal()
    try:
        c = Customer(**inp.dict())
        db.add(c)
        db.commit()
        db.refresh(c)
        return {"id": c.id, "message": "Customer created"}
    finally:
        db.close()


@app.put("/api/customers/{cid}")
def update_customer(cid: int, inp: CustomerIn, user: User = Depends(require_permission("customers", "edit"))):
    db = SessionLocal()
    try:
        c = db.query(Customer).filter(Customer.id == cid).first()
        if not c:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(c, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/customers/{cid}")
def delete_customer(cid: int, user: User = Depends(require_permission("customers", "delete"))):
    db = SessionLocal()
    try:
        c = db.query(Customer).filter(Customer.id == cid).first()
        if not c:
            raise HTTPException(404, "Not found")
        db.query(Sale).filter(Sale.customer_id == c.id).update({"customer_id": None})
        db.delete(c)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


# ---- TRANSPORTERS ----
@app.get("/api/transporters")
def list_transporters(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Transporter).all()
        return [{"id": t.id, "transporter_id": t.transporter_id, "name": t.name,
                 "phone": t.phone, "email": t.email, "address": t.address,
                 "state": t.state, "district": t.district, "city": t.city, "pincode": t.pincode,
                 "gst_number": t.gst_number, "pan_number": t.pan_number,
                 "gst_certificate": t.gst_certificate, "pan_card": t.pan_card,
                 "contact_person": t.contact_person, "contact_number": t.contact_number,
                 "tracking_url_pattern": t.tracking_url_pattern or "",
                 "blacklisted": t.blacklisted}
                for t in rows]
    finally:
        db.close()


@app.post("/api/transporters")
def create_transporter(inp: TransporterIn, user: User = Depends(require_permission("transporters", "create"))):
    db = SessionLocal()
    try:
        t = Transporter(**inp.dict())
        db.add(t)
        db.commit()
        db.refresh(t)
        return {"id": t.id, "message": "Transporter created"}
    finally:
        db.close()


@app.put("/api/transporters/{tid}")
def update_transporter(tid: int, inp: TransporterIn, user: User = Depends(require_permission("transporters", "edit"))):
    db = SessionLocal()
    try:
        t = db.query(Transporter).filter(Transporter.id == tid).first()
        if not t:
            raise HTTPException(404, "Not found")
        for k, v in inp.dict().items():
            setattr(t, k, v)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/transporters/{tid}")
def delete_transporter(tid: int, user: User = Depends(require_permission("transporters", "delete"))):
    db = SessionLocal()
    try:
        t = db.query(Transporter).filter(Transporter.id == tid).first()
        if not t:
            raise HTTPException(404, "Not found")
        db.delete(t)
        db.commit()
        return {"message": "Transporter deleted"}
    finally:
        db.close()


@app.post("/api/fix-urls")
def fix_urls(user: User = Depends(require_permission("transporters", "edit"))):
    db = SessionLocal()
    try:
        rows = db.query(Transporter).all()
        fixed = 0
        for t in rows:
            changed = False
            for field in ["gst_certificate", "pan_card"]:
                url = getattr(t, field, "")
                if url and "/image/upload/" in url and url.endswith(".pdf"):
                    setattr(t, field, url.replace("/image/upload/", "/raw/upload/"))
                    changed = True
            if changed:
                fixed += 1
        db.commit()
        return {"fixed": fixed}
    finally:
        db.close()


@app.get("/api/view-file")
async def view_file(url: str = Query(...), user: User = Depends(get_current_user)):
    from urllib.parse import urlparse
    try:
        parsed = urlparse(url)
        allowed_hosts = ["res.cloudinary.com", "cloudinary.com"]
        if parsed.scheme not in ("https", "http") or parsed.hostname not in allowed_hosts:
            raise HTTPException(400, "URL not allowed. Only Cloudinary URLs are permitted.")
        import ssl
        ctx = ssl.create_default_context()
        req = urllib.request.Request(url)
        with urllib.request.urlopen(req, context=ctx) as resp:
            data = resp.read()
            if data[:4] == b'%PDF':
                content_type = 'application/pdf'
            elif data[:2] == b'\xff\xd8':
                content_type = 'image/jpeg'
            elif data[:8] == b'\x89PNG\r\n\x1a\n':
                content_type = 'image/png'
            else:
                content_type = resp.headers.get("Content-Type", "application/octet-stream")
            return Response(content=data, media_type=content_type, headers={
                "Content-Disposition": "inline",
            })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to load file: {str(e)}")


# ---- PURCHASE RATES ----
@app.get("/api/purchase-rates")
def list_purchase_rates(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rates = db.query(PurchaseRate).all()
        # Batch-load products to avoid N+1
        product_ids = list(set(r.product_id for r in rates if r.product_id))
        product_map = {}
        if product_ids:
            products = db.query(Product).filter(Product.id.in_(product_ids)).all()
            product_map = {p.id: p for p in products}
        result = []
        for r in rates:
            p = product_map.get(r.product_id)
            result.append({
                "id": r.id, "product_id": r.product_id,
                "product_name": p.name if p else "", "part_no": p.part_no if p else "",
                "category": p.category if p else "", "size": p.size if p else "",
                "rate": r.rate, "supplier": r.supplier,
                "effective_date": r.effective_date.isoformat() if r.effective_date else ""
            })
        return result
    finally:
        db.close()


@app.post("/api/purchase-rates")
def create_purchase_rate(inp: PurchaseRateIn, user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        pr = PurchaseRate(
            product_id=inp.product_id, rate=inp.rate,
            supplier=inp.supplier,
            effective_date=date.fromisoformat(inp.effective_date) if inp.effective_date else date.today()
        )
        db.add(pr)
        db.commit()
        return {"id": pr.id, "message": "Purchase rate added"}
    finally:
        db.close()


@app.put("/api/purchase-rates/{prid}")
def update_purchase_rate(prid: int, inp: PurchaseRateUpdateIn, user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        pr = db.query(PurchaseRate).filter(PurchaseRate.id == prid).first()
        if not pr:
            raise HTTPException(404, "Not found")
        pr.rate = inp.rate
        pr.supplier = inp.supplier
        if inp.effective_date:
            pr.effective_date = date.fromisoformat(inp.effective_date)
        db.commit()
        return {"message": "Updated"}
    finally:
        db.close()


@app.delete("/api/purchase-rates/{prid}")
def delete_purchase_rate(prid: int, user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        pr = db.query(PurchaseRate).filter(PurchaseRate.id == prid).first()
        if not pr:
            raise HTTPException(404, "Not found")
        db.delete(pr)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


@app.post("/api/purchase-rates/bulk")
def bulk_create_purchase_rates(inp: BulkPurchaseRateIn, user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        rates = inp.rates
        created = 0
        for r in rates:
            pr = PurchaseRate(
                product_id=r.product_id, rate=r.rate,
                supplier=r.supplier,
                effective_date=date.fromisoformat(r.effective_date) if r.effective_date else date.today()
            )
            db.add(pr)
            created += 1
        db.commit()
        return {"message": f"{created} rates added", "count": created}
    finally:
        db.close()


# ---- GP CALCULATION ----
@app.get("/api/proforma-orders/{oid}/gp")
def calculate_gp(oid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).all()

        purchase_total = 0
        for item in items:
            pr = db.query(PurchaseRate).filter(PurchaseRate.product_id == item.product_id).order_by(PurchaseRate.effective_date.desc()).first()
            if pr:
                purchase_total += (item.final_qty or 0) * pr.rate

        sales_value = order.value_excl_gst or 0
        transport = order.transport_cost or 0
        gp = sales_value - purchase_total - transport
        gst_component = order.gst_amount or 0
        np_val = gp - gst_component

        order.purchase_total = purchase_total
        order.gross_profit = gp
        order.net_profit = np_val
        db.commit()

        return {
            "sales_value": sales_value, "purchase_total": purchase_total,
            "transport_cost": transport, "gross_profit": gp,
            "gst_amount": gst_component, "net_profit": np_val,
            "gp_percent": round(gp / sales_value * 100, 2) if sales_value else 0
        }
    finally:
        db.close()


@app.put("/api/proforma-orders/{oid}/transport")
def update_transport(oid: int, inp: TransportUpdateIn, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        order.transport_cost = inp.transport_cost
        db.commit()
        return {"message": "Transport updated"}
    finally:
        db.close()


@app.put("/api/proforma-orders/{oid}/status")
def update_order_status(oid: int, inp: OrderStatusIn, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        new_status = inp.status
        valid_statuses = ["draft", "confirmed", "po_created", "transport_pending", "transport_finalized", "billing", "completed"]
        if new_status not in valid_statuses:
            raise HTTPException(400, f"Invalid status. Must be one of: {', '.join(valid_statuses)}")
        order.status = new_status
        order.updated_at = datetime.now(timezone.utc)
        db.commit()
        return {"message": f"Order status updated to {new_status}"}
    finally:
        db.close()


# ---- WHATSAPP INTEGRATION ----

def upload_whatsapp_media(file_bytes, filename):
    """Upload a file to WhatsApp media endpoint, returns media_id"""
    url = f"{WHATSAPP_API_URL}/{WHATSAPP_PHONE_ID}/media"
    headers = {"Authorization": f"Bearer {WHATSAPP_TOKEN}"}
    files = {"file": (filename, file_bytes, "application/pdf")}
    data = {"messaging_product": "whatsapp", "type": "application/pdf"}
    try:
        resp = requests.post(url, headers=headers, files=files, data=data, timeout=30)
        result = resp.json()
        if resp.status_code == 200:
            return result.get("id")
    except Exception as e:
        logger.error("WhatsApp media upload failed: %s", e)
    return None

def send_whatsapp_message(phone_number, message, media_url=None, doc_url=None, doc_filename=None, doc_media_id=None):
    """Send a WhatsApp message using Meta Cloud API"""
    url = f"{WHATSAPP_API_URL}/{WHATSAPP_PHONE_ID}/messages"
    headers = {
        "Authorization": f"Bearer {WHATSAPP_TOKEN}",
        "Content-Type": "application/json"
    }
    
    # Format phone number (remove spaces, dashes, + sign)
    phone = phone_number.replace(" ", "").replace("-", "").replace("+", "")
    if not phone.startswith("91") and len(phone) == 10:
        phone = "91" + phone
    
    # Validate phone number
    if len(phone) < 12:
        return {"success": False, "error": "Invalid phone number. Use 10-digit Indian number."}
    
    # Send document (PDF) via media_id (preferred - no external URL needed)
    if doc_media_id:
        payload = {
            "messaging_product": "whatsapp",
            "to": phone,
            "type": "document",
            "document": {
                "id": doc_media_id,
                "filename": doc_filename or "document.pdf"
            }
        }
    # Send document (PDF) via URL
    elif doc_url:
        payload = {
            "messaging_product": "whatsapp",
            "to": phone,
            "type": "document",
            "document": {
                "link": doc_url,
                "filename": doc_filename or "document.pdf"
            }
        }
    # Send image
    elif media_url:
        payload = {
            "messaging_product": "whatsapp",
            "to": phone,
            "type": "image",
            "image": {"link": media_url}
        }
    # Send text
    else:
        payload = {
            "messaging_product": "whatsapp",
            "to": phone,
            "type": "text",
            "text": {"body": message}
        }
    
    try:
        resp = requests.post(url, json=payload, headers=headers, timeout=30)
        result = resp.json()
        if resp.status_code == 200:
            return {"success": True, "message_id": result.get("messages", [{}])[0].get("id")}
        else:
            error_msg = result.get("error", {}).get("message", "Unknown error")
            error_code = result.get("error", {}).get("code", 0)
            
            # Provide helpful error messages
            if error_code == 131047:
                error_msg = "Recipient hasn't messaged yet. Send a message to this number first, then try again."
            elif error_code == 131026:
                error_msg = "Message undeliverable. Check phone number and try again."
            
            return {"success": False, "error": error_msg, "error_code": error_code}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/whatsapp/send")
def whatsapp_send(inp: WhatsAppSendIn, user: User = Depends(require_permission("proforma_orders", "edit"))):
    """Send a WhatsApp message"""
    phone = inp.phone
    message = inp.message
    media_url = None
    
    if not phone or not message:
        raise HTTPException(400, "Phone and message are required")
    
    result = send_whatsapp_message(phone, message, media_url)
    return result


@app.post("/api/whatsapp/send-pi/{oid}")
def whatsapp_send_pi(oid: int, inp: WhatsAppSendPIIn, user: User = Depends(require_permission("proforma_orders", "edit"))):
    """Send PI PDF to a phone number via WhatsApp"""
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        
        customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).all()
        phone = inp.phone
        if not phone and customer:
            phone = customer.contact_number
        
        if not phone:
            raise HTTPException(400, "Phone number required")
        
        # Get billing site
        billing_site = None
        if order.billing_site:
            try:
                billing_site = db.query(BillingSite).filter(BillingSite.id == int(order.billing_site)).first()
            except (ValueError, TypeError):
                logger.warning("Invalid billing_site ID '%s' for order %s", order.billing_site, order.id)
        
        # Generate PI PDF
        pi_date = order.pi_date.strftime("%d-%b-%Y") if order.pi_date else ""
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Header
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, "Raksha Pipes Pvt. Ltd.", 0, 1, 'C')
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 8, f"Proforma Invoice - {order.pi_no}", 0, 1, 'C')
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 6, f"Date: {pi_date}", 0, 1, 'L')
        pdf.cell(0, 6, f"Customer: {customer.contact_name if customer else '-'}", 0, 1, 'L')
        pdf.ln(5)
        
        # Items table
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(15, 7, 'S.No', 1, 0, 'C')
        pdf.cell(30, 7, 'Part No', 1, 0, 'C')
        pdf.cell(50, 7, 'Description', 1, 0, 'C')
        pdf.cell(20, 7, 'Qty', 1, 0, 'C')
        pdf.cell(30, 7, 'Rate', 1, 0, 'C')
        pdf.cell(35, 7, 'Amount', 1, 1, 'C')
        
        pdf.set_font('Arial', '', 9)
        for i, item in enumerate(items, 1):
            pdf.cell(15, 6, str(i), 1, 0, 'C')
            pdf.cell(30, 6, str(item.part_no or '')[:15], 1, 0, 'C')
            pdf.cell(50, 6, str(item.description or '')[:25], 1, 0, 'L')
            pdf.cell(20, 6, str(item.final_qty or 0), 1, 0, 'C')
            pdf.cell(30, 6, f"Rs.{item.net_rate:,.2f}", 1, 0, 'R')
            pdf.cell(35, 6, f"Rs.{item.basic_amount:,.2f}", 1, 1, 'R')
        
        # Totals
        pdf.ln(3)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(115, 7, 'Total:', 0, 0, 'R')
        pdf.cell(35, 7, f"Rs.{order.value_excl_gst:,.2f}", 0, 1, 'R')
        pdf.cell(115, 7, 'GST (18%):', 0, 0, 'R')
        pdf.cell(35, 7, f"Rs.{order.gst_amount:,.2f}", 0, 1, 'R')
        pdf.cell(115, 7, 'Freight:', 0, 0, 'R')
        pdf.cell(35, 7, f"Rs.{order.freight_amount:,.2f}", 0, 1, 'R')
        pdf.cell(115, 7, 'Grand Total:', 0, 0, 'R')
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(35, 7, f"Rs.{order.total_amount:,.2f}", 0, 1, 'R')
        
        # Generate PDF
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        tmp_path = tmp.name
        tmp.close()
        pdf.output(tmp_path)
        
        with open(tmp_path, "rb") as f:
            pdf_bytes = f.read()
        
        # Method 1: Try WhatsApp media upload (works for permanent tokens)
        media_id = upload_whatsapp_media(pdf_bytes, f"PI_{order.pi_no}.pdf")
        if media_id:
            os.unlink(tmp_path)
            result = send_whatsapp_message(phone, "", doc_media_id=media_id, doc_filename=f"PI_{order.pi_no}.pdf")
        else:
            # Method 2: Try Cloudinary URL
            pdf_url = None
            try:
                upload_result = cloudinary.uploader.upload(tmp_path, resource_type="raw", folder="whatsapp_pi")
                pdf_url = upload_result.get("secure_url")
            except Exception as e:
                logger.error("Cloudinary upload failed for WhatsApp PI: %s", e)
            os.unlink(tmp_path)
            
            if pdf_url:
                result = send_whatsapp_message(phone, "", doc_url=pdf_url, doc_filename=f"PI_{order.pi_no}.pdf")
            else:
                return {"success": False, "error": "PDF upload failed. Check CLOUDINARY_URL env var on Render."}
        
        # Update whatsapp_status
        if result["success"]:
            order.whatsapp_status = "sent"
            db.commit()
        
        return result
    finally:
        db.close()


@app.post("/api/whatsapp/send-po/{oid}")
def whatsapp_send_po(oid: int, inp: WhatsAppSendPOIn, user: User = Depends(require_permission("proforma_orders", "edit"))):
    """Send PO PDF to a phone number via WhatsApp"""
    db = SessionLocal()
    try:
        order = db.query(ProformaOrder).filter(ProformaOrder.id == oid).first()
        if not order:
            raise HTTPException(404, "Order not found")
        
        customer = db.query(Customer).filter(Customer.id == order.customer_id).first()
        items = db.query(ProformaOrderItem).filter(ProformaOrderItem.proforma_order_id == oid).all()
        phone = inp.phone
        if not phone and customer:
            phone = customer.contact_number
        
        if not phone:
            raise HTTPException(400, "Phone number required")
        
        # Get billing site
        billing_site = None
        if order.billing_site:
            try:
                billing_site = db.query(BillingSite).filter(BillingSite.id == int(order.billing_site)).first()
            except (ValueError, TypeError):
                logger.warning("Invalid billing_site ID '%s' for order %s", order.billing_site, order.id)
        
        # Generate PO PDF
        pi_date = order.pi_date.strftime("%d-%b-%Y") if order.pi_date else ""
        
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)
        
        # Header
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, "Raksha Pipes Pvt. Ltd.", 0, 1, 'C')
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 8, f"Purchase Order - {order.po_no or order.pi_no}", 0, 1, 'C')
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 6, f"Date: {pi_date}", 0, 1, 'L')
        pdf.cell(0, 6, f"Party: {customer.contact_name if customer else '-'}", 0, 1, 'L')
        pdf.ln(5)
        
        # Items table
        pdf.set_font('Arial', 'B', 9)
        pdf.cell(15, 7, 'S.No', 1, 0, 'C')
        pdf.cell(30, 7, 'Part No', 1, 0, 'C')
        pdf.cell(50, 7, 'Description', 1, 0, 'C')
        pdf.cell(20, 7, 'Qty', 1, 0, 'C')
        pdf.cell(30, 7, 'Rate', 1, 0, 'C')
        pdf.cell(35, 7, 'Amount', 1, 1, 'C')
        
        pdf.set_font('Arial', '', 9)
        for i, item in enumerate(items, 1):
            pdf.cell(15, 6, str(i), 1, 0, 'C')
            pdf.cell(30, 6, str(item.part_no or '')[:15], 1, 0, 'C')
            pdf.cell(50, 6, str(item.description or '')[:25], 1, 0, 'L')
            pdf.cell(20, 6, str(item.final_qty or 0), 1, 0, 'C')
            pdf.cell(30, 6, f"Rs.{item.net_rate:,.2f}", 1, 0, 'R')
            pdf.cell(35, 6, f"Rs.{item.basic_amount:,.2f}", 1, 1, 'R')
        
        # Totals
        pdf.ln(3)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(115, 7, 'Total:', 0, 0, 'R')
        pdf.cell(35, 7, f"Rs.{order.value_excl_gst:,.2f}", 0, 1, 'R')
        pdf.cell(115, 7, 'GST (18%):', 0, 0, 'R')
        pdf.cell(35, 7, f"Rs.{order.gst_amount:,.2f}", 0, 1, 'R')
        pdf.cell(115, 7, 'Freight:', 0, 0, 'R')
        pdf.cell(35, 7, f"Rs.{order.freight_amount:,.2f}", 0, 1, 'R')
        pdf.cell(115, 7, 'Grand Total:', 0, 0, 'R')
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(35, 7, f"Rs.{order.total_amount:,.2f}", 0, 1, 'R')
        
        # Generate PDF
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        tmp_path = tmp.name
        tmp.close()
        pdf.output(tmp_path)
        
        with open(tmp_path, "rb") as f:
            pdf_bytes = f.read()
        os.unlink(tmp_path)
        
        # Upload to WhatsApp media (no Cloudinary needed)
        media_id = upload_whatsapp_media(pdf_bytes, f"PO_{order.po_no or order.pi_no}.pdf")
        
        if media_id:
            result = send_whatsapp_message(phone, "", doc_media_id=media_id, doc_filename=f"PO_{order.po_no or order.pi_no}.pdf")
        else:
            result = {"success": False, "error": "Failed to upload PO PDF to WhatsApp"}
        
        # Update whatsapp_status
        if result["success"]:
            order.whatsapp_status = "sent"
            db.commit()
        
        return result
    finally:
        db.close()


@app.get("/api/whatsapp/config")
def get_whatsapp_config(user: User = Depends(get_current_user)):
    """Get WhatsApp configuration status"""
    return {
        "configured": bool(WHATSAPP_TOKEN and WHATSAPP_PHONE_ID),
        "phone_id": WHATSAPP_PHONE_ID[:10] + "..." if WHATSAPP_PHONE_ID else None,
        "business_account_id": WHATSAPP_BUSINESS_ACCOUNT_ID[:10] + "..." if WHATSAPP_BUSINESS_ACCOUNT_ID else None,
        "token_set": bool(WHATSAPP_TOKEN)
    }


@app.post("/api/whatsapp/test")
def whatsapp_test(inp: WhatsAppTestIn, user: User = Depends(get_current_user)):
    """Send a simple test text message"""
    phone = inp.phone
    if not phone:
        raise HTTPException(400, "Phone number required")
    
    message = "Test message from Raksha ERP. If you received this, WhatsApp is working!"
    result = send_whatsapp_message(phone, message)
    return result


# ---- SALES ----
@app.get("/api/sales")
def list_sales(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Sale).order_by(Sale.id.desc()).all()
        # Batch-load customer names to avoid N+1
        cust_ids = list(set(s.customer_id for s in rows if s.customer_id))
        cust_map = {}
        if cust_ids:
            custs = db.query(Customer).filter(Customer.id.in_(cust_ids)).all()
            cust_map = {c.id: c.contact_name for c in custs}
        # Batch-load all sale items to avoid N+1
        sale_ids = [s.id for s in rows]
        all_items = db.query(SaleItem).filter(SaleItem.sale_id.in_(sale_ids)).order_by(SaleItem.sl_no).all()
        items_map = {}
        for si in all_items:
            items_map.setdefault(si.sale_id, []).append(si)
        out = []
        for s in rows:
            try:
                cust_name = cust_map.get(s.customer_id, "") if s.customer_id else ""
                party = s.party_name or cust_name or ""
                loc = s.location or ""
                sale_items = [
                    {
                        "id": si.id, "sl_no": si.sl_no,
                        "product_id": si.product_id,
                        "quantity": si.quantity, "unit_price": si.unit_price,
                        "discount_percent": si.discount_percent,
                        "taxable_amount": si.taxable_amount,
                        "gst_rate": si.gst_rate,
                        "cgst_amount": si.cgst_amount,
                        "sgst_amount": si.sgst_amount,
                        "total_amount": si.total_amount,
                    }
                    for si in items_map.get(s.id, [])
                ]
                out.append({
                    "id": s.id, "invoice_no": s.invoice_no or "",
                    "customer_id": s.customer_id or None,
                    "product_id": s.product_id or None,
                    "quantity": s.quantity or 0,
                    "unit_price": s.unit_price or 0,
                    "discount_percent": s.discount_percent or 0,
                    "taxable_amount": s.taxable_amount or 0,
                    "cgst_amount": s.cgst_amount or 0,
                    "sgst_amount": s.sgst_amount or 0,
                    "freight_amount": s.freight_amount or 0,
                    "payment_status": s.payment_status or "",
                    "payment_method": s.payment_method or "",
                    "notes": s.notes or "",
                    "party_name": party, "location": loc,
                    "state": s.state or "",
                    "transporter_name": s.transporter_name or "",
                    "lr_no": s.lr_no or "",
                    "weight_kgs": s.weight_kgs or 0,
                    "gp": s.gp or 0,
                    "gp_percent": s.gp_percent or 0,
                    "invoice_value": s.invoice_value or 0,
                    "total_amount": s.total_amount or 0,
                    "sale_date": s.sale_date.isoformat() if s.sale_date else None,
                    "payment_terms": s.payment_terms or "",
                    "source_csv": s.source_csv or "",
                    "lr_tracking_status": s.lr_tracking_status or "",
                    "lr_tracking_url": s.lr_tracking_url or "",
                    "lr_last_checked": s.lr_last_checked.isoformat() if s.lr_last_checked else None,
                    "items": sale_items,
                })
            except Exception as e:
                print(f"Error loading sale {s.id}: {e}")
                continue
        return out
    finally:
        db.close()


@app.get("/api/sales/{sid}")
def get_sale(sid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(status_code=404, detail="Sale not found")
        cust_name = ""
        if s.customer_id:
            cust = db.query(Customer).filter(Customer.id == s.customer_id).first()
            cust_name = cust.contact_name if cust else ""
        party = s.party_name or cust_name or ""
        loc = s.location or ""
        sale_items = [
            {
                "id": si.id, "sl_no": si.sl_no,
                "product_id": si.product_id,
                "quantity": si.quantity, "unit_price": si.unit_price,
                "discount_percent": si.discount_percent,
                "taxable_amount": si.taxable_amount,
                "gst_rate": si.gst_rate,
                "cgst_amount": si.cgst_amount,
                "sgst_amount": si.sgst_amount,
                "total_amount": si.total_amount,
            }
            for si in db.query(SaleItem).filter(SaleItem.sale_id == s.id).order_by(SaleItem.sl_no).all()
        ]
        return {
            "id": s.id, "invoice_no": s.invoice_no or "",
            "customer_id": s.customer_id or None,
            "product_id": s.product_id or None,
            "quantity": s.quantity or 0,
            "unit_price": s.unit_price or 0,
            "discount_percent": s.discount_percent or 0,
            "taxable_amount": s.taxable_amount or 0,
            "cgst_amount": s.cgst_amount or 0,
            "sgst_amount": s.sgst_amount or 0,
            "freight_amount": s.freight_amount or 0,
            "payment_status": s.payment_status or "",
            "payment_method": s.payment_method or "",
            "notes": s.notes or "",
            "party_name": party, "location": loc,
            "state": s.state or "",
            "transporter_name": s.transporter_name or "",
            "lr_no": s.lr_no or "",
            "weight_kgs": s.weight_kgs or 0,
            "gp": s.gp or 0,
            "gp_percent": s.gp_percent or 0,
            "invoice_value": s.invoice_value or 0,
            "total_amount": s.total_amount or 0,
            "sale_date": s.sale_date.isoformat() if s.sale_date else None,
            "payment_terms": s.payment_terms or "",
            "source_csv": s.source_csv or "",
            "lr_tracking_status": s.lr_tracking_status or "",
            "lr_tracking_url": s.lr_tracking_url or "",
            "lr_last_checked": s.lr_last_checked.isoformat() if s.lr_last_checked else None,
            "items": sale_items,
        }
    finally:
        db.close()


@app.get("/api/sales/freight-summary")
def freight_summary(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Sale).filter(Sale.freight_amount > 0, Sale.weight_kgs > 0).order_by(Sale.id.desc()).all()
        out = []
        for s in rows:
            out.append({
                "id": s.id, "invoice_no": s.invoice_no or "",
                "sale_date": s.sale_date.isoformat() if s.sale_date else None,
                "freight_amount": s.freight_amount or 0,
                "weight_kgs": s.weight_kgs or 0,
                "transporter_name": s.transporter_name or "",
            })
        return out
    finally:
        db.close()


@app.post("/api/sales")
def create_sale(inp: SaleIn, user: User = Depends(require_permission("sales", "create"))):
    db = SessionLocal()
    try:
        # Generate invoice_no with retry to handle race conditions
        max_retries = 5
        for attempt in range(max_retries):
            try:
                max_id = db.query(func.max(Sale.id)).scalar() or 0
                invoice_no = f"RFRP-{max_id + 1:05d}"
                break
            except Exception:
                if attempt == max_retries - 1:
                    raise
                db.rollback()
                time.sleep(0.1)

        total_taxable = 0
        total_cgst = 0
        total_sgst = 0
        total_amount = 0

        s = Sale(
            invoice_no=invoice_no, customer_id=inp.customer_id,
            freight_amount=inp.freight_amount,
            payment_status=inp.payment_status, payment_method=inp.payment_method,
            notes=inp.notes, transporter_name=inp.transporter_name, lr_no=inp.lr_no,
            invoice_value=inp.invoice_value, sale_date=datetime.now(timezone.utc)
        )
        if inp.customer_id:
            cust = db.query(Customer).filter(Customer.id == inp.customer_id).first()
            if cust:
                s.party_name = cust.contact_name or cust.name or ""
                s.location = cust.billing_address or cust.city or ""
                s.state = cust.state or ""
        db.add(s)
        db.flush()

        for idx, item in enumerate(inp.items):
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            if not prod:
                continue
            pr = db.query(Pricing).filter(Pricing.product_id == item.product_id).first()
            gst_rate = pr.gst_rate if pr else 18

            taxable = item.quantity * item.unit_price
            disc_amt = taxable * item.discount_percent / 100
            taxable -= disc_amt
            cgst = taxable * gst_rate / 200
            sgst = taxable * gst_rate / 200
            item_total = taxable + cgst + sgst

            total_taxable += taxable
            total_cgst += cgst
            total_sgst += sgst
            total_amount += item_total

            db.add(SaleItem(
                sale_id=s.id, sl_no=idx + 1,
                product_id=item.product_id, quantity=item.quantity,
                unit_price=item.unit_price, discount_percent=item.discount_percent,
                discount_amount=disc_amt, taxable_amount=taxable,
                gst_rate=gst_rate, cgst_amount=cgst, sgst_amount=sgst,
                total_amount=item_total, basic_amount=item_total
            ))

        grand_total = total_amount + inp.freight_amount
        s.taxable_amount = total_taxable
        s.cgst_amount = total_cgst
        s.sgst_amount = total_sgst
        s.total_amount = grand_total
        if not s.invoice_value:
            s.invoice_value = grand_total

        if not inp.product_id and inp.items:
            s.product_id = inp.items[0].product_id
            s.quantity = sum(i.quantity for i in inp.items)
            s.unit_price = inp.items[0].unit_price
            s.discount_percent = inp.items[0].discount_percent

        db.commit()
        return {"invoice_no": invoice_no, "total": grand_total}
    finally:
        db.close()


@app.delete("/api/sales/{sid}")
def delete_sale(sid: int, user: User = Depends(require_permission("sales", "delete"))):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Not found")
        # Delete child sale items first
        db.query(SaleItem).filter(SaleItem.sale_id == sid).delete()
        db.delete(s)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


@app.patch("/api/sales/{sid}/invoice")
def patch_sale_invoice(sid: int, body: SaleInvoiceIn, user: User = Depends(require_permission("sales", "edit"))):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Not found")
        if body.invoice_value:
            s.invoice_value = body.invoice_value
        db.commit()
        return {"message": "Patched", "id": sid, "invoice_value": s.invoice_value}
    finally:
        db.close()


@app.put("/api/sales/{sid}")
def update_sale(sid: int, inp: SaleIn, user: User = Depends(require_permission("sales", "edit"))):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Not found")

        total_taxable = 0
        total_cgst = 0
        total_sgst = 0
        total_amount = 0

        db.query(SaleItem).filter(SaleItem.sale_id == sid).delete()

        for idx, item in enumerate(inp.items):
            prod = db.query(Product).filter(Product.id == item.product_id).first()
            if not prod:
                continue
            pr = db.query(Pricing).filter(Pricing.product_id == item.product_id).first()
            gst_rate = pr.gst_rate if pr else 18

            taxable = item.quantity * item.unit_price
            disc_amt = taxable * item.discount_percent / 100
            taxable -= disc_amt
            cgst = taxable * gst_rate / 200
            sgst = taxable * gst_rate / 200
            item_total = taxable + cgst + sgst

            total_taxable += taxable
            total_cgst += cgst
            total_sgst += sgst
            total_amount += item_total

            db.add(SaleItem(
                sale_id=s.id, sl_no=idx + 1,
                product_id=item.product_id, quantity=item.quantity,
                unit_price=item.unit_price, discount_percent=item.discount_percent,
                discount_amount=disc_amt, taxable_amount=taxable,
                gst_rate=gst_rate, cgst_amount=cgst, sgst_amount=sgst,
                total_amount=item_total, basic_amount=item_total
            ))

        grand_total = total_amount + inp.freight_amount
        s.customer_id = inp.customer_id
        if inp.customer_id:
            cust = db.query(Customer).filter(Customer.id == inp.customer_id).first()
            if cust:
                s.party_name = cust.contact_name or cust.name or ""
                s.location = cust.billing_address or cust.city or ""
                s.state = cust.state or ""
        s.freight_amount = inp.freight_amount
        s.taxable_amount = total_taxable
        s.cgst_amount = total_cgst
        s.sgst_amount = total_sgst
        s.total_amount = grand_total
        s.invoice_value = inp.invoice_value or grand_total
        s.payment_status = inp.payment_status
        s.payment_method = inp.payment_method
        s.notes = inp.notes
        s.transporter_name = inp.transporter_name or s.transporter_name or ""
        s.lr_no = inp.lr_no or s.lr_no or ""

        if inp.items:
            s.product_id = inp.items[0].product_id
            s.quantity = sum(i.quantity for i in inp.items)
            s.unit_price = inp.items[0].unit_price
            s.discount_percent = inp.items[0].discount_percent

        db.commit()
        return {"message": "Sale updated", "total": grand_total}
    finally:
        db.close()


@app.patch("/api/sales/bulk-payment")
def bulk_update_payment_status(body: BulkPaymentIn, user: User = Depends(require_permission("sales", "bulk_edit"))):
    status = body.status
    db = SessionLocal()
    try:
        updated = db.query(Sale).filter(Sale.id.in_(body.ids)).update({"payment_status": status}, synchronize_session=False)
        db.commit()
        return {"updated": updated, "message": f"Updated {updated} sales to {status}"}
    finally:
        db.close()


@app.patch("/api/sales/bulk-lr-status")
def bulk_update_lr_status(body: BulkLRIn, user: User = Depends(require_permission("sales", "bulk_edit"))):
    status = body.lr_no or "Delivered"
    db = SessionLocal()
    try:
        q = db.query(Sale)
        if body.ids:
            q = q.filter(Sale.id.in_(body.ids))
        updated = q.filter(Sale.lr_tracking_status != status).update({"lr_tracking_status": status}, synchronize_session=False)
        db.commit()
        return {"updated": updated, "message": f"Updated {updated} sales to {status}"}
    finally:
        db.close()


@app.delete("/api/customers/by-id/{customer_id}")
def delete_customer_by_customer_id(customer_id: str, user: User = Depends(require_permission("customers", "delete"))):
    db = SessionLocal()
    try:
        c = db.query(Customer).filter(Customer.customer_id == customer_id).first()
        if not c:
            raise HTTPException(404, "Not found")
        db.query(Sale).filter(Sale.customer_id == c.id).update({"customer_id": None})
        db.delete(c)
        db.commit()
        return {"message": f"Deleted {customer_id}"}
    finally:
        db.close()


# ---- LR TRACKING ----
@app.put("/api/sales/{sid}/lr-tracking")
def update_lr_tracking(sid: int, body: LRTrackingIn, user: User = Depends(require_permission("sales", "edit"))):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Sale not found")
        if body.lr_no:
            s.lr_no = body.lr_no
        if body.tracking_url:
            s.lr_tracking_url = body.tracking_url
        s.lr_last_checked = datetime.now(timezone.utc)
        db.commit()
        return {"message": "LR tracking updated"}
    finally:
        db.close()


@app.post("/api/sales/{sid}/generate-tracking-url")
def generate_tracking_url(sid: int, user: User = Depends(require_permission("sales", "edit"))):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Sale not found")
        if not s.transporter_name or not s.lr_no:
            return {"tracking_url": "", "message": "No transporter or LR number found"}
        t = db.query(Transporter).filter(Transporter.name.ilike(s.transporter_name)).first()
        if t and t.tracking_url_pattern:
            tracking_url = t.tracking_url_pattern.replace("{lr_no}", s.lr_no).replace("{LR_NO}", s.lr_no)
            s.lr_tracking_url = tracking_url
            s.lr_last_checked = datetime.now(timezone.utc)
            db.commit()
            return {"tracking_url": tracking_url}
        name_lower = s.transporter_name.lower()
        default_patterns = {
            "vrl": "https://vrlgroup.in/Track/LRNumber/{lr_no}",
            "v trans": "https://vrlgroup.in/Track/LRNumber/{lr_no}",
            "dtdc": "https://www.dtdc.in/tracking/dtdc-tracking-results.asp?Lrnos={lr_no}",
            "safexpress": "https://www.safexpress.com/track-trace/{lr_no}",
            "gati": "https://www.gati.com/shipmentTracking/{lr_no}",
            "professional": "https://www.professional.couriers.in/tracking/{lr_no}",
            "ecom": "https://www.ecomexpress.in/tracking/{lr_no}",
            "delhivery": "https://www.delhivery.com/tracking/package/{lr_no}",
        }
        for key, pattern in default_patterns.items():
            if key in name_lower:
                tracking_url = pattern.replace("{lr_no}", s.lr_no)
                s.lr_tracking_url = tracking_url
                s.lr_last_checked = datetime.now(timezone.utc)
                db.commit()
                return {"tracking_url": tracking_url}
        return {"tracking_url": "", "message": "No tracking URL pattern configured for this transporter. Add one in Transporter settings."}
    finally:
        db.close()


@app.get("/api/sales/{sid}/lr-tracking")
def get_lr_tracking(sid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Sale not found")
        return {
            "lr_tracking_status": s.lr_tracking_status or "",
            "lr_tracking_url": s.lr_tracking_url or "",
            "lr_last_checked": s.lr_last_checked.isoformat() if s.lr_last_checked else None,
            "transporter_name": s.transporter_name or "",
            "lr_no": s.lr_no or "",
        }
    finally:
        db.close()


@app.post("/api/auto-generate-tracking-urls")
def auto_generate_tracking_urls(user: User = Depends(require_permission("sales", "edit"))):
    db = SessionLocal()
    try:
        updated = 0
        sales = db.query(Sale).filter(Sale.lr_no != "", Sale.lr_no != None).all()
        transporters = {t.name.lower(): t for t in db.query(Transporter).all()}
        for s in sales:
            if s.lr_tracking_url:
                continue
            t = transporters.get((s.transporter_name or "").lower())
            if t and t.tracking_url_pattern:
                url = t.tracking_url_pattern.replace("{lr_no}", s.lr_no or "").replace("{LR_NO}", s.lr_no or "")
                s.lr_tracking_url = url
                updated += 1
        db.commit()
        return {"message": f"Generated tracking URLs for {updated} sales", "updated": updated}
    finally:
        db.close()


# ---- AUTO FETCH TRACKING STATUS ----
http_requests = requests
try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False
    class BeautifulSoup:
        def __init__(self, *a, **kw): pass
        def get_text(self): return ""
        def select(self, *a, **kw): return []

TRACKING_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
}

def fetch_vrl_tracking(lr_no):
    try:
        url = f"https://vrlgroup.in/Track/LRNumber/{lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        rows = soup.select('table tr')
        statuses = []
        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 3:
                date_text = cells[0].get_text(strip=True)
                location = cells[1].get_text(strip=True)
                status = cells[2].get_text(strip=True)
                if status:
                    statuses.append({"date": date_text, "location": location, "status": status})
        if statuses:
            latest = statuses[-1]
            return {"status": latest["status"], "location": latest["location"], "date": latest["date"], "history": statuses, "source": "VRL Logistics"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


def fetch_dtdc_tracking(lr_no):
    try:
        url = f"https://www.dtdc.in/tracking/dtdc-tracking-results.asp?Lrnos={lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        rows = soup.select('table tr')
        statuses = []
        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 3:
                date_text = cells[0].get_text(strip=True)
                location = cells[1].get_text(strip=True)
                status = cells[2].get_text(strip=True)
                if status and date_text:
                    statuses.append({"date": date_text, "location": location, "status": status})
        if statuses:
            latest = statuses[-1]
            return {"status": latest["status"], "location": latest["location"], "date": latest["date"], "history": statuses, "source": "DTDC"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


def fetch_safexpress_tracking(lr_no):
    try:
        url = f"https://www.safexpress.com/track-trace/{lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text()
        statuses = []
        for keyword in ["Picked Up", "In Transit", "Out for Delivery", "Delivered", "Exception", "Not Delivered"]:
            if keyword.lower() in text.lower():
                statuses.append(keyword)
        if statuses:
            return {"status": statuses[-1], "location": "", "date": "", "history": [], "source": "Safexpress"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


def fetch_gati_tracking(lr_no):
    try:
        url = f"https://www.gati.com/shipmentTracking/{lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text()
        statuses = []
        for keyword in ["Picked Up", "In Transit", "Out for Delivery", "Delivered", "Reached Destination", "Dispatched"]:
            if keyword.lower() in text.lower():
                statuses.append(keyword)
        if statuses:
            return {"status": statuses[-1], "location": "", "date": "", "history": [], "source": "Gati"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


def fetch_professional_tracking(lr_no):
    try:
        url = f"https://www.professional.couriers.in/tracking/{lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text()
        statuses = []
        for keyword in ["Picked Up", "In Transit", "Out for Delivery", "Delivered", "Exception"]:
            if keyword.lower() in text.lower():
                statuses.append(keyword)
        if statuses:
            return {"status": statuses[-1], "location": "", "date": "", "history": [], "source": "Professional Couriers"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


def fetch_ecom_express_tracking(lr_no):
    try:
        url = f"https://www.ecomexpress.in/tracking/{lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text()
        statuses = []
        for keyword in ["Picked", "In Transit", "Out for Delivery", "Delivered", "Reached"]:
            if keyword.lower() in text.lower():
                statuses.append(keyword)
        if statuses:
            return {"status": statuses[-1], "location": "", "date": "", "history": [], "source": "Ecom Express"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


def fetch_delhivery_tracking(lr_no):
    try:
        url = f"https://www.delhivery.com/tracking/package/{lr_no}"
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False)
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text()
        statuses = []
        for keyword in ["Picked Up", "In Transit", "Out for Delivery", "Delivered", "Reached Destination Hub"]:
            if keyword.lower() in text.lower():
                statuses.append(keyword)
        if statuses:
            return {"status": statuses[-1], "location": "", "date": "", "history": [], "source": "Delhivery"}
        return {"status": "", "message": "No tracking data found"}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


TRANSPORTER_TRACKERS = {
    "vrl": fetch_vrl_tracking,
    "vrl logistics": fetch_vrl_tracking,
    "vrl group": fetch_vrl_tracking,
    "v trans": fetch_vrl_tracking,
    "v xpress": fetch_vrl_tracking,
    "dtdc": fetch_dtdc_tracking,
    "dtdc courier": fetch_dtdc_tracking,
    "dtdc express": fetch_dtdc_tracking,
    "safexpress": fetch_safexpress_tracking,
    "saf express": fetch_safexpress_tracking,
    "gati": fetch_gati_tracking,
    "gati courier": fetch_gati_tracking,
    "professional": fetch_professional_tracking,
    "professional couriers": fetch_professional_tracking,
    "professional courier": fetch_professional_tracking,
    "ecom": fetch_ecom_express_tracking,
    "ecom express": fetch_ecom_express_tracking,
    "delhivery": fetch_delhivery_tracking,
}


def get_tracker_for_transporter(transporter_name):
    name_lower = (transporter_name or "").lower().strip()
    for key, func in TRANSPORTER_TRACKERS.items():
        if key in name_lower:
            return func
    return None


def fetch_generic_tracking(lr_no, tracking_url_pattern):
    try:
        if not tracking_url_pattern:
            return {"status": "", "message": "No tracking URL pattern configured"}
        url = tracking_url_pattern.replace("{lr_no}", lr_no).replace("{LR_NO}", lr_no)
        r = http_requests.get(url, headers=TRACKING_HEADERS, timeout=15, verify=False, allow_redirects=True)
        soup = BeautifulSoup(r.text, 'html.parser')
        text = soup.get_text().lower()
        statuses = []
        for keyword in ["delivered", "out for delivery", "in transit", "picked up", "dispatched", "reached destination", "arrived", "exception", "delayed", "not delivered"]:
            if keyword in text:
                statuses.append(keyword.title())
        if statuses:
            return {"status": statuses[0], "location": "", "date": "", "history": [], "source": "Web", "url": url}
        return {"status": "", "message": "Could not parse tracking status from page", "url": url}
    except Exception as e:
        return {"status": "", "message": f"Failed to fetch: {str(e)}"}


@app.post("/api/fetch-tracking/{sid}")
def fetch_tracking_status(sid: int, user: User = Depends(require_permission("sales", "edit"))):
    db = SessionLocal()
    try:
        s = db.query(Sale).filter(Sale.id == sid).first()
        if not s:
            raise HTTPException(404, "Sale not found")
        if not s.lr_no:
            return {"status": "", "message": "No LR number set for this sale"}
        tracker = get_tracker_for_transporter(s.transporter_name)
        if tracker:
            result = tracker(s.lr_no)
        else:
            t = db.query(Transporter).filter(Transporter.name.ilike(s.transporter_name)).first()
            if t and t.tracking_url_pattern:
                result = fetch_generic_tracking(s.lr_no, t.tracking_url_pattern)
            else:
                return {"status": "", "message": f"No auto-tracking available for transporter: {s.transporter_name}. Add a tracking URL pattern in Transporter settings."}
        if result.get("status"):
            status = result["status"]
            if "deliver" in status.lower():
                s.lr_tracking_status = "Delivered"
            elif "transit" in status.lower():
                s.lr_tracking_status = "In Transit"
            elif "out for delivery" in status.lower():
                s.lr_tracking_status = "Out for Delivery"
            elif "delay" in status.lower() or "exception" in status.lower():
                s.lr_tracking_status = "Delayed"
            else:
                s.lr_tracking_status = status
            s.lr_last_checked = datetime.now(timezone.utc)
            db.commit()
        if result.get("url") and not s.lr_tracking_url:
            s.lr_tracking_url = result["url"]
            db.commit()
        return result
    finally:
        db.close()


@app.post("/api/fetch-tracking-bulk")
def fetch_tracking_bulk(user: User = Depends(require_permission("sales", "bulk_edit"))):
    db = SessionLocal()
    try:
        sales = db.query(Sale).filter(Sale.lr_no != "", Sale.lr_no != None, Sale.lr_tracking_status != "Delivered").all()
        updated = 0
        results = []
        for s in sales:
            tracker = get_tracker_for_transporter(s.transporter_name)
            if tracker:
                result = tracker(s.lr_no)
            else:
                t = db.query(Transporter).filter(Transporter.name.ilike(s.transporter_name)).first()
                if t and t.tracking_url_pattern:
                    result = fetch_generic_tracking(s.lr_no, t.tracking_url_pattern)
                else:
                    continue
            try:
                if result.get("status"):
                    status = result["status"]
                    if "deliver" in status.lower():
                        s.lr_tracking_status = "Delivered"
                    elif "transit" in status.lower():
                        s.lr_tracking_status = "In Transit"
                    elif "out for delivery" in status.lower():
                        s.lr_tracking_status = "Out for Delivery"
                    elif "delay" in status.lower() or "exception" in status.lower():
                        s.lr_tracking_status = "Delayed"
                    else:
                        s.lr_tracking_status = status
                    s.lr_last_checked = datetime.now(timezone.utc)
                    if result.get("url") and not s.lr_tracking_url:
                        s.lr_tracking_url = result["url"]
                    updated += 1
                    results.append({"sale_id": s.id, "lr_no": s.lr_no, "status": s.lr_tracking_status})
            except Exception:
                continue
        db.commit()
        return {"message": f"Updated {updated} shipments", "updated": updated, "results": results}
    finally:
        db.close()


# ---- EXPENSES ----
@app.get("/api/expenses")
def list_expenses(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Expense).order_by(Expense.expense_date.desc()).all()
        return [{"id": e.id, "category": e.category, "description": e.description,
                 "amount": e.amount, "vendor": e.vendor,
                 "expense_date": e.expense_date.isoformat() if e.expense_date else None}
                for e in rows]
    finally:
        db.close()


@app.get("/api/expenses/{eid}")
def get_expense(eid: int, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        e = db.query(Expense).filter(Expense.id == eid).first()
        if not e:
            raise HTTPException(404, "Expense not found")
        return {"id": e.id, "category": e.category, "description": e.description,
                "amount": e.amount, "vendor": e.vendor,
                "expense_date": e.expense_date.isoformat() if e.expense_date else None}
    finally:
        db.close()


@app.post("/api/expenses")
def create_expense(inp: ExpenseIn, user: User = Depends(require_permission("expenses", "create"))):
    db = SessionLocal()
    try:
        dt = datetime.strptime(inp.expense_date, "%Y-%m-%d") if inp.expense_date else datetime.now(timezone.utc)
        e = Expense(category=inp.category, description=inp.description,
                    amount=inp.amount, vendor=inp.vendor, expense_date=dt)
        db.add(e)
        db.commit()
        return {"message": "Expense added"}
    finally:
        db.close()


@app.delete("/api/expenses/{eid}")
def delete_expense(eid: int, user: User = Depends(require_permission("expenses", "delete"))):
    db = SessionLocal()
    try:
        e = db.query(Expense).filter(Expense.id == eid).first()
        if not e:
            raise HTTPException(404, "Not found")
        db.delete(e)
        db.commit()
        return {"message": "Deleted"}
    finally:
        db.close()


@app.put("/api/expenses/{eid}")
def update_expense(eid: int, inp: ExpenseIn, user: User = Depends(require_permission("expenses", "edit"))):
    db = SessionLocal()
    try:
        e = db.query(Expense).filter(Expense.id == eid).first()
        if not e:
            raise HTTPException(404, "Not found")
        e.category = inp.category
        e.description = inp.description
        e.amount = inp.amount
        e.vendor = inp.vendor
        if inp.expense_date:
            try:
                e.expense_date = datetime.strptime(inp.expense_date, "%Y-%m-%d")
            except Exception:
                logger.warning("Invalid expense_date '%s' for expense %s", inp.expense_date, eid)
        db.commit()
        return {"message": "Expense updated"}
    finally:
        db.close()


# ---- REPORTS ----
@app.get("/api/reports/profit-loss")
def profit_loss(start_date: str = None, end_date: str = None, user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        sales_q = db.query(Sale)
        expenses_q = db.query(Expense)
        orders_q = db.query(Order)

        if start_date:
            try:
                sd = datetime.strptime(start_date, "%Y-%m-%d")
                sales_q = sales_q.filter(Sale.sale_date >= sd)
                expenses_q = expenses_q.filter(Expense.expense_date >= sd)
                orders_q = orders_q.filter(Order.entry_date >= start_date)
            except Exception:
                logger.warning("Invalid start_date for P&L: %s", start_date)
        if end_date:
            try:
                ed = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                sales_q = sales_q.filter(Sale.sale_date <= ed)
                expenses_q = expenses_q.filter(Expense.expense_date <= ed)
                orders_q = orders_q.filter(Order.entry_date <= end_date)
            except Exception:
                logger.warning("Invalid end_date for P&L: %s", end_date)

        sales = sales_q.all()
        expenses = expenses_q.all()
        orders = orders_q.all()

        # Revenue from Sales
        sale_revenue = sum(s.invoice_value or s.total_amount or 0 for s in sales)
        units = sum(s.quantity or 0 for s in sales)
        gst = sum((s.cgst_amount or 0) + (s.sgst_amount or 0) for s in sales)
        gp_from_sales = sum(s.gp or 0 for s in sales)
        gp_values = [s.gp_percent for s in sales if s.gp_percent and s.gp_percent > 0]
        gp_avg = sum(gp_values) / len(gp_values) if gp_values else 0

        # Freight from sales (rate per kg * weight = total freight cost)
        sale_freight = sum((s.freight_amount or 0) * (s.weight_kgs or 0) for s in sales)

        # COGS from Orders (cost of goods)
        order_cogs = sum(o.value_excl_gst_freight or 0 for o in orders)
        order_freight_cost = sum(o.transport_charges or 0 for o in orders)
        order_invoice_total = sum(o.invoice_amount or 0 for o in orders)
        order_boxes = sum(o.no_of_boxes or 0 for o in orders)
        order_weight = sum(o.weight_kgs or 0 for o in orders)
        order_credit_notes = sum(o.credit_note_amount or 0 for o in orders)

        # Total revenue = sales invoice values (which may or may not include freight)
        total_revenue = sale_revenue

        # Total COGS (order cost + transport)
        total_cogs = order_cogs + order_freight_cost

        exp_by_cat = {}
        for e in expenses:
            exp_by_cat[e.category] = exp_by_cat.get(e.category, 0) + e.amount
        total_opex = sum(exp_by_cat.values())

        # Gross Profit = Revenue - COGS - Credit Notes
        gross_profit = total_revenue - total_cogs - order_credit_notes
        gross_margin = (gross_profit / total_revenue * 100) if total_revenue else 0

        ebitda = gross_profit - total_opex
        tax_rate = float(get_setting("tax_rate", "25"))
        tax = ebitda * tax_rate / 100 if ebitda > 0 else 0
        pat = ebitda - tax

        return {
            "total_revenue": total_revenue,
            "sale_revenue": sale_revenue, "sale_freight": sale_freight,
            "total_cogs": total_cogs,
            "order_cogs": order_cogs, "order_freight_cost": order_freight_cost,
            "order_invoice_total": order_invoice_total,
            "order_boxes": order_boxes, "order_weight": order_weight,
            "order_credit_notes": order_credit_notes,
            "gst": gst, "units": units,
            "gross_profit": gross_profit, "gross_margin": gross_margin,
            "gp_from_sales": gp_from_sales, "gp_avg": gp_avg,
            "expenses": exp_by_cat, "total_opex": total_opex,
            "ebitda": ebitda, "ebitda_margin": (ebitda / total_revenue * 100) if total_revenue else 0,
            "tax_rate": float(get_setting("tax_rate", "25")), "tax": tax, "pat": pat,
            "total_orders": len(orders), "total_sales": len(sales),
        }
    finally:
        db.close()


# ---- DASHBOARD ----
@app.get("/api/dashboard")
def dashboard(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        # Use SQL aggregations instead of loading all records
        revenue = db.query(func.coalesce(func.sum(Sale.invoice_value), 0)).scalar() or 0
        # Fallback to total_amount if invoice_value is not set
        if revenue == 0:
            revenue = db.query(func.coalesce(func.sum(Sale.total_amount), 0)).scalar() or 0
        
        freight = db.query(func.coalesce(func.sum(Sale.freight_amount * Sale.weight_kgs), 0)).scalar() or 0
        gp_total = db.query(func.coalesce(func.sum(Sale.gp), 0)).scalar() or 0
        pending = db.query(func.coalesce(func.sum(Sale.total_amount), 0)).filter(Sale.payment_status == "Pending").scalar() or 0
        
        total_order_value = db.query(func.coalesce(func.sum(Order.invoice_amount), 0)).scalar() or 0
        total_order_cost = db.query(func.coalesce(func.sum(Order.value_excl_gst_freight), 0)).scalar() or 0

        lr_in_transit = db.query(func.count(Sale.id)).filter(Sale.lr_tracking_status == "In Transit").scalar() or 0
        lr_delivered = db.query(func.count(Sale.id)).filter(Sale.lr_tracking_status == "Delivered").scalar() or 0
        lr_delayed = db.query(func.count(Sale.id)).filter(Sale.lr_tracking_status == "Delayed").scalar() or 0
        lr_pending = db.query(func.count(Sale.id)).filter(Sale.lr_no.isnot(None), Sale.lr_tracking_status.is_(None)).scalar() or 0

        # Recent sales with batch-loaded customers
        recent_sales_raw = db.query(Sale).order_by(Sale.id.desc()).limit(5).all()
        customer_ids = list(set(s.customer_id for s in recent_sales_raw if s.customer_id))
        customers_map = {}
        if customer_ids:
            customers = db.query(Customer).filter(Customer.id.in_(customer_ids)).all()
            customers_map = {c.id: c for c in customers}
        
        recent_sales = []
        for s in recent_sales_raw:
            try:
                cust_name = ""
                if s.customer_id and s.customer_id in customers_map:
                    cust_name = customers_map[s.customer_id].contact_name or ""
                dt_str = ""
                if s.sale_date:
                    if hasattr(s.sale_date, 'strftime'):
                        dt_str = s.sale_date.strftime("%d %b %Y")
                    else:
                        dt_str = str(s.sale_date)[:10]
                recent_sales.append({
                    "id": s.id, "invoice": s.invoice_no or "",
                    "customer": s.party_name or cust_name or "",
                    "amount": s.total_amount or 0, "status": s.payment_status or "",
                    "date": dt_str
                })
            except Exception:
                continue

        recent_orders = []
        for o in db.query(Order).order_by(Order.id.desc()).limit(5).all():
            recent_orders.append({
                "id": o.id, "sl_no": o.sl_no or 0,
                "po_no": o.po_no or "", "customer_name": o.customer_name or "",
                "billing_site": o.billing_site or "",
                "invoice_no": o.invoice_no or "",
                "invoice_amount": o.invoice_amount or 0,
                "entry_date": o.entry_date or ""
            })

        # Monthly revenue chart using SQL
        monthly_revenue = {}
        try:
            monthly_rows = db.execute(text("""
                SELECT substr(sale_date::text, 1, 7) as month, 
                       COALESCE(sum(total_amount), 0) as total
                FROM sales 
                WHERE sale_date IS NOT NULL 
                GROUP BY month 
                ORDER BY month
            """)).fetchall()
            for row in monthly_rows:
                monthly_revenue[row[0]] = row[1]
        except Exception as e:
            logger.error("Dashboard monthly revenue query failed: %s", e)
        sorted_months = sorted(monthly_revenue.keys())[-12:]
        revenue_chart = {"labels": sorted_months, "data": [monthly_revenue[m] for m in sorted_months]}

        party_revenue = {}
        party_rows = db.query(
            Sale.party_name,
            func.coalesce(func.sum(Sale.total_amount), 0).label("total")
        ).filter(Sale.party_name.isnot(None)).group_by(Sale.party_name).order_by(func.sum(Sale.total_amount).desc()).limit(8).all()
        for row in party_rows:
            party_revenue[row.party_name] = row.total
        party_chart = {"labels": list(party_revenue.keys()), "data": list(party_revenue.values())}

        location_revenue = {}
        loc_rows = db.query(
            Sale.location,
            func.coalesce(func.sum(Sale.total_amount), 0).label("total")
        ).filter(Sale.location.isnot(None)).group_by(Sale.location).order_by(func.sum(Sale.total_amount).desc()).limit(8).all()
        for row in loc_rows:
            location_revenue[row.location] = row.total
        location_chart = {"labels": list(location_revenue.keys()), "data": list(location_revenue.values())}

        return {
            "total_products": db.query(Product).count(),
            "total_customers": db.query(Customer).count(),
            "total_orders": db.query(Order).count(),
            "total_sales": db.query(Sale).count(),
            "revenue": revenue,
            "freight": freight,
            "gp_total": gp_total,
            "pending": pending,
            "total_order_value": total_order_value,
            "total_order_cost": total_order_cost,
            "recent_sales": recent_sales,
            "recent_orders": recent_orders,
            "lr_in_transit": lr_in_transit,
            "lr_delivered": lr_delivered,
            "lr_delayed": lr_delayed,
            "lr_pending": lr_pending,
            "revenue_chart": revenue_chart,
            "party_chart": party_chart,
            "location_chart": location_chart,
        }
    except Exception as e:
        return {"error": str(e), "total_products": 0, "total_customers": 0,
                "total_orders": 0, "total_sales": 0, "revenue": 0, "freight": 0,
                "gp_total": 0, "pending": 0, "total_order_value": 0,
                "total_order_cost": 0, "recent_sales": [], "recent_orders": [],
                "lr_in_transit": 0, "lr_delivered": 0, "lr_delayed": 0, "lr_pending": 0,
                "revenue_chart": {"labels": [], "data": []},
                "party_chart": {"labels": [], "data": []},
                "location_chart": {"labels": [], "data": []}}
    finally:
        db.close()


# ---- BILLING SITES ----
@app.get("/api/billing-sites")
def list_billing_sites(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        sites = db.query(BillingSite).order_by(BillingSite.name).all()
        return [{"id": s.id, "name": s.name, "address": s.address, "phone": s.phone,
                 "email": s.email, "website": s.website, "gstin": s.gstin,
                 "state_code": s.state_code, "pan": s.pan} for s in sites]
    finally:
        db.close()


# ---- SETTINGS ----
@app.get("/api/settings")
def get_settings(user: User = Depends(get_current_user)):
    db = SessionLocal()
    try:
        rows = db.query(Settings).all()
        return {s.key: s.value for s in rows}
    finally:
        db.close()


@app.put("/api/settings")
def update_settings(body: SettingsUpdateIn, user: User = Depends(require_permission("settings", "edit"))):
    db = SessionLocal()
    try:
        for k, v in body.settings.items():
            row = db.query(Settings).filter(Settings.key == k).first()
            if row:
                row.value = str(v)
            else:
                db.add(Settings(key=k, value=str(v)))
        db.commit()
        return {"message": "Settings updated"}
    finally:
        db.close()


# ---- CSV IMPORT ----
def parse_csv_amount(val):
    if not val or str(val).strip() in ('-', '–', '', 'None', 'none'):
        return 0
    return float(str(val).replace('₹', '').replace(',', '').replace(' ', '').strip() or 0)


def parse_csv_date(val):
    if not val or val.strip() in ('-', '–', ''):
        return None
    from dateutil import parser as dateparser
    try:
        return dateparser.parse(val.strip(), dayfirst=False).strftime('%Y-%m-%d')
    except Exception:
        return val.strip() if val else None


@app.post("/api/import/orders")
async def import_orders_csv(file: UploadFile = File(...), user: User = Depends(require_permission("orders", "import"))):
    content = await file.read()
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            sl_raw = row.get('Sl No.', '').strip()
            if not sl_raw:
                skipped += 1
                continue
            try:
                sl_no = int(sl_raw)
            except ValueError:
                skipped += 1
                continue
            existing = db.query(Order).filter(Order.sl_no == sl_no).first()
            data = {
                "sl_no": sl_no,
                "po_no": "",
                "po_date": parse_csv_date(row.get('PO Date', '')),
                "customer_name": row.get('Customer Name', '').strip(),
                "billing_site": row.get('Billing Site', '').strip(),
                "shipping_site": row.get('Shipping Site', '').strip(),
                "no_of_boxes": int(parse_csv_amount(row.get('No. Of Boxes', '0'))),
                "value_excl_gst_freight": parse_csv_amount(row.get('Value (excl. GST & Freight)', '0')),
                "invoice_no": row.get('Invoice No.', '').strip().replace('-', '') if row.get('Invoice No.', '').strip() not in ('-', '–', '') else '',
                "invoice_date": parse_csv_date(row.get('Invoice Date', '')),
                "invoice_amount_excl_gst": parse_csv_amount(row.get('Invoice Amount (ex. GST)', '0')),
                "weight_kgs": parse_csv_amount(row.get('Weight (Kg)', '0')),
                "freight_rate_per_kg": parse_csv_amount(row.get('Freight (Rate / Kg)', '0')),
                "transport_charges": parse_csv_amount(row.get('Transport Charges', '0')),
                "invoice_amount": parse_csv_amount(row.get('Invoice Amount', '0')),
                "eway_bill_no": row.get('E-way Bill No', '').strip() if row.get('E-way Bill No', '').strip() not in ('-', '–', '') else '',
                "lr_no": row.get('LR Copy', '').strip() if row.get('LR Copy', '').strip() not in ('-', '–', '') else '',
                "entry_date": parse_csv_date(row.get('ERP Entry Date', '')),
                "credit_note_amount": parse_csv_amount(row.get('Credit Note Amount (If any)', '0')),
                "credit_note_no": row.get('Credit Note No.', '').strip() if row.get('Credit Note No.', '').strip() not in ('-', '–', '') else '',
                "transporter": row.get('Transporter', '').strip(),
                "transporter_no": "",
            }
            if existing:
                for k, v in data.items():
                    if k != "sl_no":
                        setattr(existing, k, v)
            else:
                o = Order(**data)
                db.add(o)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} orders, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/sales")
async def import_sales_csv(file: UploadFile = File(...), user: User = Depends(require_permission("sales", "import"))):
    content = await file.read()
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            sl_raw = row.get('Sl No.', '').strip()
            if not sl_raw:
                skipped += 1
                continue
            try:
                sl_no = int(sl_raw)
            except ValueError:
                skipped += 1
                continue
            invoice_no = row.get('Raksha Invoice NO', '').strip()
            if invoice_no in ('-', '–', ''):
                invoice_no = f"INDORE-{sl_no:04d}"
            sale_date = parse_csv_date(row.get('Date ', '') or row.get('Date', ''))
            sale_date_dt = None
            if sale_date:
                try:
                    from datetime import datetime as dt
                    sale_date_dt = dt.strptime(sale_date, '%Y-%m-%d')
                except Exception:
                    logger.warning("Invalid sale_date in CSV import: %s", sale_date)
            freight = parse_csv_amount(row.get('Freight', '0'))
            gp = parse_csv_amount(row.get('GP', '0'))
            gp_pct_raw = row.get('GP%', '0').replace('%', '').strip()
            try:
                gp_pct = float(gp_pct_raw) if gp_pct_raw and gp_pct_raw not in ('-', '–', '', 'None') else 0
            except ValueError:
                gp_pct = 0
            invoice_value = parse_csv_amount(row.get('Raksha Invoice Value', '') or row.get('Invoice Value', '') or row.get('invoice_value', '0'))
            s = Sale(
                invoice_no=invoice_no,
                sale_date=sale_date_dt,
                payment_terms=row.get('Payment Terms', '').strip(),
                party_name=row.get('Party Name ', '') or row.get('Party Name', '').strip(),
                location=row.get('Location', '').strip(),
                pincode=row.get('Pincode', '').strip(),
                state=row.get('State', '').strip(),
                transporter_name=row.get('Transporter Name', '').strip(),
                lr_no=row.get('LR No', '').strip(),
                freight_amount=freight,
                weight_kgs=parse_csv_amount(row.get('Weight', '0')),
                weight_pg_fiber=parse_csv_amount(row.get('Weight on PG Fiber Bill', '0')),
                sales_person=row.get('Sales Ex Person / Person In-Charge', '').strip(),
                pg_fiber_invoice_no=(row.get('P.G.Fiber Invoice No', '') or row.get('P.G.Fiber Invoice No ', '') or '').strip(),
                pg_fiber_invoice_value=parse_csv_amount(row.get('P.G.Fiber Invoice Value', '') or row.get('P.G.Fiber Invoice Value ', '0')),
                gp=gp,
                gp_percent=gp_pct,
                invoice_value=invoice_value,
                total_amount=invoice_value if invoice_value else 0,
                source_csv="From Indore",
            )
            db.add(s)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} sales, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


def map_csv_col(row, keys, default=""):
    for k in keys:
        v = row.get(k, "")
        if v and v.strip() and v.strip() not in ('-', '–'):
            return v.strip()
    return default


@app.post("/api/import/products")
async def import_products_csv(file: UploadFile = File(...), user: User = Depends(require_permission("products", "import"))):
    content = await file.read()
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            name = map_csv_col(row, ['Name', 'Description', 'Product Name', 'name', 'description'])
            if not name:
                skipped += 1
                continue
            part_no = map_csv_col(row, ['Part No', 'Part Number', 'SKU', 'part_no'])
            category = map_csv_col(row, ['Category', 'category'], 'Manhole Cover')
            size = map_csv_col(row, ['Size', 'size'])
            load_rating = map_csv_col(row, ['Load Rating', 'Load', 'load_rating'], '5 Ton')
            material = map_csv_col(row, ['Material', 'material'], 'FRP')
            color = map_csv_col(row, ['Color', 'color'], 'Grey')
            unit = map_csv_col(row, ['Unit', 'unit'], 'Nos')
            hsn_code = map_csv_col(row, ['HSN Code', 'HSN', 'hsn_code'])
            mrp = parse_csv_amount(map_csv_col(row, ['MRP', 'Price', 'mrp'], '0'))

            if part_no:
                existing = db.query(Product).filter(Product.part_no == part_no).first()
            else:
                existing = db.query(Product).filter(Product.name == name).first()
            if existing:
                existing.name = name
                existing.category = category
                existing.size = size
                existing.load_rating = load_rating
                existing.material = material
                existing.color = color
                existing.unit = unit
                existing.hsn_code = hsn_code
                if part_no:
                    existing.part_no = part_no
                if mrp and existing.pricing:
                    existing.pricing.mrp = mrp
                    existing.pricing.raw_material_cost = mrp
                    existing.pricing.total_cost = mrp
                imported += 1
                continue

            p = Product(part_no=part_no, name=name, category=category,
                        size=size, load_rating=load_rating, material=material,
                        color=color, unit=unit, hsn_code=hsn_code)
            db.add(p)
            db.flush()
            db.add(Pricing(product_id=p.id, raw_material_cost=mrp, total_cost=mrp, mrp=mrp))
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} products, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/products/dedup")
def dedup_products(user: User = Depends(require_permission("products", "edit"))):
    db = SessionLocal()
    try:
        products = db.query(Product).order_by(Product.id).all()
        seen = {}
        removed = 0
        for p in products:
            key = (p.part_no or '', p.name or '')
            if key in seen:
                # Merge pricing if needed
                if p.pricing and not seen[key].pricing:
                    seen[key].pricing = p.pricing
                    p.pricing = None
                elif p.pricing and seen[key].pricing:
                    db.delete(p.pricing)
                # Re-point foreign keys referencing this product
                for sale in db.query(Sale).filter(Sale.product_id == p.id).all():
                    sale.product_id = seen[key].id
                db.delete(p)
                removed += 1
            else:
                seen[key] = p
        db.commit()
        return {"removed": removed, "remaining": len(seen), "message": f"Removed {removed} duplicate products"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Dedup failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/customers")
async def import_customers_csv(file: UploadFile = File(...), user: User = Depends(require_permission("customers", "import"))):
    content = await file.read()
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            customer_id = map_csv_col(row, ['Customer ID', 'ID', 'Cust ID', 'customer_id'])
            if not customer_id:
                skipped += 1
                continue
            existing = db.query(Customer).filter(Customer.customer_id == customer_id).first()
            data = {
                "customer_id": customer_id,
                "name": map_csv_col(row, ['Contact Name', 'Name', 'Customer Name', 'contact_name']),
                "gstin": map_csv_col(row, ['GSTIN', 'GST Number', 'GST', 'gstin']),
                "billing_address": map_csv_col(row, ['Billing Address', 'Address', 'billing_address']),
                "shipping_address": map_csv_col(row, ['Shipping Address', 'shipping_address']),
                "state": map_csv_col(row, ['State', 'state']),
                "district": map_csv_col(row, ['District', 'district']),
                "city": map_csv_col(row, ['City', 'city']),
                "pincode": map_csv_col(row, ['Pincode', 'Pin Code', 'ZIP', 'pincode']),
                "contact_name": map_csv_col(row, ['Contact Name', 'Name', 'contact_name']),
                "contact_number": map_csv_col(row, ['Contact Number', 'Phone', 'Mobile', 'contact_number']),
                "contact_email": map_csv_col(row, ['Contact Email', 'Email', 'contact_email']),
                "exec_code": map_csv_col(row, ['Executive Code', 'Exec Code', 'exec_code']),
                "exec_name": map_csv_col(row, ['Executive Name', 'Exec Name', 'exec_name']),
                "exec_number": map_csv_col(row, ['Executive Number', 'Exec Number', 'exec_number']),
                "exec_email": map_csv_col(row, ['Executive Email', 'Exec Email', 'exec_email']),
            }
            if existing:
                for k, v in data.items():
                    if k != "customer_id":
                        setattr(existing, k, v)
            else:
                c = Customer(**data)
                db.add(c)
                try:
                    db.flush()
                except Exception:
                    db.rollback()
                    ex = db.query(Customer).filter(Customer.customer_id == customer_id).first()
                    if ex:
                        for k, v in data.items():
                            if k != "customer_id":
                                setattr(ex, k, v)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} customers, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/transporters")
async def import_transporters_csv(file: UploadFile = File(...), user: User = Depends(require_permission("transporters", "import"))):
    content = await file.read()
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            transporter_id = map_csv_col(row, ['Transporter ID', 'ID', 'transporter_id'])
            name = map_csv_col(row, ['Name', 'Transporter Name', 'name'])
            if not transporter_id or not name:
                skipped += 1
                continue
            existing = db.query(Transporter).filter(Transporter.transporter_id == transporter_id).first()
            data = {
                "transporter_id": transporter_id,
                "name": name,
                "phone": map_csv_col(row, ['Phone', 'Mobile', 'phone']),
                "email": map_csv_col(row, ['Email', 'email']),
                "address": map_csv_col(row, ['Address', 'address']),
                "state": map_csv_col(row, ['State', 'state']),
                "district": map_csv_col(row, ['District', 'district']),
                "city": map_csv_col(row, ['City', 'city']),
                "pincode": map_csv_col(row, ['Pincode', 'Pin Code', 'pincode']),
                "gst_number": map_csv_col(row, ['GST Number', 'GSTIN', 'GST', 'gst_number']),
                "pan_number": map_csv_col(row, ['PAN Number', 'PAN', 'pan_number']),
                "contact_person": map_csv_col(row, ['Contact Person', 'contact_person']),
                "contact_number": map_csv_col(row, ['Contact Number', 'contact_number']),
            }
            if existing:
                for k, v in data.items():
                    if k != "transporter_id":
                        setattr(existing, k, v)
            else:
                t = Transporter(**data)
                db.add(t)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} transporters, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/import/expenses")
async def import_expenses_csv(file: UploadFile = File(...), user: User = Depends(require_permission("expenses", "import"))):
    content = await file.read()
    text = content.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    db = SessionLocal()
    imported = 0
    skipped = 0
    try:
        for row in reader:
            category = map_csv_col(row, ['Category', 'category'])
            amount_raw = map_csv_col(row, ['Amount', 'amount'], '0')
            amount = parse_csv_amount(amount_raw)
            if not category or amount == 0:
                skipped += 1
                continue
            date_str = map_csv_col(row, ['Date', 'Expense Date', 'expense_date'])
            dt = None
            if date_str:
                try:
                    from dateutil import parser as dateparser
                    dt = dateparser.parse(date_str, dayfirst=False)
                except Exception:
                    logger.warning("Invalid expense date in CSV import: %s", date_str)
            if not dt:
                dt = datetime.now(timezone.utc)
            e = Expense(
                category=category,
                description=map_csv_col(row, ['Description', 'description']),
                amount=amount,
                vendor=map_csv_col(row, ['Vendor', 'vendor']),
                expense_date=dt
            )
            db.add(e)
            imported += 1
        db.commit()
        return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} expenses, skipped {skipped}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


# ---- FILE UPLOAD (Cloudinary) ----
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".pdf"}

@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...), user: User = Depends(require_permission("products", "create"))):
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, "Only .jpg, .png, .pdf files allowed")
    try:
        content = await file.read()
        rtype = "raw" if ext == ".pdf" else "image"
        result = cloudinary.uploader.upload(
            content,
            folder="raksha_erp",
            resource_type=rtype
        )
        url = result["secure_url"]
        if rtype == "raw" and "/image/" in url:
            url = url.replace("/image/upload/", "/raw/upload/")
        return {"filename": result["public_id"], "url": url, "original": file.filename}
    except Exception as e:
        raise HTTPException(500, f"Upload failed: {str(e)}")


@app.post("/api/import-standard-packaging")
async def import_standard_packaging(file: UploadFile = File(...), user: User = Depends(require_permission("products", "import"))):
    db = SessionLocal()
    try:
        content = await file.read()
        filename = (file.filename or "").lower()

        if filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        else:
            text_content = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text_content))
            rows = list(reader)

        if not rows:
            raise HTTPException(400, "File is empty")

        header_row_idx = 0
        for i, row in enumerate(rows):
            row_lower = [str(c).strip().lower() for c in row]
            if any("part no" in c for c in row_lower):
                header_row_idx = i
                break

        header = [h.strip().lower() for h in rows[header_row_idx]]
        part_no_idx = None
        box_idx = None
        mrp_idx = None
        desc_idx = None
        size_idx = None
        grp_idx = None
        for i, h in enumerate(header):
            if "part no" in h:
                part_no_idx = i
            if h in ("box", "boxes") or "std pack" in h or "std pkg" in h:
                box_idx = i
            if h in ("gen", "mrp") or h == "gen":
                mrp_idx = i
            if "product spec" in h or "description" in h:
                desc_idx = i
            if "size" in h:
                size_idx = i
            if "item grp" in h or "item group" in h or "grp" in h:
                grp_idx = i
        if part_no_idx is None or box_idx is None:
            raise HTTPException(400, f"File must have 'Part No' and 'Box'/'Std Packing' columns. Found: {header}")

        desc_idx = desc_idx or 1
        size_idx = size_idx or 2
        grp_idx = grp_idx or 4

        created = 0
        updated = 0
        not_found = []
        for row in rows[header_row_idx+1:]:
            if len(row) <= max(part_no_idx, box_idx):
                continue
            part_no = row[part_no_idx].strip().replace(" ", "")
            box_val = row[box_idx].strip()
            if not part_no or not box_val:
                continue
            try:
                pieces = int(float(box_val))
            except ValueError:
                continue

            desc = str(row[desc_idx] or "").strip() if len(row) > desc_idx else ""
            size_raw = str(row[size_idx] or "").strip() if len(row) > size_idx else ""
            item_grp = str(row[grp_idx] or "").strip().upper() if len(row) > grp_idx else "FRP"

            size_mm = size_raw.lower().replace(" ", "").replace("x", "x")
            tonnage = "10 Ton" if "10 ton" in desc.lower() else "5 Ton"
            color = "White" if "white" in desc.lower() or part_no.upper().endswith("-WH") else "Grey"
            has_lock = "lock" in desc.lower() or part_no.upper().endswith("L") and "GRY" not in part_no.upper()
            has_hinges = "hinge" in desc.lower()
            category = "Gully Cover" if "gully" in desc.lower() or item_grp == "GULLY" else "Manhole Cover"

            product = db.query(Product).filter(Product.part_no == part_no).first()
            if product:
                product.pieces_per_box = pieces
                product.std_packaging = pieces
                if mrp_idx is not None and len(row) > mrp_idx:
                    mrp_val = row[mrp_idx].strip().replace(",", "").replace("₹", "").replace("?", "")
                    try:
                        mrp = float(mrp_val)
                        if mrp > 0:
                            pr = db.query(Pricing).filter(Pricing.product_id == product.id).first()
                            if pr:
                                pr.mrp = mrp
                            else:
                                db.add(Pricing(product_id=product.id, mrp=mrp, gst_rate=18))
                    except ValueError:
                        logger.warning("Invalid MRP value '%s' for product %s", mrp_val, part_no)
                updated += 1
            else:
                if not desc:
                    continue
                new_p = Product(
                    part_no=part_no, name=desc, category=category,
                    size=size_mm, load_rating=tonnage, material="FRP",
                    color=color, hsn_code="39259090",
                    pieces_per_box=pieces, std_packaging=pieces
                )
                db.add(new_p)
                db.flush()
                mrp_val = 0
                if mrp_idx is not None and len(row) > mrp_idx:
                    try:
                        mrp_val = float(row[mrp_idx].strip().replace(",", "").replace("₹", "").replace("?", ""))
                    except ValueError:
                        logger.warning("Invalid MRP value for new product: %s", row[mrp_idx])
                db.add(Pricing(product_id=new_p.id, mrp=mrp_val, gst_rate=18))
                created += 1

        db.commit()
        return {"updated": updated, "created": created, "not_found": not_found, "total_rows": len(rows) - header_row_idx - 1}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Import failed: {str(e)}")
    finally:
        db.close()


# ---- XLSX IMPORT (Orders & Sales) ----
def read_xlsx_sheet(file_content, sheet_name=None):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    ws = None
    if sheet_name:
        exact = [s for s in wb.sheetnames if s.strip().lower() == sheet_name.strip().lower()]
        if exact:
            ws = wb[exact[0]]
        else:
            partial = [s for s in wb.sheetnames if sheet_name.strip().lower() in s.strip().lower()]
            if partial:
                ws = wb[partial[0]]
    if ws is None:
        ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append([str(c) if c is not None else "" for c in row])
    return rows


def rows_to_csv_string(rows):
    output = io.StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    return output.getvalue()


@app.post("/api/import/orders-xlsx")
async def import_orders_xlsx(file: UploadFile = File(...), user: User = Depends(require_permission("orders", "import"))):
    content = await file.read()
    try:
        rows = read_xlsx_sheet(content, sheet_name='Orders')
        if not rows:
            raise HTTPException(400, "File is empty")
        csv_text = rows_to_csv_string(rows)
        text_content = csv_text
        reader = csv.DictReader(io.StringIO(text_content))
        db = SessionLocal()
        imported = 0
        skipped = 0
        try:
            for row in reader:
                sl_raw = row.get('Sl No.', '').strip()
                if not sl_raw:
                    skipped += 1
                    continue
                try:
                    sl_no = int(float(sl_raw))
                except ValueError:
                    skipped += 1
                    continue
                existing = db.query(Order).filter(Order.sl_no == sl_no).first()
                data = {
                    "sl_no": sl_no,
                    "po_no": "",
                    "po_date": parse_csv_date(row.get('PO Date', '')),
                    "customer_name": (row.get('Customer Name', '') or row.get('Billing Site', '') or '').strip(),
                    "billing_site": row.get('Billing Site', '').strip(),
                    "shipping_site": row.get('Shipping Site', '').strip(),
                    "no_of_boxes": int(parse_csv_amount(row.get('No. Of Boxes', '0'))),
                    "value_excl_gst_freight": parse_csv_amount(row.get('Value (excl. GST & Freight)', '0')),
                    "invoice_no": row.get('Invoice No.', '').strip() if row.get('Invoice No.', '').strip() not in ('-', '–', '') else '',
                    "invoice_date": parse_csv_date(row.get('Invoice Date', '')),
                    "invoice_amount_excl_gst": parse_csv_amount(row.get('Invoice Amount (ex. GST)', '0')),
                    "weight_kgs": parse_csv_amount(row.get('Weight (Kg)', '0')),
                    "freight_rate_per_kg": parse_csv_amount(row.get('Freight (Rate / Kg)', '0')),
                    "transport_charges": parse_csv_amount(row.get('Transport Charges', '0')),
                    "invoice_amount": parse_csv_amount(row.get('Invoice Amount', '0')),
                    "eway_bill_no": row.get('E-way Bill No', '').strip() if row.get('E-way Bill No', '').strip() not in ('-', '–', '') else '',
                    "lr_no": row.get('LR Copy', '').strip() if row.get('LR Copy', '').strip() not in ('-', '–', '') else '',
                    "entry_date": parse_csv_date(row.get('ERP Entry Date', '')),
                    "credit_note_amount": parse_csv_amount(row.get('Credit Note Amount (If any)', '0')),
                    "credit_note_no": row.get('Credit Note No.', '').strip() if row.get('Credit Note No.', '').strip() not in ('-', '–', '') else '',
                    "transporter": row.get('Transporter', '').strip(),
                    "transporter_no": "",
                }
                if existing:
                    for k, v in data.items():
                        if k != "sl_no":
                            setattr(existing, k, v)
                else:
                    o = Order(**data)
                    db.add(o)
                imported += 1
            db.commit()
            return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} orders from XLSX"}
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to read XLSX: {str(e)}")


@app.post("/api/import/sales-xlsx")
async def import_sales_xlsx(file: UploadFile = File(...), user: User = Depends(require_permission("sales", "import"))):
    content = await file.read()
    try:
        rows = read_xlsx_sheet(content, sheet_name='Sales')
        if not rows:
            raise HTTPException(400, "File is empty")
        csv_text = rows_to_csv_string(rows)
        reader = csv.DictReader(io.StringIO(csv_text))
        db = SessionLocal()
        imported = 0
        skipped = 0
        try:
            for row in reader:
                sl_raw = row.get('Sl No.', '').strip()
                if not sl_raw:
                    skipped += 1
                    continue
                try:
                    int(float(sl_raw))
                except ValueError:
                    skipped += 1
                    continue
                invoice_no = row.get('Raksha Invoice NO', '').strip()
                if invoice_no in ('-', '–', ''):
                    invoice_no = None
                sale_date = parse_csv_date(row.get('Date ', '') or row.get('Date', ''))
                sale_date_dt = None
                if sale_date:
                    try:
                        sale_date_dt = datetime.strptime(sale_date, '%Y-%m-%d')
                    except Exception:
                        logger.warning("Invalid sale_date in CSV update: %s", sale_date)
                freight = parse_csv_amount(row.get('Freight', '0'))
                gp = parse_csv_amount(row.get('GP', '0'))
                gp_pct_raw = row.get('GP%', '0').replace('%', '').strip()
                try:
                    gp_pct = float(gp_pct_raw) if gp_pct_raw and gp_pct_raw not in ('-', '–', '', 'None') else 0
                except ValueError:
                    gp_pct = 0
                invoice_value = parse_csv_amount(row.get('Raksha Invoice Value', '') or row.get('Invoice Value', '') or row.get('invoice_value', '0'))

                existing = None
                if invoice_no:
                    existing = db.query(Sale).filter(Sale.invoice_no == invoice_no).first()
                if not existing:
                    existing = db.query(Sale).filter(
                        Sale.party_name == (row.get('Party Name ', '') or row.get('Party Name', '') or '').strip(),
                        Sale.sale_date == sale_date_dt
                    ).first() if sale_date_dt else None

                data = dict(
                    invoice_no=invoice_no,
                    sale_date=sale_date_dt,
                    payment_terms=row.get('Payment Terms', '').strip(),
                    party_name=(row.get('Party Name ', '') or row.get('Party Name', '') or '').strip(),
                    location=row.get('Location', '').strip(),
                    pincode=row.get('Pincode', '').strip(),
                    state=row.get('State', '').strip(),
                    transporter_name=row.get('Transporter Name', '').strip(),
                    lr_no=row.get('LR No', '').strip(),
                    freight_amount=freight,
                    weight_kgs=parse_csv_amount(row.get('Weight', '0')),
                    weight_pg_fiber=parse_csv_amount(row.get('Weight on PG Fiber Bill', '0')),
                    sales_person=row.get('Sales Ex Person / Person In-Charge', '').strip(),
                    pg_fiber_invoice_no=(row.get('P.G.Fiber Invoice No', '') or row.get('P.G.Fiber Invoice No ', '') or '').strip(),
                    pg_fiber_invoice_value=parse_csv_amount(row.get('P.G.Fiber Invoice Value', '') or row.get('P.G.Fiber Invoice Value ', '0')),
                    gp=gp,
                    gp_percent=gp_pct,
                    invoice_value=invoice_value,
                    total_amount=invoice_value if invoice_value else 0,
                    source_csv="From Indore",
                )

                if existing:
                    for k, v in data.items():
                        setattr(existing, k, v)
                else:
                    s = Sale(**data)
                    db.add(s)
                imported += 1
            db.commit()
            return {"imported": imported, "skipped": skipped, "message": f"Imported {imported} sales from XLSX"}
        except Exception as e:
            db.rollback()
            raise HTTPException(500, f"Import failed: {str(e)}")
        finally:
            db.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to read XLSX: {str(e)}")


# ---- EXPORT (CSV, XLSX, PDF) ----
@app.get("/api/export/orders")
def export_orders(format: str = "csv", user: User = Depends(require_permission("orders", "export"))):
    db = SessionLocal()
    try:
        rows = db.query(Order).order_by(Order.sl_no).all()
        headers = ["Sl No.", "PO No.", "PO Date", "Customer Name", "Billing Site", "Shipping Site", "No. Of Boxes",
                    "Value (excl. GST & Freight)", "Invoice No.", "Invoice Date",
                    "Invoice Amount (ex. GST)", "Weight (Kg)", "Freight (Rate / Kg)",
                    "Transport Charges", "Invoice Amount", "E-way Bill No", "LR No.",
                    "Entry Date", "Credit Note Amount", "Credit Note No.", "Transporter"]
        data = []
        for o in rows:
            data.append([o.sl_no, o.po_no or "", o.po_date or "", o.customer_name or "", o.billing_site or "", o.shipping_site or "",
                         o.no_of_boxes or 0, o.value_excl_gst_freight or 0, o.invoice_no or "",
                         o.invoice_date or "", o.invoice_amount_excl_gst or 0, o.weight_kgs or 0,
                         o.freight_rate_per_kg or 0, o.transport_charges or 0, o.invoice_amount or 0,
                         o.eway_bill_no or "", o.lr_no or "", o.entry_date or "",
                         o.credit_note_amount or 0, o.credit_note_no or "", o.transporter or ""])

        if format == "xlsx":
            return export_xlsx("Orders", headers, data)
        elif format == "pdf":
            return export_pdf("Orders", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


@app.get("/api/export/sales")
def export_sales(format: str = "csv", user: User = Depends(require_permission("sales", "export"))):
    db = SessionLocal()
    try:
        rows = db.query(Sale).order_by(Sale.id.desc()).all()
        headers = ["Invoice No.", "Date", "Party Name", "Location", "State", "Transporter",
                    "Freight", "Weight", "Weight PG Fiber", "Invoice Value", "GP", "GP%",
                    "Payment Terms", "Sales Person", "PG Fiber Invoice No", "PG Fiber Invoice Value"]
        data = []
        for s in rows:
            dt = ""
            if s.sale_date:
                try:
                    dt = s.sale_date.strftime("%Y-%m-%d")
                except Exception:
                    dt = str(s.sale_date)[:10]
            data.append([s.invoice_no or "", dt, s.party_name or "", s.location or "",
                         s.state or "", s.transporter_name or "", s.freight_amount or 0,
                         s.weight_kgs or 0, s.weight_pg_fiber or 0, s.invoice_value or 0,
                         s.gp or 0, s.gp_percent or 0, s.payment_terms or "",
                         s.sales_person or "", s.pg_fiber_invoice_no or "",
                         s.pg_fiber_invoice_value or 0])

        if format == "xlsx":
            return export_xlsx("Sales", headers, data)
        elif format == "pdf":
            return export_pdf("Sales", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


@app.get("/api/export/proforma-orders")
def export_proforma_orders(format: str = "csv", order_type: str = None, user: User = Depends(require_permission("proforma_orders", "export"))):
    db = SessionLocal()
    try:
        query = db.query(ProformaOrder)
        if order_type:
            query = query.filter(ProformaOrder.order_type == order_type)
        rows = query.order_by(ProformaOrder.created_at.desc()).all()
        headers = ["PI No", "Date", "Customer", "Type", "Billing Site", "Shipping Site",
                    "Boxes", "Total Qty", "Value (excl GST)", "GST", "Freight",
                    "Total Amount", "Payment Status", "Delivery Days"]
        data = []
        for o in rows:
            cust = db.query(Customer).filter(Customer.id == o.customer_id).first()
            data.append([o.pi_no or "", o.pi_date.strftime("%Y-%m-%d") if o.pi_date else "",
                         cust.contact_name if cust else "", o.order_type or "",
                         o.billing_site or "", o.shipping_site or "",
                         o.no_of_boxes or 0, o.total_qty or 0, o.value_excl_gst or 0,
                         o.gst_amount or 0, o.freight_amount or 0, o.total_amount or 0,
                         o.payment_status or "", o.delivery_days or 0])

        if format == "xlsx":
            return export_xlsx("PI-PO Orders", headers, data)
        elif format == "pdf":
            return export_pdf("PI-PO Orders", headers, data)
        else:
            return export_csv(headers, data)
    finally:
        db.close()


def export_csv(headers, data):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in data:
        writer.writerow(row)
    csv_bytes = output.getvalue().encode('utf-8-sig')
    return Response(content=csv_bytes, media_type="text/csv",
                    headers={"Content-Disposition": "attachment; filename=export.csv"})


def export_xlsx(sheet_name, headers, data):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    for row in data:
        ws.append(row)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                logger.debug("Column width calc error for cell in %s", col_letter)
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={sheet_name}.xlsx"})


def export_pdf(title, headers, data):
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, f"Raksha ERP - {title}", ln=True, align="C")
    pdf.set_font("Helvetica", "", 8)
    pdf.cell(0, 6, f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}", ln=True, align="C")
    pdf.ln(4)
    num_cols = len(headers)
    col_width = max(277 / num_cols, 20)
    pdf.set_font("Helvetica", "B", 7)
    for h in headers:
        short_h = str(h)[:20]
        pdf.cell(col_width, 7, short_h, border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "", 6)
    for row in data:
        for val in row:
            s = str(val)[:22]
            pdf.cell(col_width, 5, s, border=1)
        pdf.ln()
    pdf_bytes = pdf.output()
    return Response(content=bytes(pdf_bytes), media_type="application/pdf",
                    headers={"Content-Disposition": f"attachment; filename={title}.pdf"})


# ---- AUTH ----
@app.post("/api/auth/login")
@limiter.limit("5/minute")
def login(request: Request, body: LoginIn):
    username = body.username
    password = body.password
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.username == username).first()
        if not user or not bcrypt.checkpw(password.encode(), user.password_hash.encode()):
            raise HTTPException(status_code=401, detail="Invalid credentials")
        if not user.is_active:
            raise HTTPException(status_code=403, detail="Account disabled")
        user.last_login = datetime.now(timezone.utc)
        db.commit()
        access_token = jwt.encode(
            {"user_id": user.id, "role": user.role, "type": "access",
             "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)},
            JWT_SECRET, algorithm=JWT_ALGORITHM
        )
        refresh_token = jwt.encode(
            {"user_id": user.id, "type": "refresh",
             "exp": datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)},
            JWT_SECRET, algorithm=JWT_ALGORITHM
        )
        audit_log(user, "login", details=f"User {username} logged in", request=request)
        return {
            "access_token": access_token,
            "refresh_token": refresh_token,
            "token_type": "bearer",
            "user": {"id": user.id, "username": user.username, "full_name": user.full_name, "role": user.role}
        }
    finally:
        db.close()


@app.post("/api/auth/refresh")
@limiter.limit("10/minute")
def refresh_token(request: Request, body: RefreshIn):
    token = body.refresh_token
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "refresh":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload.get("user_id")
        jti = payload.get("jti")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Refresh token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    db = SessionLocal()
    try:
        # Check if token is blacklisted (fix #21 - token revocation)
        if jti:
            bl = db.query(TokenBlacklist).filter(TokenBlacklist.token == token).first()
            if bl:
                raise HTTPException(status_code=401, detail="Token has been revoked")
        user = db.query(User).filter(User.id == user_id, User.is_active == 1).first()
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        # Blacklist old refresh token (rotation)
        if jti:
            try:
                old_exp = datetime.fromtimestamp(payload.get("exp", 0), tz=timezone.utc)
                db.add(TokenBlacklist(token=token, user_id=user_id, expires_at=old_exp))
                db.commit()
            except Exception as e:
                logger.warning(f"Failed to blacklist old refresh token: {e}")
        new_jti = str(uuid.uuid4())
        access_token = jwt.encode(
            {"user_id": user.id, "role": user.role, "type": "access",
             "exp": datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)},
            JWT_SECRET, algorithm=JWT_ALGORITHM
        )
        new_refresh_token = jwt.encode(
            {"user_id": user.id, "type": "refresh", "jti": new_jti,
             "exp": datetime.now(timezone.utc) + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)},
            JWT_SECRET, algorithm=JWT_ALGORITHM
        )
        return {"access_token": access_token, "refresh_token": new_refresh_token, "token_type": "bearer"}
    finally:
        db.close()


@app.get("/api/auth/me")
def get_me(user: User = Depends(get_current_user)):
    return {
        "id": user.id, "username": user.username, "full_name": user.full_name,
        "email": user.email, "role": user.role, "is_active": user.is_active,
        "last_login": str(user.last_login) if user.last_login else None,
        "permissions": ROLE_PERMISSIONS.get(user.role, {}),
    }


@app.get("/api/users")
def list_users(user: User = Depends(require_permission("users", "view"))):
    db = SessionLocal()
    try:
        users = db.query(User).all()
        return [{"id": u.id, "username": u.username, "full_name": u.full_name,
                 "email": u.email, "role": u.role, "is_active": u.is_active,
                 "last_login": str(u.last_login) if u.last_login else None,
                 "created_at": str(u.created_at) if u.created_at else None} for u in users]
    finally:
        db.close()


@app.get("/api/users/{uid}")
def get_user(uid: int, user: User = Depends(require_permission("users", "view"))):
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == uid).first()
        if not u:
            raise HTTPException(404, "User not found")
        return {"id": u.id, "username": u.username, "full_name": u.full_name,
                "email": u.email, "role": u.role, "is_active": u.is_active,
                "last_login": str(u.last_login) if u.last_login else None,
                "created_at": str(u.created_at) if u.created_at else None}
    finally:
        db.close()


@app.post("/api/users")
def create_user(body: UserCreateIn, user: User = Depends(require_permission("users", "create")), request: Request = None):
    db = SessionLocal()
    try:
        if db.query(User).filter(User.username == body.username).first():
            raise HTTPException(status_code=400, detail="Username already exists")
        pw = body.password
        pw_hash = bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()
        new_user = User(
            username=body.username, password_hash=pw_hash,
            full_name=body.full_name, email=body.email,
            role=body.role, is_active=1,
        )
        db.add(new_user)
        db.commit()
        audit_log(user, "create_user", resource="users", resource_id=new_user.id,
                  details=f"Created user {body.username} with role {body.role}", request=request)
        return {"message": "User created", "id": new_user.id}
    finally:
        db.close()


@app.put("/api/users/{uid}")
def update_user(uid: int, body: UserUpdateIn, user: User = Depends(require_permission("users", "edit")), request: Request = None):
    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == uid).first()
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        if target.username == "admin" and body.role and body.role != "admin":
            raise HTTPException(status_code=400, detail="Cannot change admin role")
        if body.full_name is not None:
            target.full_name = body.full_name
        if body.email is not None:
            target.email = body.email
        if body.role is not None:
            target.role = body.role
        if body.is_active is not None:
            target.is_active = body.is_active
        if body.password:
            target.password_hash = bcrypt.hashpw(body.password.encode(), bcrypt.gensalt()).decode()
        target.updated_at = datetime.now(timezone.utc)
        db.commit()
        audit_log(user, "update_user", resource="users", resource_id=uid,
                  details=f"Updated user {target.username}", request=request)
        return {"message": "User updated"}
    finally:
        db.close()


@app.delete("/api/users/{uid}")
def delete_user(uid: int, user: User = Depends(require_permission("users", "delete")), request: Request = None):
    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == uid).first()
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        # Prevent deleting the last admin (role-based, not username-based)
        if target.role == "admin":
            admin_count = db.query(User).filter(User.role == "admin", User.is_active == 1).count()
            if admin_count <= 1:
                raise HTTPException(status_code=400, detail="Cannot delete the last active admin user")
        deleted_username = target.username
        db.delete(target)
        db.commit()
        audit_log(user, "delete_user", resource="users", resource_id=uid,
                  details=f"Deleted user {deleted_username}", request=request)
        return {"message": "User deleted"}
    finally:
        db.close()


@app.put("/api/users/{uid}/password")
def change_password(uid: int, body: ChangePasswordIn, user: User = Depends(get_current_user)):
    if user.id != uid and user.role != "admin":
        raise HTTPException(status_code=403, detail="Can only change own password")
    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == uid).first()
        if not target:
            raise HTTPException(status_code=404, detail="User not found")
        if user.role != "admin":
            if not bcrypt.checkpw(body.current_password.encode(), target.password_hash.encode()):
                raise HTTPException(status_code=400, detail="Current password incorrect")
        elif user.id == uid:
            if not body.current_password:
                raise HTTPException(status_code=400, detail="Current password required")
            if not bcrypt.checkpw(body.current_password.encode(), target.password_hash.encode()):
                raise HTTPException(status_code=400, detail="Current password incorrect")
        pw = body.new_password
        target.password_hash = bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()
        db.commit()
        return {"message": "Password changed"}
    finally:
        db.close()


# ---- HEALTH CHECK ----
@app.get("/health")
def health_check():
    try:
        db = SessionLocal()
        db.execute(text("SELECT 1"))
        db.close()
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        return JSONResponse(status_code=503, content={"status": "unhealthy", "database": "disconnected"})


# ---- REQUEST SIZE LIMITS ----
MAX_UPLOAD_SIZE = int(os.environ.get("MAX_UPLOAD_SIZE", 10 * 1024 * 1024))  # 10MB default


@app.post("/api/products/import")
@app.post("/api/customers/import")
@app.post("/api/transporters/import")
@app.post("/api/purchase-rates/import")
@app.post("/api/sales/import")
@app.post("/api/expenses/import")
async def check_upload_size(request: Request):
    content_length = request.headers.get("content-length")
    if content_length and int(content_length) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_UPLOAD_SIZE // (1024*1024)}MB")
    return await request.body()


# ---- FRONTEND ----
frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.isdir(frontend_path):
    app.mount("/static", StaticFiles(directory=frontend_path), name="static")


@app.get("/")
def index():
    index_path = os.path.join(frontend_path, "index.html")
    if os.path.isfile(index_path):
        return FileResponse(index_path)
    return {"message": "Raksha ERP API is running. Frontend not found."}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
